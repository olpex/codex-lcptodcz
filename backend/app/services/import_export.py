import csv
import os
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

from docx import Document as DocxDocument
from fpdf import FPDF
from fpdf.fonts import FontFace
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from sqlalchemy import or_
from sqlalchemy.orm import Session
import xlrd
from xlrd.xldate import xldate_as_datetime

from app.core.crypto import cipher
from app.models import (
    Document,
    DocumentType,
    ExportJob,
    Group,
    GroupStatus,
    GroupMembership,
    ImportJob,
    JobStatus,
    MembershipStatus,
    Performance,
    ScheduleSlot,
    Teacher,
    Trainee,
)
from app.services.storage import storage_path

PREFERRED_TRAINEE_SHEET_NAMES = ("Додаток", "додаток")
HEADER_SCAN_LIMIT = 30
ROW_SAMPLE_LIMIT = 20
IMPORT_UPDATE_MODES = {"skip_existing", "missing_only", "overwrite"}
PDF_REPORT_ROW_LIMIT = 500
PDF_FONT_CANDIDATES = (
    "PDF_FONT_PATH",
    "backend/assets/fonts/DejaVuSans.ttf",
    "assets/fonts/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
    "C:/Windows/Fonts/arial.ttf",
    "C:/Windows/Fonts/segoeui.ttf",
)
REPORT_TYPE_LABELS = {
    "trainees": "Слухачі",
    "teacher_workload": "Навантаження викладачів",
    "kpi": "Показники",
    "form_1pa": "Форма 1-ПА",
    "employment": "Працевлаштування",
    "financial": "Фінансовий звіт",
}

FIRST_NAME_ALIASES = {
    "first_name",
    "first name",
    "firstname",
    "ім'я",
    "iм'я",
    "імя",
    "имя",
    "імя слухача",
}
LAST_NAME_ALIASES = {
    "last_name",
    "last name",
    "lastname",
    "surname",
    "прізвище",
    "прiзвище",
    "фамилия",
}
MIDDLE_NAME_ALIASES = {
    "middle_name",
    "middle name",
    "patronymic",
    "по батькові",
    "по батькови",
    "по-батькові",
    "по батьковi",
}
FULL_NAME_ALIASES = {
    "піб",
    "пiб",
    "п.i.б",
    "п.і.б",
    "пiб слухача",
    "піб безробітного",
    "пiб безробітного",
    "прізвище ім'я по батькові",
    "прізвище, ім'я, по батькові",
    "фио",
    "full_name",
    "full name",
}
BIRTH_DATE_ALIASES = {
    "birth_date",
    "birth date",
    "дата народження",
    "дата народж.",
    "дн",
}
STATUS_ALIASES = {"status", "статус"}
GROUP_CODE_ALIASES = {"group_code", "код групи", "номер групи", "group", "група"}
GROUP_NAME_ALIASES = {"group_name", "назва групи", "найменування групи"}
ROW_NUMBER_ALIASES = {"№", "no", "номер", "№ з/п", "п/п", "n"}
EMPLOYMENT_CENTER_ALIASES = {
    "центр зайнятості, який направив безробітного на професійне навчання",
    "центр зайнятості",
    "цз",
}
CONTRACT_NUMBER_ALIASES = {"№ договору", "номер договору", "договору", "contract_number"}
CERTIFICATE_NUMBER_ALIASES = {"сертифікат", "номер сертифікату", "certificate", "certificate_number"}
CERTIFICATE_ISSUE_DATE_ALIASES = {"дата видачі сертифікату", "дата видачі сертифіката", "certificate_issue_date"}
POSTAL_INDEX_ALIASES = {"індекс", "поштовий індекс", "postal_index"}
ADDRESS_ALIASES = {"адреса", "address"}
PASSPORT_SERIES_ALIASES = {"паспорт: серія", "паспорт серія", "серія паспорта", "passport_series"}
PASSPORT_NUMBER_ALIASES = {"паспорт: №", "паспорт №", "номер паспорта", "passport_number"}
PASSPORT_ISSUED_BY_ALIASES = {"ким виданий", "кем выдан", "passport_issued_by"}
PASSPORT_ISSUED_DATE_ALIASES = {"коли виданий", "дата видачі паспорта", "passport_issued_date"}
TAX_ID_ALIASES = {"ідентифікаційний код", "ідентифікаційний номер", "інн", "рнокпп", "tax_id"}
PHONE_ALIASES = {"телефон", "номер телефону", "phone"}

TRAINEE_HEADER_HINTS = (
    FIRST_NAME_ALIASES
    | LAST_NAME_ALIASES
    | FULL_NAME_ALIASES
    | BIRTH_DATE_ALIASES
    | STATUS_ALIASES
    | GROUP_CODE_ALIASES
    | GROUP_NAME_ALIASES
)
GROUP_CONTEXT_PATTERN = re.compile(r"\bгрупа\s*([0-9a-zа-яіїєґ\/\-]+)\b", re.IGNORECASE)


def _normalize_header(value: Any) -> str:
    raw = str(value or "").strip().lower()
    raw = " ".join(raw.replace("\n", " ").replace("\r", " ").split())
    return raw.replace("’", "'").replace("`", "'")


def _normalize_text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.replace("\n", " ").replace("\r", " ").split()).strip()
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
    return str(value).strip()


def _pick_sheet_name(sheet_names: list[str]) -> str | None:
    if not sheet_names:
        return None
    lowered = {name.lower(): name for name in sheet_names}
    for preferred in PREFERRED_TRAINEE_SHEET_NAMES:
        if preferred.lower() in lowered:
            return lowered[preferred.lower()]
    return sheet_names[0]


def _row_is_empty(row: list[Any]) -> bool:
    return all(_normalize_text_value(value) == "" for value in row)


def _find_header_row_index(rows: list[list[Any]]) -> int:
    best_index = 0
    best_score = -1
    for idx, row in enumerate(rows[:HEADER_SCAN_LIMIT]):
        normalized = [_normalize_header(value) for value in row]
        non_empty = [value for value in normalized if value]
        if not non_empty:
            continue
        score = sum(1 for value in non_empty if value in TRAINEE_HEADER_HINTS)
        if score > best_score:
            best_score = score
            best_index = idx
            if score >= 2:
                break
    return best_index


def _make_unique_headers(raw_headers: list[Any]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for idx, header in enumerate(raw_headers):
        candidate = _normalize_text_value(header) or f"column_{idx + 1}"
        base = candidate
        count = used.get(base, 0)
        if count:
            candidate = f"{base}_{count + 1}"
        used[base] = count + 1
        headers.append(candidate)
    return headers


def _rows_from_xlsx(file_path: str) -> tuple[str | None, list[list[Any]]]:
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet_name = _pick_sheet_name(workbook.sheetnames)
    if not sheet_name:
        workbook.close()
        return None, []
    worksheet = workbook[sheet_name]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    workbook.close()
    return sheet_name, rows


def _rows_from_xls(file_path: str) -> tuple[str | None, list[list[Any]]]:
    workbook = xlrd.open_workbook(file_path)
    sheet_name = _pick_sheet_name(workbook.sheet_names())
    if not sheet_name:
        return None, []
    sheet = workbook.sheet_by_name(sheet_name)
    rows: list[list[Any]] = []
    for row_index in range(sheet.nrows):
        row: list[Any] = []
        for col_index in range(sheet.ncols):
            cell = sheet.cell(row_index, col_index)
            value: Any = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    value = xldate_as_datetime(cell.value, workbook.datemode)
                except Exception:
                    value = cell.value
            row.append(value)
        rows.append(row)
    return sheet_name, rows


def _tabular_preview(data: list[dict[str, Any]]) -> list[dict[str, Any]]:
    preview: list[dict[str, Any]] = []
    for row in data[:ROW_SAMPLE_LIMIT]:
        payload = {str(key): _normalize_text_value(value) for key, value in row.items()}
        preview.append(payload)
    return preview


def _parse_tabular_content(file_path: str) -> dict[str, Any]:
    extension = Path(file_path).suffix.lower()
    if extension == ".xls":
        sheet_name, rows = _rows_from_xls(file_path)
    else:
        sheet_name, rows = _rows_from_xlsx(file_path)

    if not rows:
        return {"rows": 0, "headers": [], "data": [], "sheet_name": sheet_name}

    header_idx = _find_header_row_index(rows)
    group_context = _extract_group_context(rows, header_idx)
    headers = _make_unique_headers(rows[header_idx])
    data: list[dict[str, Any]] = []
    for raw_row in rows[header_idx + 1 :]:
        if _row_is_empty(raw_row):
            continue
        payload: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            payload[header] = raw_row[idx] if idx < len(raw_row) else None
        data.append(payload)
    return {
        "rows": len(data),
        "headers": headers,
        "data": data,
        "sheet_name": sheet_name,
        "header_row_index": header_idx + 1,
        **group_context,
    }


def _extract_group_context(rows: list[list[Any]], header_idx: int) -> dict[str, Any]:
    if not rows:
        return {}

    for row in reversed(rows[:header_idx]):
        for cell in row:
            raw_text = _normalize_text_value(cell)
            if not raw_text:
                continue
            match = GROUP_CONTEXT_PATTERN.search(raw_text)
            if not match:
                continue
            group_code = match.group(1).strip(" \"'«».,:;()[]{}")
            group_name = raw_text[match.end() :].strip(" \"'«».,:;()[]{}-")
            payload: dict[str, Any] = {"default_group_code": group_code}
            if group_name:
                payload["default_group_name"] = group_name
            payload["group_context_source"] = raw_text
            return payload
    return {}


def _parse_date_value(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial dates are usually in this interval.
        if 20000 <= float(value) <= 60000:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        return None
    text = _normalize_text_value(value)
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _split_full_name(full_name: str) -> tuple[str, str, str]:
    parts = [part for part in full_name.split(" ") if part]
    if not parts:
        return "", "", ""
    if len(parts) == 1:
        return parts[0], "", ""
    if len(parts) == 2:
        return parts[0], parts[1], ""
    return parts[0], parts[1], " ".join(parts[2:])


def _first_non_empty(keymap: dict[str, Any], aliases: set[str]) -> Any:
    for alias in aliases:
        value = keymap.get(alias)
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    for key, value in keymap.items():
        for alias in aliases:
            if alias in key:
                if value is None:
                    continue
                if isinstance(value, str) and not value.strip():
                    continue
                return value
    return None


def _has_any_alias(headers: set[str], aliases: set[str]) -> bool:
    for header in headers:
        for alias in aliases:
            if header == alias or alias in header:
                return True
    return False


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _set_plain_value(obj: Any, field: str, value: Any, overwrite: bool) -> bool:
    if _is_missing(value):
        return False
    current = getattr(obj, field)
    if overwrite:
        if current != value:
            setattr(obj, field, value)
            return True
        return False
    if _is_missing(current):
        setattr(obj, field, value)
        return True
    return False


def _set_encrypted_value(obj: Any, field: str, plain_value: str | None, overwrite: bool) -> bool:
    normalized = _normalize_text_value(plain_value)
    if not normalized:
        return False

    current_cipher = getattr(obj, field)
    if overwrite:
        current_plain = _normalize_text_value(cipher.decrypt(current_cipher))
        if current_plain == normalized:
            return False
        setattr(obj, field, cipher.encrypt(normalized))
        return True

    if _is_missing(current_cipher):
        setattr(obj, field, cipher.encrypt(normalized))
        return True
    return False


def _ensure_group_for_trainee(
    db: Session,
    trainee: Trainee,
    branch_id: str,
    group_cache: dict[str, Group],
    group_code_raw: str | None,
    group_name_raw: str | None,
    overwrite_group: bool = False,
) -> tuple[int, bool]:
    group_code = _normalize_text_value(group_code_raw)
    group_name = _normalize_text_value(group_name_raw)
    if not group_code and not group_name:
        return 0, False

    incoming_code = (group_code or f"AUTO-{group_name[:32] or trainee.id}")[:50]
    effective_code = incoming_code
    if trainee.group_code and not overwrite_group:
        effective_code = trainee.group_code

    trainee_group_changed = trainee.group_code != effective_code
    trainee.group_code = effective_code

    cache_key = effective_code.lower()
    group = group_cache.get(cache_key)
    if not group:
        group = (
            db.query(Group)
            .filter(Group.branch_id == branch_id, Group.code == effective_code)
            .first()
        )
        if not group:
            group = Group(
                branch_id=branch_id,
                code=effective_code,
                name=(group_name or effective_code)[:255],
                status=GroupStatus.ACTIVE,
                capacity=30,
            )
            db.add(group)
            db.flush()
        group_cache[cache_key] = group

    membership_exists = (
        db.query(GroupMembership)
        .filter(
            GroupMembership.group_id == group.id,
            GroupMembership.trainee_id == trainee.id,
        )
        .first()
    )
    if membership_exists:
        return 0, trainee_group_changed

    db.add(
        GroupMembership(
            group_id=group.id,
            trainee_id=trainee.id,
            status=MembershipStatus.ACTIVE,
        )
    )
    return 1, True


def parse_document_content(file_path: str, doc_type: DocumentType) -> dict:
    if doc_type == DocumentType.XLSX:
        return _parse_tabular_content(file_path)

    if doc_type == DocumentType.DOCX:
        doc = DocxDocument(file_path)
        text = "\n".join(paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip())
        return {"rows": 1, "text_preview": text[:3000]}

    if doc_type == DocumentType.PDF:
        reader = PdfReader(file_path)
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
        return {"rows": len(reader.pages), "text_preview": text[:3000]}

    if doc_type == DocumentType.CSV:
        for encoding in ("utf-8-sig", "cp1251", "utf-8"):
            try:
                with open(file_path, "r", encoding=encoding, newline="") as handle:
                    reader = csv.DictReader(handle)
                    data = list(reader)
                return {"rows": len(data), "headers": reader.fieldnames or [], "data": data}
            except UnicodeDecodeError:
                continue
        return {"rows": 0, "headers": [], "data": [], "note": "Не вдалося декодувати CSV"}

    return {"rows": 0, "data": []}


def _looks_like_trainee_registry(parsed: dict) -> bool:
    headers = {_normalize_header(h) for h in parsed.get("headers", [])}
    return bool(
        (_has_any_alias(headers, FIRST_NAME_ALIASES) and _has_any_alias(headers, LAST_NAME_ALIASES))
        or _has_any_alias(headers, FULL_NAME_ALIASES)
    )


def _extract_trainee_payload(
    row: dict[str, Any],
    default_group_code: str = "",
    default_group_name: str = "",
) -> dict[str, Any] | None:
    keymap = {_normalize_header(k): v for k, v in row.items()}

    first_name = _normalize_text_value(_first_non_empty(keymap, FIRST_NAME_ALIASES))
    last_name = _normalize_text_value(_first_non_empty(keymap, LAST_NAME_ALIASES))
    middle_name = _normalize_text_value(_first_non_empty(keymap, MIDDLE_NAME_ALIASES))

    if not first_name or not last_name:
        full_name_raw = _normalize_text_value(_first_non_empty(keymap, FULL_NAME_ALIASES))
        if full_name_raw:
            parsed_last, parsed_first, parsed_middle = _split_full_name(full_name_raw)
            last_name = last_name or parsed_last
            first_name = first_name or parsed_first
            middle_name = middle_name or parsed_middle

    if not first_name or not last_name:
        return None

    source_row_number_raw = _first_non_empty(keymap, ROW_NUMBER_ALIASES)
    source_row_number: int | None = None
    try:
        if source_row_number_raw is not None and _normalize_text_value(source_row_number_raw):
            source_row_number = int(float(_normalize_text_value(source_row_number_raw)))
    except ValueError:
        source_row_number = None

    status_value = _normalize_text_value(_first_non_empty(keymap, STATUS_ALIASES)).lower() or "active"
    return {
        "first_name": first_name,
        "last_name": last_name,
        "middle_name": middle_name,
        "birth_date": _parse_date_value(_first_non_empty(keymap, BIRTH_DATE_ALIASES)),
        "source_row_number": source_row_number,
        "employment_center": _normalize_text_value(_first_non_empty(keymap, EMPLOYMENT_CENTER_ALIASES)) or None,
        "contract_number": _normalize_text_value(_first_non_empty(keymap, CONTRACT_NUMBER_ALIASES)) or None,
        "certificate_number": _normalize_text_value(_first_non_empty(keymap, CERTIFICATE_NUMBER_ALIASES)) or None,
        "certificate_issue_date": _parse_date_value(_first_non_empty(keymap, CERTIFICATE_ISSUE_DATE_ALIASES)),
        "postal_index": _normalize_text_value(_first_non_empty(keymap, POSTAL_INDEX_ALIASES)) or None,
        "address": _normalize_text_value(_first_non_empty(keymap, ADDRESS_ALIASES)) or None,
        "passport_series": _normalize_text_value(_first_non_empty(keymap, PASSPORT_SERIES_ALIASES)) or None,
        "passport_number": _normalize_text_value(_first_non_empty(keymap, PASSPORT_NUMBER_ALIASES)) or None,
        "passport_issued_by": _normalize_text_value(_first_non_empty(keymap, PASSPORT_ISSUED_BY_ALIASES)) or None,
        "passport_issued_date": _parse_date_value(_first_non_empty(keymap, PASSPORT_ISSUED_DATE_ALIASES)),
        "tax_id": _normalize_text_value(_first_non_empty(keymap, TAX_ID_ALIASES)) or None,
        "phone_value": _normalize_text_value(_first_non_empty(keymap, PHONE_ALIASES)) or None,
        "status": status_value if status_value in {"active", "completed", "expelled"} else "active",
        "group_code": _normalize_text_value(_first_non_empty(keymap, GROUP_CODE_ALIASES)) or default_group_code,
        "group_name": _normalize_text_value(_first_non_empty(keymap, GROUP_NAME_ALIASES)) or default_group_name,
    }


def _find_existing_trainee(
    db: Session,
    branch_id: str,
    first_name: str,
    last_name: str,
    middle_name: str,
    birth_date: date | None,
    contract_number: str | None,
) -> tuple[Trainee | None, str | None]:
    if contract_number:
        existing = (
            db.query(Trainee)
            .filter(
                Trainee.branch_id == branch_id,
                Trainee.contract_number == contract_number,
            )
            .first()
        )
        if existing:
            return existing, "contract_number"

    normalized_first_name = first_name if not middle_name else f"{first_name} {middle_name}"
    existing_query = db.query(Trainee).filter(
        Trainee.branch_id == branch_id,
        Trainee.first_name == normalized_first_name,
        Trainee.last_name == last_name,
    )
    if birth_date:
        existing_query = existing_query.filter(
            or_(Trainee.birth_date == birth_date, Trainee.birth_date.is_(None))
        )
    existing = existing_query.first()
    if existing:
        return existing, "name_birth_date"

    if middle_name:
        fallback_query = db.query(Trainee).filter(
            Trainee.branch_id == branch_id,
            Trainee.last_name == last_name,
            Trainee.first_name.ilike(f"{first_name}%"),
        )
        if birth_date:
            fallback_query = fallback_query.filter(
                or_(Trainee.birth_date == birth_date, Trainee.birth_date.is_(None))
            )
        existing = fallback_query.first()
        if existing:
            return existing, "partial_name_birth_date"

    return None, None


def analyze_trainee_import_duplicates(
    db: Session,
    parsed: dict,
    branch_id: str,
    preview_limit: int = 10,
) -> dict[str, Any]:
    if not _looks_like_trainee_registry(parsed):
        return {
            "new_count": 0,
            "duplicate_count": 0,
            "invalid_count": 0,
            "duplicate_preview": [],
            "note": "Структура не схожа на реєстр слухачів",
        }

    default_group_code = _normalize_text_value(parsed.get("default_group_code"))
    default_group_name = _normalize_text_value(parsed.get("default_group_name"))
    new_count = 0
    duplicate_count = 0
    invalid_count = 0
    duplicate_preview: list[dict[str, Any]] = []

    for index, row in enumerate(parsed.get("data", []), start=1):
        payload = _extract_trainee_payload(row, default_group_code, default_group_name)
        if not payload:
            invalid_count += 1
            continue

        existing, match_reason = _find_existing_trainee(
            db,
            branch_id,
            payload["first_name"],
            payload["last_name"],
            payload["middle_name"],
            payload["birth_date"],
            payload["contract_number"],
        )
        if not existing:
            new_count += 1
            continue

        duplicate_count += 1
        if len(duplicate_preview) < preview_limit:
            incoming_name = f"{payload['last_name']} {payload['first_name']}"
            if payload["middle_name"]:
                incoming_name = f"{incoming_name} {payload['middle_name']}"
            duplicate_preview.append(
                {
                    "row_number": payload["source_row_number"] or index,
                    "incoming_name": incoming_name,
                    "contract_number": payload["contract_number"],
                    "group_code": payload["group_code"],
                    "existing_id": existing.id,
                    "existing_name": f"{existing.last_name} {existing.first_name}",
                    "match_reason": match_reason,
                }
            )

    return {
        "new_count": new_count,
        "duplicate_count": duplicate_count,
        "invalid_count": invalid_count,
        "duplicate_preview": duplicate_preview,
    }


def try_import_trainees(
    db: Session,
    parsed: dict,
    branch_id: str,
    update_existing_mode: str = "missing_only",
) -> dict:
    if update_existing_mode not in IMPORT_UPDATE_MODES:
        update_existing_mode = "missing_only"
    overwrite_existing = update_existing_mode == "overwrite"

    if not _looks_like_trainee_registry(parsed):
        return {"inserted": 0, "skipped_invalid": 0, "skipped_existing": 0, "note": "Структура не схожа на реєстр слухачів"}

    inserted = 0
    updated_existing = 0
    skipped_invalid = 0
    skipped_existing = 0
    memberships_created = 0
    inserted_ids: list[int] = []

    default_group_code = _normalize_text_value(parsed.get("default_group_code"))
    default_group_name = _normalize_text_value(parsed.get("default_group_name"))

    group_cache: dict[str, Group] = {}
    for row in parsed.get("data", []):
        payload = _extract_trainee_payload(row, default_group_code, default_group_name)
        if not payload:
            skipped_invalid += 1
            continue

        first_name = payload["first_name"]
        last_name = payload["last_name"]
        middle_name = payload["middle_name"]
        birth_date = payload["birth_date"]
        source_row_number = payload["source_row_number"]
        employment_center = payload["employment_center"]
        contract_number = payload["contract_number"]
        certificate_number = payload["certificate_number"]
        certificate_issue_date = payload["certificate_issue_date"]
        postal_index = payload["postal_index"]
        address = payload["address"]
        passport_series = payload["passport_series"]
        passport_number = payload["passport_number"]
        passport_issued_by = payload["passport_issued_by"]
        passport_issued_date = payload["passport_issued_date"]
        tax_id = payload["tax_id"]
        phone_value = payload["phone_value"]
        status = payload["status"]
        group_code = payload["group_code"]
        group_name = payload["group_name"]

        existing, _match_reason = _find_existing_trainee(
            db,
            branch_id,
            first_name,
            last_name,
            middle_name,
            birth_date,
            contract_number,
        )
        if existing:
            if update_existing_mode == "skip_existing":
                skipped_existing += 1
                continue

            changed = False
            changed = _set_plain_value(existing, "source_row_number", source_row_number, overwrite_existing) or changed
            changed = _set_plain_value(existing, "birth_date", birth_date, overwrite_existing) or changed
            changed = _set_plain_value(existing, "contract_number", contract_number, overwrite_existing) or changed
            changed = _set_plain_value(existing, "certificate_number", certificate_number, overwrite_existing) or changed
            changed = _set_plain_value(existing, "certificate_issue_date", certificate_issue_date, overwrite_existing) or changed
            changed = _set_plain_value(existing, "postal_index", postal_index, overwrite_existing) or changed
            changed = _set_plain_value(existing, "passport_issued_date", passport_issued_date, overwrite_existing) or changed
            changed = _set_plain_value(existing, "status", status, overwrite_existing) or changed

            changed = _set_encrypted_value(existing, "employment_center_encrypted", employment_center, overwrite_existing) or changed
            changed = _set_encrypted_value(existing, "address_encrypted", address, overwrite_existing) or changed
            changed = _set_encrypted_value(existing, "passport_series_encrypted", passport_series, overwrite_existing) or changed
            changed = _set_encrypted_value(existing, "passport_number_encrypted", passport_number, overwrite_existing) or changed
            changed = _set_encrypted_value(existing, "passport_issued_by_encrypted", passport_issued_by, overwrite_existing) or changed
            changed = _set_encrypted_value(existing, "tax_id_encrypted", tax_id, overwrite_existing) or changed
            changed = _set_encrypted_value(existing, "phone_encrypted", phone_value, overwrite_existing) or changed

            memberships_added, group_changed = _ensure_group_for_trainee(
                db,
                existing,
                branch_id,
                group_cache,
                group_code,
                group_name,
                overwrite_group=overwrite_existing,
            )
            memberships_created += memberships_added
            changed = changed or group_changed

            if changed:
                db.add(existing)
                updated_existing += 1
            else:
                skipped_existing += 1
            continue

        trainee = Trainee(
            branch_id=branch_id,
            source_row_number=source_row_number,
            first_name=first_name if not middle_name else f"{first_name} {middle_name}",
            last_name=last_name,
            birth_date=birth_date,
            employment_center_encrypted=cipher.encrypt(employment_center),
            contract_number=contract_number,
            certificate_number=certificate_number,
            certificate_issue_date=certificate_issue_date,
            postal_index=postal_index,
            address_encrypted=cipher.encrypt(address),
            passport_series_encrypted=cipher.encrypt(passport_series),
            passport_number_encrypted=cipher.encrypt(passport_number),
            passport_issued_by_encrypted=cipher.encrypt(passport_issued_by),
            passport_issued_date=passport_issued_date,
            tax_id_encrypted=cipher.encrypt(tax_id),
            status=status,
            phone_encrypted=cipher.encrypt(phone_value),
        )
        db.add(trainee)
        db.flush()
        inserted += 1
        inserted_ids.append(trainee.id)

        memberships_added, _ = _ensure_group_for_trainee(
            db,
            trainee,
            branch_id,
            group_cache,
            group_code,
            group_name,
            overwrite_group=True,
        )
        memberships_created += memberships_added

    db.commit()
    return {
        "inserted": inserted,
        "updated_existing": updated_existing,
        "inserted_ids": inserted_ids,
        "skipped_invalid": skipped_invalid,
        "skipped_existing": skipped_existing,
        "memberships_created": memberships_created,
        "already_loaded": skipped_existing > 0 and inserted == 0 and updated_existing == 0,
        "sheet_name": parsed.get("sheet_name"),
        "default_group_code": default_group_code or None,
        "default_group_name": default_group_name or None,
        "update_existing_mode": update_existing_mode,
        "preview": _tabular_preview(parsed.get("data", [])),
    }


def collect_teacher_detailed_workload(
    db: Session,
    branch_id: str,
    teacher_ids: list[int],
    date_from: date | None = None,
    date_to: date | None = None,
) -> dict[str, list[dict]]:
    teachers = (
        db.query(Teacher)
        .filter(Teacher.branch_id == branch_id, Teacher.id.in_(teacher_ids))
        .all()
    )
    
    teacher_map = {t.id: f"{t.last_name} {t.first_name}" for t in teachers}
    result: dict[str, list[dict]] = {name: [] for name in teacher_map.values()}
    
    query = (
        db.query(ScheduleSlot, Group)
        .join(Group, Group.id == ScheduleSlot.group_id)
        .filter(Group.branch_id == branch_id, ScheduleSlot.teacher_id.in_(teacher_ids))
    )
    if date_from:
        query = query.filter(ScheduleSlot.starts_at >= datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc))
    if date_to:
        query = query.filter(ScheduleSlot.starts_at <= datetime.combine(date_to, datetime.max.time(), tzinfo=timezone.utc))
        
    query = query.order_by(ScheduleSlot.starts_at)
    slots = query.all()

    for slot, group in slots:
        teacher_name = teacher_map.get(slot.teacher_id)
        if not teacher_name:
            continue
            
        academic_hours = slot.academic_hours
        if academic_hours is None:
            academic_hours = (slot.ends_at - slot.starts_at).total_seconds() / 3600
            
        result[teacher_name].append({
            "Номер групи": group.code,
            "Назва групи": group.name or "",
            "Дата (день)": slot.starts_at.astimezone().strftime("%Y-%m-%d"),
            "Пара": slot.pair_number if slot.pair_number is not None else "",
            "Кількість годин": round(float(academic_hours), 2)
        })
        
    return result

def collect_teacher_workload_summary(
    db: Session,
    branch_id: str,
    date_from: date | None = None,
    date_to: date | None = None,
) -> list[dict]:
    teachers = (
        db.query(Teacher)
        .filter(Teacher.branch_id == branch_id, Teacher.is_active.is_(True))
        .all()
    )
    query = (
        db.query(ScheduleSlot)
        .join(Group, Group.id == ScheduleSlot.group_id)
        .filter(Group.branch_id == branch_id)
    )
    if date_from:
        query = query.filter(ScheduleSlot.starts_at >= datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc))
    if date_to:
        query = query.filter(ScheduleSlot.starts_at <= datetime.combine(date_to, datetime.max.time(), tzinfo=timezone.utc))
    slots = query.all()

    totals: dict[int, float] = {}
    for slot in slots:
        totals.setdefault(slot.teacher_id, 0.0)
        if slot.academic_hours is not None:
            totals[slot.teacher_id] += float(slot.academic_hours)
        else:
            totals[slot.teacher_id] += (slot.ends_at - slot.starts_at).total_seconds() / 3600

    rows: list[dict] = []
    for teacher in teachers:
        total_hours = round(totals.get(teacher.id, 0.0), 2)
        annual_load = round(float(teacher.annual_load_hours or 0.0), 2)
        # For report visibility include teachers who already teach or have planned annual load.
        if total_hours <= 0 and annual_load <= 0:
            continue
        remaining = round(max(annual_load - total_hours, 0.0), 2)
        rows.append(
            {
                "teacher_id": teacher.id,
                "teacher_name": f"{teacher.last_name} {teacher.first_name}",
                "total_hours": total_hours,
                "annual_load_hours": annual_load,
                "remaining_hours": remaining,
            }
        )
    rows.sort(key=lambda item: item["teacher_name"].lower())
    for idx, row in enumerate(rows, start=1):
        row["row_number"] = idx
    return rows


def collect_report_rows(db: Session, report_type: str, branch_id: str, request_payload: dict | None = None) -> list[dict] | dict[str, list[dict]]:
    if report_type == "trainees":
        trainees = db.query(Trainee).filter(Trainee.branch_id == branch_id).all()
        return [
            {
                "id": trainee.id,
                "first_name": trainee.first_name,
                "last_name": trainee.last_name,
                "status": trainee.status,
                "created_at": trainee.created_at.isoformat(),
            }
            for trainee in trainees
        ]

    if report_type == "teacher_workload":
        if request_payload and request_payload.get("teacher_ids"):
            teacher_ids = request_payload["teacher_ids"]
            start_date_str = request_payload.get("start_date")
            end_date_str = request_payload.get("end_date")
            start_date = date.fromisoformat(start_date_str) if start_date_str else None
            end_date = date.fromisoformat(end_date_str) if end_date_str else None
            return collect_teacher_detailed_workload(db, branch_id, teacher_ids, start_date, end_date)
            
        summary = collect_teacher_workload_summary(db, branch_id)
        return [
            {
                "Номер за порядком": row["row_number"],
                "Прізвище, ім'я та по батькові викладача": row["teacher_name"],
                "Загальна кількість годин": row["total_hours"],
                "Річне педнавантаження": row["annual_load_hours"],
                "Залишок годин": row["remaining_hours"],
            }
            for row in summary
        ]

    if report_type == "kpi":
        active_groups = (
            db.query(GroupMembership)
            .join(Group, Group.id == GroupMembership.group_id)
            .filter(GroupMembership.status == MembershipStatus.ACTIVE)
            .filter(Group.branch_id == branch_id)
            .count()
        )
        progress = db.query(Performance).filter(Performance.branch_id == branch_id).all()
        avg_progress = round(sum(record.progress_pct for record in progress) / len(progress), 2) if progress else 0.0
        return [
            {"metric": "active_memberships", "value": active_groups},
            {"metric": "avg_training_progress_pct", "value": avg_progress},
            {"metric": "generated_at", "value": datetime.now(timezone.utc).isoformat()},
        ]

    if report_type == "form_1pa":
        trainees_total = db.query(Trainee).filter(Trainee.branch_id == branch_id).count()
        completed = (
            db.query(Trainee)
            .filter(Trainee.branch_id == branch_id, Trainee.status == "completed")
            .count()
        )
        employed = (
            db.query(Performance)
            .filter(Performance.branch_id == branch_id, Performance.employment_flag.is_(True))
            .count()
        )
        return [
            {"field": "period", "value": date.today().isoformat()},
            {"field": "trainees_total", "value": trainees_total},
            {"field": "trainees_completed", "value": completed},
            {"field": "employed_after_training", "value": employed},
            {"field": "employment_rate_estimate", "value": round((employed / max(completed, 1)) * 100, 2)},
        ]

    if report_type == "employment":
        rows = (
            db.query(Performance, Trainee, Group)
            .join(Trainee, Trainee.id == Performance.trainee_id)
            .join(Group, Group.id == Performance.group_id)
            .filter(Performance.branch_id == branch_id, Group.branch_id == branch_id, Trainee.branch_id == branch_id)
            .all()
        )
        return [
            {
                "trainee_id": performance.trainee_id,
                "trainee_name": f"{trainee.last_name} {trainee.first_name}",
                "group_code": group.code,
                "progress_pct": performance.progress_pct,
                "attendance_pct": performance.attendance_pct,
                "employment_flag": performance.employment_flag,
            }
            for performance, trainee, group in rows
        ]

    if report_type == "financial":
        teachers = db.query(Teacher).filter(Teacher.branch_id == branch_id).all()
        slots = (
            db.query(ScheduleSlot)
            .join(Group, Group.id == ScheduleSlot.group_id)
            .filter(Group.branch_id == branch_id)
            .all()
        )
        rows: list[dict] = []
        total_amount = 0.0
        for teacher in teachers:
            total_hours = 0.0
            for slot in slots:
                if slot.teacher_id == teacher.id:
                    if slot.academic_hours is not None:
                        total_hours += float(slot.academic_hours)
                    else:
                        total_hours += (slot.ends_at - slot.starts_at).total_seconds() / 3600
            amount = round(total_hours * teacher.hourly_rate, 2)
            total_amount += amount
            rows.append(
                {
                    "teacher_id": teacher.id,
                    "teacher_name": f"{teacher.last_name} {teacher.first_name}",
                    "hourly_rate": teacher.hourly_rate,
                    "total_hours": round(total_hours, 2),
                    "amount_uah": amount,
                }
            )
        rows.append(
            {
                "teacher_id": "",
                "teacher_name": "TOTAL",
                "hourly_rate": "",
                "total_hours": "",
                "amount_uah": round(total_amount, 2),
            }
        )
        return rows

    return []


def _resolve_pdf_font_path() -> Path:
    for candidate in PDF_FONT_CANDIDATES:
        value = os.environ.get(candidate) if candidate == "PDF_FONT_PATH" else candidate
        if not value:
            continue
        path = Path(value)
        if not path.is_absolute():
            path = Path.cwd() / path
        if path.exists():
            return path
    raise RuntimeError(
        "Не знайдено Unicode-шрифт для PDF. Вкажіть PDF_FONT_PATH або додайте DejaVuSans.ttf до backend/assets/fonts."
    )


def _configure_pdf(pdf: FPDF) -> str:
    font_path = _resolve_pdf_font_path()
    font_family = "ReportFont"
    pdf.add_font(font_family, "", str(font_path))
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_creator("SUPTC")
    pdf.set_author("SUPTC")
    return font_family


def _format_report_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Так" if value else "Ні"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(round(value, 2))
    return str(value)


def _prepare_pdf_sections(report_rows: list[dict] | dict[str, list[dict]]) -> list[tuple[str | None, list[dict]]]:
    if isinstance(report_rows, dict):
        return [(str(section_name), rows) for section_name, rows in report_rows.items()]
    return [(None, report_rows)]


def _pdf_column_widths(headers: list[str], rows: list[dict], available_width: float) -> list[float]:
    if not headers:
        return [available_width]
    weights: list[float] = []
    for header in headers:
        sample_lengths = [len(header)]
        for row in rows[:30]:
            sample_lengths.append(min(len(_format_report_value(row.get(header))), 60))
        weights.append(max(10, min(max(sample_lengths), 42)))

    total = sum(weights) or 1
    min_width = 18
    widths = [max(min_width, available_width * weight / total) for weight in weights]
    width_sum = sum(widths)
    if width_sum > available_width:
        scale = available_width / width_sum
        widths = [width * scale for width in widths]
    return widths


def _write_pdf_table(pdf: FPDF, headers: list[str], rows: list[dict], font_family: str) -> None:
    if not rows:
        pdf.set_font(font_family, size=9)
        pdf.multi_cell(0, 6, text="Дані відсутні", new_x="LMARGIN", new_y="NEXT")
        return

    body_rows = rows[:PDF_REPORT_ROW_LIMIT]
    font_size = 7 if len(headers) > 6 else 8
    line_height = 4.6 if len(headers) > 6 else 5.2
    pdf.set_font(font_family, size=font_size)
    col_widths = _pdf_column_widths(headers, body_rows, pdf.epw)
    headings_style = FontFace(fill_color=(232, 238, 244), color=(15, 23, 42))

    with pdf.table(
        width=pdf.epw,
        col_widths=col_widths,
        line_height=line_height,
        headings_style=headings_style,
        text_align="LEFT",
        v_align="TOP",
        wrapmode="CHAR",
        padding=(1.2, 1.1),
        repeat_headings=1,
    ) as table:
        header_row = table.row()
        for header in headers:
            header_row.cell(header)

        for source_row in body_rows:
            row = table.row()
            for header in headers:
                row.cell(_format_report_value(source_row.get(header))[:500])

    if len(rows) > PDF_REPORT_ROW_LIMIT:
        pdf.ln(2)
        pdf.set_font(font_family, size=8)
        pdf.multi_cell(
            0,
            5,
            text=f"Показано перші {PDF_REPORT_ROW_LIMIT} рядків із {len(rows)}.",
            new_x="LMARGIN",
            new_y="NEXT",
        )


def _save_report_pdf(report_rows: list[dict] | dict[str, list[dict]], report_type: str, out_file: Path) -> None:
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_margins(10, 10, 10)
    pdf.add_page()
    font_family = _configure_pdf(pdf)

    title = REPORT_TYPE_LABELS.get(report_type, report_type)
    pdf.set_title(title)
    pdf.set_font(font_family, size=15)
    pdf.multi_cell(0, 8, text=f"Звіт: {title}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(font_family, size=8)
    pdf.set_text_color(71, 85, 105)
    pdf.multi_cell(
        0,
        5,
        text=f"Сформовано: {datetime.now(timezone.utc).astimezone().strftime('%d.%m.%Y %H:%M')}",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    for index, (section_name, rows) in enumerate(_prepare_pdf_sections(report_rows)):
        if index > 0:
            pdf.add_page()
        if section_name:
            pdf.set_font(font_family, size=11)
            pdf.multi_cell(0, 6, text=section_name, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)

        headers = list(rows[0].keys()) if rows else ["Дані"]
        _write_pdf_table(pdf, headers, rows, font_family)

    pdf.output(str(out_file))


def save_report_file(report_rows: list[dict] | dict[str, list[dict]], report_type: str, export_format: str, request_payload: dict | None = None) -> tuple[str, DocumentType]:
    out_dir = storage_path()
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    base_name = f"{report_type}_{stamp}_{uuid4().hex[:8]}"

    flat_report_rows = []
    if isinstance(report_rows, dict):
        for name, rows in report_rows.items():
            for r in rows:
                flat_report_rows.append({**r, "Вкладка": name})
    else:
        flat_report_rows = report_rows

    if export_format == "csv":
        out_file = out_dir / f"{base_name}.csv"
        fieldnames = list(flat_report_rows[0].keys()) if flat_report_rows else ["empty"]
        with out_file.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for row in flat_report_rows:
                writer.writerow(row)
        return str(out_file), DocumentType.CSV

    if export_format == "xlsx":
        workbook = Workbook()
        
        if isinstance(report_rows, dict):
            # Handle multi-sheet dictionary
            first_sheet = True
            for sheet_name, rows in report_rows.items():
                if first_sheet:
                    sheet = workbook.active
                    sheet.title = str(sheet_name)[:31]  # Excel limits sheet names to 31 chars
                    first_sheet = False
                else:
                    sheet = workbook.create_sheet(title=str(sheet_name)[:31])
                    
                if rows:
                    headers = list(rows[0].keys())
                    sheet.append(headers)
                    for row in rows:
                        sheet.append([row.get(col) for col in headers])
                else:
                    sheet.append(["empty"])
        else:
            # Handle single list
            sheet = workbook.active
            if report_rows:
                headers = list(report_rows[0].keys())
                sheet.append(headers)
                for row in report_rows:
                    sheet.append([row.get(col) for col in headers])
            else:
                sheet.append(["empty"])
                
        out_file = out_dir / f"{base_name}.xlsx"
        workbook.save(out_file)
        return str(out_file), DocumentType.XLSX

    out_file = out_dir / f"{base_name}.pdf"
    _save_report_pdf(report_rows, report_type, out_file)
    return str(out_file), DocumentType.PDF


def mark_job_running(job: ImportJob | ExportJob) -> None:
    job.status = JobStatus.RUNNING
    job.started_at = datetime.now(timezone.utc)


def mark_job_success(job: ImportJob | ExportJob, result_payload: dict, message: str | None = None) -> None:
    job.status = JobStatus.SUCCEEDED
    job.result_payload = result_payload
    job.message = message
    job.finished_at = datetime.now(timezone.utc)


def mark_job_failed(job: ImportJob | ExportJob, message: str) -> None:
    job.status = JobStatus.FAILED
    job.message = message
    job.finished_at = datetime.now(timezone.utc)
    job.retries = (job.retries or 0) + 1
