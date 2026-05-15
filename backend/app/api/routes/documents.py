from uuid import uuid4
from pathlib import Path
import shutil
import tempfile

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, UploadFile, status
from fastapi.responses import FileResponse
from celery.utils.log import get_task_logger
from openpyxl import load_workbook

from app.api.deps import CurrentUser, DbSession, apply_branch_scope, ensure_same_branch, require_roles
from app.models import Document, ExportJob, Group, ImportJob, JobStatus, RoleName, ScheduleSlot
from app.schemas.api import (
    BatchImportFormatsResponse,
    BatchImportResponse,
    BatchImportSkippedFile,
    ExportRequest,
    ImportPreviewGroup,
    ImportPreviewResponse,
    JobResponse,
)
from app.services.audit import write_audit
from app.services.cache import cache_get_json, cache_set_json
from app.services.import_registry import get_batch_import_format, supported_batch_import_extensions
from app.services.import_export import IMPORT_UPDATE_MODES, analyze_trainee_import_duplicates, parse_document_content
from app.services.schedule_import import parse_schedule_docx
from app.services.storage import detect_document_type, persist_upload
from app.tasks.worker import process_export_job_task, process_import_job_task

router = APIRouter()
logger = get_task_logger(__name__)
MAX_BATCH_IMPORT_FILES = 100
INLINE_IMPORT_ROW_LIMIT = 500
BATCH_IMPORT_FORMATS_CACHE_KEY = "documents:batch_import_formats:v1"
BATCH_IMPORT_FORMATS_CACHE_TTL_SECONDS = 300
IMPORTABLE_DOCUMENT_TYPES = {"xlsx", "docx", "csv"}
IMPORTABLE_DOCUMENT_TYPES_LABEL = ".xls/.xlsx, .csv, .docx"
PDF_IMPORT_UNSUPPORTED_MESSAGE = (
    "PDF не підтримується для автоматичного імпорту. "
    "Для слухачів завантажте .xls/.xlsx або .csv, для розкладу - .docx."
)


def _dispatch_with_fallback(task, job_id: int, job_kind: str, *, allow_inline: bool = True) -> str:
    """
    Try queue-first dispatch. If broker is unavailable (common in serverless),
    run synchronously in-process to avoid API 500 on import/export actions.
    """
    try:
        task.delay(job_id)
        return "queued"
    except Exception as queue_exc:
        logger.warning("Queue dispatch failed for %s job %s: %s", job_kind, job_id, queue_exc)
        if not allow_inline:
            return "queue_unavailable"
        try:
            task.run(job_id)
            return "inline"
        except Exception as inline_exc:
            logger.exception("Inline execution failed for %s job %s: %s", job_kind, job_id, inline_exc)
            return "inline_failed"


def _with_dispatch_notice(job: JobResponse, dispatch_mode: str) -> JobResponse:
    if dispatch_mode == "inline":
        suffix = f" {job.message}" if job.message else ""
        job.message = f"Черга тимчасово недоступна. Операцію виконано одразу в API.{suffix}"
    elif dispatch_mode == "inline_failed":
        suffix = f" {job.message}" if job.message else ""
        job.message = f"Черга недоступна, а inline-виконання завершилось помилкою.{suffix}"
    elif dispatch_mode == "queue_unavailable":
        suffix = f" {job.message}" if job.message else ""
        job.message = f"Черга тимчасово недоступна. Великий імпорт залишено у статусі queued для фонової обробки.{suffix}"
    return job


def _preview_rows(rows: list[dict], limit: int = 10) -> list[dict[str, str]]:
    preview: list[dict[str, str]] = []
    for row in rows[:limit]:
        preview.append({str(key): "" if value is None else str(value) for key, value in row.items()})
    return preview


def _write_upload_to_temp(file: UploadFile) -> str:
    suffix = f".{file.filename.rsplit('.', 1)[1].lower()}" if file.filename and "." in file.filename else ""
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        return tmp.name


def _display_upload_name(file: UploadFile, relative_path: str | None = None) -> str:
    raw_name = (relative_path or file.filename or "uploaded_file").replace("\\", "/").strip()
    parts = [part for part in raw_name.split("/") if part and part not in {".", ".."}]
    safe_name = " / ".join(parts) or file.filename or "uploaded_file"
    return safe_name[-255:]


def _ensure_importable_document_type(doc_type) -> None:
    if doc_type.value == "pdf":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=PDF_IMPORT_UNSUPPORTED_MESSAGE)
    if doc_type.value not in IMPORTABLE_DOCUMENT_TYPES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Підтримуються {IMPORTABLE_DOCUMENT_TYPES_LABEL}")


def _is_large_tabular_import(path: str, doc_type) -> bool:
    if doc_type.value == "xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return any((sheet.max_row or 0) > INLINE_IMPORT_ROW_LIMIT for sheet in workbook.worksheets)
        finally:
            workbook.close()
    if doc_type.value == "csv":
        try:
            with Path(path).open("r", encoding="utf-8-sig", errors="ignore", newline="") as handle:
                for index, _line in enumerate(handle, start=1):
                    if index > INLINE_IMPORT_ROW_LIMIT:
                        return True
        except OSError:
            return False
    return False


def _create_import_job(
    db: DbSession,
    current_user: CurrentUser,
    file: UploadFile,
    *,
    update_existing_mode: str,
    idempotency_key: str,
    source: str,
    message: str,
    request_payload: dict | None = None,
    relative_path: str | None = None,
) -> ImportJob:
    doc_type = detect_document_type(file.filename)
    _ensure_importable_document_type(doc_type)
    path, sha256 = persist_upload(file)
    document = Document(
        file_name=_display_upload_name(file, relative_path),
        file_path=path,
        file_type=doc_type,
        mime_type=file.content_type,
        hash_sha256=sha256,
        source=source,
        created_by=current_user.id,
        branch_id=current_user.branch_id,
    )
    db.add(document)
    db.flush()

    job = ImportJob(
        branch_id=current_user.branch_id,
        idempotency_key=idempotency_key,
        document_id=document.id,
        status=JobStatus.QUEUED,
        message=message,
        request_payload=request_payload,
        result_payload={"import_mode": update_existing_mode},
    )
    db.add(job)
    db.flush()
    return job


@router.post(
    "/import/preview",
    response_model=ImportPreviewResponse,
    dependencies=[Depends(require_roles(RoleName.ADMIN, RoleName.METHODIST))],
)
def preview_import_document(
    db: DbSession,
    current_user: CurrentUser,
    file: UploadFile = File(...),
) -> ImportPreviewResponse:
    doc_type = detect_document_type(file.filename)
    _ensure_importable_document_type(doc_type)

    temp_path = _write_upload_to_temp(file)
    try:
        if doc_type.value in {"xlsx", "csv"}:
            parsed = parse_document_content(temp_path, doc_type)
            duplicate_analysis = analyze_trainee_import_duplicates(db, parsed, current_user.branch_id)
            warnings: list[str] = []
            if not parsed.get("headers"):
                warnings.append("Не знайдено заголовків таблиці")
            if not parsed.get("rows"):
                warnings.append("Не знайдено рядків для імпорту")
            if duplicate_analysis.get("duplicate_count"):
                warnings.append("Знайдено наявних слухачів. Перед імпортом оберіть дію для дублікатів.")
            return ImportPreviewResponse(
                filename=file.filename or "uploaded_file",
                file_type=doc_type.value,
                import_kind="contracts",
                rows=int(parsed.get("rows") or 0),
                sheet_name=parsed.get("sheet_name"),
                headers=[str(item) for item in parsed.get("headers", [])],
                default_group_code=parsed.get("default_group_code"),
                default_group_name=parsed.get("default_group_name"),
                new_count=int(duplicate_analysis.get("new_count") or 0),
                duplicate_count=int(duplicate_analysis.get("duplicate_count") or 0),
                invalid_count=int(duplicate_analysis.get("invalid_count") or 0),
                duplicate_preview=duplicate_analysis.get("duplicate_preview", []),
                preview=_preview_rows(parsed.get("data", [])),
                warnings=warnings,
            )

        if doc_type.value == "docx":
            try:
                schedules = parse_schedule_docx(temp_path)
            except Exception as exc:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"DOCX не схожий на розклад: {exc}") from exc

            groups: list[ImportPreviewGroup] = []
            for item in schedules:
                entries = item.get("entries") or []
                group_code = str(item.get("group_code") or "")
                existing_group = (
                    db.query(Group)
                    .filter(Group.branch_id == current_user.branch_id, Group.code == group_code)
                    .first()
                )
                existing_lessons = 0
                if existing_group and entries:
                    min_start = min(entry["starts_at"] for entry in entries)
                    max_end = max(entry["ends_at"] for entry in entries)
                    existing_lessons = (
                        db.query(ScheduleSlot)
                        .filter(
                            ScheduleSlot.group_id == existing_group.id,
                            ScheduleSlot.starts_at >= min_start,
                            ScheduleSlot.starts_at <= max_end,
                        )
                        .count()
                    )
                groups.append(
                    ImportPreviewGroup(
                        code=group_code,
                        name=str(item.get("group_name") or ""),
                        start_date=item.get("start_date"),
                        end_date=item.get("end_date"),
                        lessons=len(entries),
                        teachers=len({entry.get("teacher_name") for entry in entries if entry.get("teacher_name")}),
                        subjects=len({entry.get("subject_name") for entry in entries if entry.get("subject_name")}),
                        total_hours=round(float(item.get("group_total_hours") or 0), 2),
                        already_exists=existing_group is not None,
                        existing_lessons=existing_lessons,
                    )
                )
            warnings = [] if groups else ["У документі не знайдено груп для імпорту"]
            if any(group.existing_lessons > 0 for group in groups):
                warnings.append("У вибраному періоді вже є заняття. Перед імпортом оберіть режим оновлення розкладу.")
            return ImportPreviewResponse(
                filename=file.filename or "uploaded_file",
                file_type=doc_type.value,
                import_kind="schedule",
                rows=sum(group.lessons for group in groups),
                groups=groups,
                warnings=warnings,
            )

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Підтримуються {IMPORTABLE_DOCUMENT_TYPES_LABEL}")
    finally:
        try:
            Path(temp_path).unlink(missing_ok=True)
        except OSError:
            pass


@router.post(
    "/import",
    response_model=JobResponse,
    status_code=status.HTTP_202_ACCEPTED,
    dependencies=[Depends(require_roles(RoleName.ADMIN, RoleName.METHODIST))],
)
def import_document(
    db: DbSession,
    current_user: CurrentUser,
    file: UploadFile = File(...),
    update_existing_mode: str = Form(default="skip_existing"),
    x_idempotency_key: str | None = Header(default=None),
) -> JobResponse:
    doc_type = detect_document_type(file.filename)
    _ensure_importable_document_type(doc_type)
    if update_existing_mode not in IMPORT_UPDATE_MODES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Некоректний режим імпорту")

    raw_idem_key = x_idempotency_key or f"import-{uuid4().hex}"
    idem_key = f"{current_user.branch_id}:{raw_idem_key}"
    existing = (
        apply_branch_scope(db.query(ImportJob), ImportJob, current_user.branch_id)
        .filter(ImportJob.idempotency_key == idem_key)
        .first()
    )
    if existing:
        return JobResponse.model_validate(existing)

    job = _create_import_job(
        db,
        current_user,
        file,
        update_existing_mode=update_existing_mode,
        idempotency_key=idem_key,
        source="upload",
        message="Заявку на імпорт створено",
    )
    db.commit()
    db.refresh(job)

    allow_inline = not _is_large_tabular_import(job.document.file_path, job.document.file_type) if job.document else True
    dispatch_mode = _dispatch_with_fallback(process_import_job_task, job.id, "import", allow_inline=allow_inline)
    db.refresh(job)
    write_audit(
        db,
        actor_user_id=current_user.id,
        action="documents.import.create_job",
        entity_type="import_job",
        entity_id=str(job.id),
        details={
            "document_id": job.document_id,
            "file_name": job.document.file_name if job.document else file.filename,
            "dispatch_mode": dispatch_mode,
            "import_mode": update_existing_mode,
        },
    )
    return _with_dispatch_notice(JobResponse.model_validate(job), dispatch_mode)


@router.get(
    "/import/batch/formats",
    response_model=BatchImportFormatsResponse,
    dependencies=[Depends(require_roles(RoleName.ADMIN, RoleName.METHODIST))],
)
def get_batch_import_formats() -> BatchImportFormatsResponse:
    cached = cache_get_json(BATCH_IMPORT_FORMATS_CACHE_KEY)
    if isinstance(cached, dict):
        return BatchImportFormatsResponse.model_validate(cached)

    response = BatchImportFormatsResponse(supported_extensions=supported_batch_import_extensions())
    cache_set_json(
        BATCH_IMPORT_FORMATS_CACHE_KEY,
        response.model_dump(mode="json"),
        BATCH_IMPORT_FORMATS_CACHE_TTL_SECONDS,
    )
    return response


@router.post(
    "/import/batch",
    response_model=BatchImportResponse,
    status_code=status.HTTP_202_ACCEPTED,
    dependencies=[Depends(require_roles(RoleName.ADMIN, RoleName.METHODIST))],
)
def import_document_batch(
    db: DbSession,
    current_user: CurrentUser,
    files: list[UploadFile] = File(...),
    update_existing_mode: str = Form(default="skip_existing"),
    relative_paths: list[str] | None = Form(default=None),
) -> BatchImportResponse:
    if update_existing_mode not in IMPORT_UPDATE_MODES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Некоректний режим імпорту")
    if not files:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Оберіть папку або кілька файлів для імпорту")
    if len(files) > MAX_BATCH_IMPORT_FILES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"За один пакет можна імпортувати не більше {MAX_BATCH_IMPORT_FILES} файлів",
        )

    batch_id = uuid4().hex
    paths = relative_paths or []
    skipped_files: list[BatchImportSkippedFile] = []
    queued_jobs: list[ImportJob] = []
    supported_extensions = supported_batch_import_extensions()

    for index, file in enumerate(files):
        relative_path = paths[index] if index < len(paths) else None
        display_name = _display_upload_name(file, relative_path)
        import_format = get_batch_import_format(file.filename)
        if not import_format:
            skipped_files.append(
                BatchImportSkippedFile(
                    filename=display_name,
                    reason=f"Непідтримуваний формат. Підтримуються: {', '.join(supported_extensions)}",
                )
            )
            continue

        job = _create_import_job(
            db,
            current_user,
            file,
            update_existing_mode=update_existing_mode,
            idempotency_key=f"{current_user.branch_id}:batch-{batch_id}-{index}",
            source="batch_upload",
            message=f"Заявку на пакетний імпорт створено: {display_name}",
            request_payload={
                "batch_id": batch_id,
                "batch_index": index,
                "batch_total": len(files),
                "relative_path": relative_path,
                "import_kind": import_format.import_kind,
                "import_mode": update_existing_mode,
            },
            relative_path=relative_path,
        )
        queued_jobs.append(job)

    if not queued_jobs:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"У вибраній папці не знайдено файлів підтримуваних форматів: {', '.join(supported_extensions)}",
        )

    db.commit()

    job_responses: list[JobResponse] = []
    dispatch_modes: dict[str, int] = {}
    for job in queued_jobs:
        allow_inline = not _is_large_tabular_import(job.document.file_path, job.document.file_type) if job.document else True
        dispatch_mode = _dispatch_with_fallback(process_import_job_task, job.id, "batch_import", allow_inline=allow_inline)
        dispatch_modes[dispatch_mode] = dispatch_modes.get(dispatch_mode, 0) + 1
        db.refresh(job)
        job_responses.append(_with_dispatch_notice(JobResponse.model_validate(job), dispatch_mode))

    write_audit(
        db,
        actor_user_id=current_user.id,
        action="documents.import.batch_create_jobs",
        entity_type="import_batch",
        entity_id=batch_id,
        details={
            "accepted_count": len(queued_jobs),
            "skipped_count": len(skipped_files),
            "job_ids": [job.id for job in queued_jobs],
            "dispatch_modes": dispatch_modes,
            "import_mode": update_existing_mode,
        },
    )
    return BatchImportResponse(
        batch_id=batch_id,
        total_files=len(files),
        accepted_count=len(queued_jobs),
        skipped_count=len(skipped_files),
        supported_extensions=supported_extensions,
        jobs=job_responses,
        skipped_files=skipped_files,
    )


@router.post(
    "/export",
    response_model=JobResponse,
    status_code=status.HTTP_202_ACCEPTED,
    dependencies=[Depends(require_roles(RoleName.ADMIN, RoleName.METHODIST))],
)
def export_report(
    payload: ExportRequest,
    db: DbSession,
    current_user: CurrentUser,
    x_idempotency_key: str | None = Header(default=None),
) -> JobResponse:
    raw_idem_key = x_idempotency_key or f"export-{payload.report_type}-{payload.export_format}-{uuid4().hex}"
    idem_key = f"{current_user.branch_id}:{raw_idem_key}"
    existing = apply_branch_scope(db.query(ExportJob), ExportJob, current_user.branch_id).filter(ExportJob.idempotency_key == idem_key).first()
    if existing:
        return JobResponse.model_validate(existing)

    job = ExportJob(
        idempotency_key=idem_key,
        report_type=payload.report_type,
        export_format=payload.export_format,
        branch_id=current_user.branch_id,
        status=JobStatus.QUEUED,
        message="Заявку на експорт створено",
        request_payload={
            "teacher_ids": payload.teacher_ids,
            "start_date": payload.start_date.isoformat() if payload.start_date else None,
            "end_date": payload.end_date.isoformat() if payload.end_date else None,
        },
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    dispatch_mode = _dispatch_with_fallback(process_export_job_task, job.id, "export")
    db.refresh(job)
    write_audit(
        db,
        actor_user_id=current_user.id,
        action="documents.export.create_job",
        entity_type="export_job",
        entity_id=str(job.id),
        details={"report_type": payload.report_type, "format": payload.export_format, "dispatch_mode": dispatch_mode},
    )
    return _with_dispatch_notice(JobResponse.model_validate(job), dispatch_mode)


@router.get("/{document_id}/download")
def download_document(document_id: int, db: DbSession, current_user: CurrentUser) -> FileResponse:
    document = db.get(Document, document_id)
    if not document:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Документ не знайдено")
    ensure_same_branch(current_user, document, "Документ")
    if not Path(document.file_path).exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл документа відсутній у сховищі")
    return FileResponse(
        path=document.file_path,
        filename=document.file_name,
        media_type=document.mime_type or "application/octet-stream",
    )
