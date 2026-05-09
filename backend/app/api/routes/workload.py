from datetime import date, datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.api.deps import CurrentUser, DbSession, require_roles
from app.models import RoleName, ScheduleSlot, Teacher
from app.schemas.api import TeacherMergeRequest, TeacherMergeResponse, WorkloadResponse
from app.services.audit import write_audit
from app.services.import_export import collect_teacher_workload_summary

router = APIRouter()


@router.get("", response_model=list[WorkloadResponse])
def get_workload(
    db: DbSession,
    current_user: CurrentUser,
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
) -> list[WorkloadResponse]:
    summary = collect_teacher_workload_summary(db, current_user.branch_id, date_from=date_from, date_to=date_to)
    return [
        WorkloadResponse(
            teacher_id=row["teacher_id"],
            row_number=row["row_number"],
            teacher_name=row["teacher_name"],
            total_hours=row["total_hours"],
            annual_load_hours=row["annual_load_hours"],
            remaining_hours=row["remaining_hours"],
            groups=row.get("groups", []),
        )
        for row in summary
    ]


@router.get("/export-summary")
def export_workload_summary(
    db: DbSession,
    current_user: CurrentUser,
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
) -> StreamingResponse:
    rows = collect_teacher_workload_summary(db, current_user.branch_id, date_from=date_from, date_to=date_to)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Педнавантаження"
    generated_at = datetime.now(ZoneInfo("Europe/Kyiv"))
    sheet.merge_cells("A1:D1")
    sheet["A1"] = f"Дата формування: {generated_at.strftime('%d.%m.%Y %H:%M')}"
    sheet["A1"].font = Font(bold=True, color="1F3349")
    sheet["A1"].alignment = Alignment(horizontal="left")
    headers = ["Викладач", "Поточні години", "Річний план", "Залишок годин"]
    sheet.append(headers)
    sheet.insert_rows(2)

    for row in rows:
        sheet.append(
            [
                row["teacher_name"],
                row["total_hours"],
                row["annual_load_hours"],
                row["remaining_hours"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="E8F1F4")
    for cell in sheet[3]:
        cell.font = Font(bold=True, color="1F3349")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=4, min_col=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")
        remaining_cell = row[2]
        if isinstance(remaining_cell.value, (int, float)) and remaining_cell.value < 0:
            remaining_cell.font = Font(bold=True, color="DC2626")
        else:
            remaining_cell.font = Font(bold=True, color="0F5132")

    widths = {
        "A": 42,
        "B": 16,
        "C": 16,
        "D": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A4"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    if rows:
        table = Table(displayName="TeacherWorkloadSummary", ref=f"A3:D{len(rows) + 3}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    date_from_part = date_from.isoformat() if date_from else "all"
    date_to_part = date_to.isoformat() if date_to else "all"
    filename = f"teacher_workload_summary_{date_from_part}_{date_to_part}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/merge-teachers",
    response_model=TeacherMergeResponse,
    dependencies=[Depends(require_roles(RoleName.ADMIN, RoleName.METHODIST))],
)
def merge_teachers(
    payload: TeacherMergeRequest,
    db: DbSession,
    current_user: CurrentUser,
) -> TeacherMergeResponse:
    source_ids = sorted({teacher_id for teacher_id in payload.source_teacher_ids if teacher_id != payload.target_teacher_id})
    if not source_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Оберіть щонайменше одного викладача для об'єднання")

    target = (
        db.query(Teacher)
        .filter(Teacher.id == payload.target_teacher_id, Teacher.branch_id == current_user.branch_id)
        .first()
    )
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Основного викладача не знайдено")

    sources = (
        db.query(Teacher)
        .filter(Teacher.branch_id == current_user.branch_id, Teacher.id.in_(source_ids))
        .all()
    )
    if len(sources) != len(source_ids):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Одного з викладачів для об'єднання не знайдено")

    corrected_last_name = payload.last_name.strip() if payload.last_name is not None else None
    corrected_first_name = payload.first_name.strip() if payload.first_name is not None else None
    if payload.last_name is not None and not corrected_last_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Прізвище викладача не може бути порожнім")
    if payload.first_name is not None and not corrected_first_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ім'я та по батькові викладача не можуть бути порожніми")

    reassigned_slots = (
        db.query(ScheduleSlot)
        .filter(ScheduleSlot.teacher_id.in_(source_ids))
        .update({ScheduleSlot.teacher_id: target.id}, synchronize_session=False)
    )
    target.annual_load_hours = float(target.annual_load_hours or 0) + sum(float(source.annual_load_hours or 0) for source in sources)
    if not target.hourly_rate:
        target.hourly_rate = max(float(source.hourly_rate or 0) for source in sources + [target])
    if corrected_last_name is not None:
        target.last_name = corrected_last_name
    if corrected_first_name is not None:
        target.first_name = corrected_first_name
    db.add(target)

    merged_names = [f"{source.last_name} {source.first_name}".strip() for source in sources]
    for source in sources:
        db.delete(source)

    db.commit()
    db.refresh(target)

    write_audit(
        db,
        actor_user_id=current_user.id,
        action="teacher.merge",
        entity_type="teacher",
        entity_id=str(target.id),
        details={
            "merged_teacher_ids": source_ids,
            "merged_teacher_names": merged_names,
            "reassigned_slots": reassigned_slots,
            "annual_load_hours": target.annual_load_hours,
            "final_teacher_name": f"{target.last_name} {target.first_name}".strip(),
        },
    )
    return TeacherMergeResponse(
        target_teacher_id=target.id,
        teacher_name=f"{target.last_name} {target.first_name}".strip(),
        merged_teacher_ids=source_ids,
        reassigned_slots=reassigned_slots,
        annual_load_hours=target.annual_load_hours,
    )
