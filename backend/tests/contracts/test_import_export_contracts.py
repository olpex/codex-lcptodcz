import csv
import json
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path

from docx import Document as DocxDocument
from fpdf import FPDF
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader

from app.models import DocumentType, Group, Room, ScheduleSlot, Subject, Teacher, Trainee
from app.services.import_export import collect_report_rows, parse_document_content, save_report_file


FIXTURES_DIR = Path(__file__).resolve().parents[1] / "fixtures" / "contracts"


def _read_fixture(name: str) -> dict:
    with (FIXTURES_DIR / name).open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def test_import_contract_xlsx(tmp_path: Path):
    golden = _read_fixture("import_xlsx.golden.json")

    file_path = tmp_path / "import.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(golden["headers"])
    for row in golden["data"]:
        sheet.append([row["first_name"], row["last_name"], row["status"]])
    workbook.save(file_path)

    parsed = parse_document_content(str(file_path), DocumentType.XLSX)
    assert parsed["rows"] == golden["rows"]
    assert parsed["headers"] == golden["headers"]
    assert parsed["data"] == golden["data"]


def test_import_contract_docx(tmp_path: Path):
    golden = _read_fixture("import_docx.golden.json")

    file_path = tmp_path / "import.docx"
    document = DocxDocument()
    for line in golden["text_preview"].split("\n"):
        document.add_paragraph(line)
    document.save(file_path)

    parsed = parse_document_content(str(file_path), DocumentType.DOCX)
    assert parsed["rows"] == 1
    assert _normalize_text(parsed["text_preview"]) == _normalize_text(golden["text_preview"])


def test_import_contract_pdf(tmp_path: Path):
    golden = _read_fixture("import_pdf.golden.json")

    file_path = tmp_path / "import.pdf"
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)
    for line in golden["text_contains"]:
        pdf.cell(0, 8, text=line, new_x="LMARGIN", new_y="NEXT")
    pdf.output(str(file_path))

    parsed = parse_document_content(str(file_path), DocumentType.PDF)
    preview = _normalize_text(parsed["text_preview"])
    for expected in golden["text_contains"]:
        assert _normalize_text(expected) in preview


def test_export_contract_structures(db_session):
    golden = _read_fixture("export_structure.golden.json")
    db_session.add(Trainee(first_name="Олена", last_name="Коваль", status="active"))

    teacher = Teacher(first_name="Марія Іванівна", last_name="Бондар", annual_load_hours=120)
    group = Group(code="EXP-001", name="Експортна група", status="active")
    subject = Subject(name="Тестовий предмет", hours_total=20)
    room = Room(name="Тестова аудиторія", capacity=20)
    db_session.add_all([teacher, group, subject, room])
    db_session.flush()
    starts_at = datetime(2026, 4, 1, 9, 30, tzinfo=timezone.utc)
    db_session.add(
        ScheduleSlot(
            group_id=group.id,
            teacher_id=teacher.id,
            subject_id=subject.id,
            room_id=room.id,
            starts_at=starts_at,
            ends_at=starts_at + timedelta(minutes=95),
            academic_hours=2.0,
            pair_number=1,
        )
    )
    db_session.commit()

    for report_type, expected_headers in golden.items():
        rows = collect_report_rows(db_session, report_type, "main")
        assert rows, f"empty rows for {report_type}"

        csv_path, csv_type = save_report_file(rows, report_type, "csv")
        assert csv_type == DocumentType.CSV
        with Path(csv_path).open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            assert reader.fieldnames == expected_headers

        xlsx_path, xlsx_type = save_report_file(rows, report_type, "xlsx")
        assert xlsx_type == DocumentType.XLSX
        workbook = load_workbook(xlsx_path, data_only=True)
        sheet = workbook.active
        xlsx_headers = [cell.value for cell in sheet[1]]
        assert xlsx_headers == expected_headers

        pdf_path, pdf_type = save_report_file(rows, report_type, "pdf")
        assert pdf_type == DocumentType.PDF
        pdf_text = _normalize_text("\n".join((page.extract_text() or "") for page in PdfReader(pdf_path).pages))
        assert f"Звіт: " in pdf_text
        assert "????" not in pdf_text
        if report_type == "teacher_workload":
            assert "Бондар Марія Іванівна" in pdf_text
