from datetime import datetime, timezone
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.core.crypto import cipher
from app.models import Group, GroupStatus, JournalWorkloadEntry, Room, ScheduleSlot, Subject, Teacher, Trainee
from app.services.import_export import collect_teacher_workload_summary
from app.services import journal_monitor


def _seed_schedule(db_session, group: Group, suffix: str) -> None:
    teacher = Teacher(branch_id="main", first_name="Тест", last_name="Викладач", hourly_rate=0, is_active=True)
    subject = Subject(branch_id="main", name=f"Предмет {suffix}", hours_total=4)
    room = Room(branch_id="main", name=f"Аудиторія {suffix}", capacity=20)
    db_session.add_all([teacher, subject, room])
    db_session.flush()
    db_session.add(
        ScheduleSlot(
            group_id=group.id,
            teacher_id=teacher.id,
            subject_id=subject.id,
            room_id=room.id,
            starts_at=datetime(2026, 5, 1, 9, 30, tzinfo=timezone.utc),
            ends_at=datetime(2026, 5, 1, 11, 5, tzinfo=timezone.utc),
            pair_number=1,
            academic_hours=2,
        )
    )


def _journal_workbook_bytes(
    rows: list[tuple[str, float, str, str]],
    *,
    sheet_name: str = "Дисципліни",
    discipline_sheet_index: int = 2,
) -> bytes:
    workbook = Workbook()
    workbook.active.title = "Загальні"
    while len(workbook.worksheets) < discipline_sheet_index:
        workbook.create_sheet(f"Аркуш {len(workbook.worksheets) + 1}")
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(["Назва дисципліни", "Кількість годин", "Сторінки", "Прізвище, ім'я, по батькові викладача"])
    for row in rows:
        sheet.append(list(row))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _journal_workbook_with_total_row_bytes(rows: list[tuple[str, float, str, str]], total_hours: float) -> bytes:
    workbook = Workbook()
    workbook.active.title = "Загальні"
    sheet = workbook.create_sheet("Дисципліни")
    sheet.append(["№", "назва предмету", "годин", "сторінка", "прізвище, ім'я, по батькові викладача"])
    for index, row in enumerate(rows, start=1):
        subject, hours, pages, teacher = row
        sheet.append([index, subject, hours, pages, teacher])
    sheet.append(["", "Загальний обсяг навчального часу", total_hours, "", ""])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _journal_workbook_without_subject_column_bytes(rows: list[tuple[float, str, str]]) -> bytes:
    workbook = Workbook()
    workbook.active.title = "Загальні"
    sheet = workbook.create_sheet("Дисципліни")
    sheet.append(["Кількість годин", "Сторінки", "Прізвище, ім'я, по батькові викладача"])
    for row in rows:
        sheet.append(list(row))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _journal_zv_workbook_bytes(rows: list[tuple[int, str, str, str, str, str, str, str]]) -> bytes:
    workbook = Workbook()
    workbook.active.title = "Дисципліни"
    sheet = workbook.create_sheet("ЗВ")
    sheet.append(
        [
            "Номер за порядком",
            "Номер в журналі З-СНН",
            "Прізвище, ім'я та по батькові слухача",
            "Стать",
            "Дата народження",
            "Ідентифікаційний номер",
            "Домашня адреса",
            "Телефон",
        ]
    )
    for row in rows:
        sheet.append(list(row))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _journal_zv_workbook_with_title_bytes(rows: list[tuple[int, str, str, str, str, str, str]]) -> bytes:
    workbook = Workbook()
    workbook.active.title = "ЗВ"
    sheet = workbook.active
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "ЗАГАЛЬНІ ВІДОМОСТІ ПРО СЛУХАЧІВ"
    sheet.append(["№ п/п", "№ договору", "Прізвище, ім'я, по батькові", "Стать", "Дата народження", "ІПН", "Адреса"])
    for row in rows:
        sheet.append(list(row))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@pytest.fixture(autouse=True)
def _default_drive_workbook_lister(monkeypatch):
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {
                "id": f"{folder_id}-xlsx",
                "name": f"{folder_id}.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "modifiedTime": "2026-05-01T10:00:00Z",
            }
        ],
        raising=False,
    )


def _journal_zv_workbook_with_combined_address_phone_bytes(rows: list[tuple[int, str, str, str, str, str, str]]) -> bytes:
    workbook = Workbook()
    workbook.active.title = "ЗВ"
    sheet = workbook.active
    sheet.append(["№ п/п", "№ договору", "Прізвище, ім'я, по батькові", "Стать", "Дата народження", "ІПН", "Адреса, телефон"])
    for row in rows:
        sheet.append(list(row))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _journal_combined_workbook_bytes() -> bytes:
    workbook = Workbook()
    workbook.active.title = "Загальні"
    disciplines = workbook.create_sheet("Дисципліни")
    disciplines.append(["Назва дисципліни", "Кількість годин", "Сторінки", "Прізвище, ім'я, по батькові викладача"])
    disciplines.append(["Основи безпеки", 8, "1-8", "Коваль Олена Петрівна"])
    zv = workbook.create_sheet("ЗВ")
    zv.append(
        [
            "Номер за порядком",
            "Номер в журналі З-СНН",
            "Прізвище, ім'я та по батькові слухача",
            "Стать",
            "Дата народження",
            "Ідентифікаційний номер",
            "Домашня адреса",
            "Телефон",
        ]
    )
    zv.append((1, "З-СНН-001", "Петренко Іван Іванович", "ч", "01.02.1990", "1234567890", "м. Львів", "+380501112233"))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_journal_zv_parser_skips_title_and_imports_all_real_trainees():
    surnames = [
        "Андрущенко",
        "Бойко",
        "Василенко",
        "Гнатюк",
        "Данилюк",
        "Єфименко",
        "Жук",
        "Захаренко",
        "Іванчук",
        "Климчук",
        "Левченко",
        "Мельник",
        "Назаренко",
        "Онищенко",
        "Петренко",
        "Романюк",
        "Савчук",
        "Ткаченко",
        "Удовенко",
        "Федоренко",
        "Хоменко",
        "Цимбалюк",
        "Чорний",
        "Шевченко",
        "Юрченко",
        "Яценко",
        "Коваль",
        "Марченко",
        "Лисенко",
        "Мороз",
        "Поліщук",
        "Руденко",
        "Семенюк",
    ]
    rows = [
        (
            index,
            "1(З-СНН) від 02.01.2026",
            f"{surname} Наталія Іванівна",
            "ж" if index % 2 else "ч",
            "02.01.1990",
            f"10000000{index:02d}",
            f"Адреса {index}",
        )
        for index, surname in enumerate(surnames, start=1)
    ]

    parsed = journal_monitor.parse_journal_zv_trainees_xlsx(
        _journal_zv_workbook_with_title_bytes(rows),
        group_code="1-26",
        group_name="1-26 Організація трудових відносин",
    )

    assert parsed["rows"] == 33
    assert parsed["data"][0]["Прізвище"] == "Андрущенко"
    assert parsed["data"][0]["Ідентифікаційний номер"] == "1000000001"
    assert parsed["data"][0]["№ договору"] == "1(З-СНН) від 02.01.2026"
    assert parsed["data"][-1]["Номер за порядком"] == 33
    assert all(row["Прізвище"] != "№" for row in parsed["data"])


def test_journal_zv_parser_does_not_copy_address_into_phone_from_combined_column():
    parsed = journal_monitor.parse_journal_zv_trainees_xlsx(
        _journal_zv_workbook_with_combined_address_phone_bytes(
            [
                (1, "З-СНН-001", "Петренко Іван Іванович", "ч", "01.02.1990", "1234567890", "м. Львів, вул. Зелена 1"),
                (2, "З-СНН-002", "Коваль Олена Петрівна", "ж", "03.04.1992", "0987654321", "м. Київ, вул. Хрещатик 1, +380501112233"),
            ]
        ),
        group_code="46-26",
        group_name="46-26 Журнал",
    )

    assert parsed["data"][0]["Домашня адреса"] == "м. Львів, вул. Зелена 1"
    assert parsed["data"][0]["Телефон"] is None
    assert parsed["data"][1]["Домашня адреса"] == "м. Київ, вул. Хрещатик 1"
    assert parsed["data"][1]["Телефон"] == "+380501112233"


def test_journal_zv_parser_extracts_ukrainian_mobile_from_address_cell():
    parsed = journal_monitor.parse_journal_zv_trainees_xlsx(
        _journal_zv_workbook_with_combined_address_phone_bytes(
            [
                (
                    1,
                    "З-СНН-001",
                    "Петренко Іван Іванович",
                    "ч",
                    "01.02.1990",
                    "1234567890",
                    "Львівська область, Жовтанці, Центральна 7, 097 831 94 50",
                ),
                (
                    2,
                    "З-СНН-002",
                    "Коваль Олена Петрівна",
                    "ж",
                    "03.04.1992",
                    "0987654321",
                    "м. Київ, вул. Хрещатик 1, 067.222.33.44",
                ),
                (
                    3,
                    "З-СНН-003",
                    "Сидоренко Марія Іванівна",
                    "ж",
                    "05.06.1994",
                    "1111111111",
                    "м. Луцьк, +38 (050) 111-22-33",
                ),
            ]
        ),
        group_code="46-26",
        group_name="46-26 Журнал",
    )

    assert parsed["data"][0]["Домашня адреса"] == "Львівська область, Жовтанці, Центральна 7"
    assert parsed["data"][0]["Телефон"] == "+380978319450"
    assert parsed["data"][1]["Домашня адреса"] == "м. Київ, вул. Хрещатик 1"
    assert parsed["data"][1]["Телефон"] == "+380672223344"
    assert parsed["data"][2]["Домашня адреса"] == "м. Луцьк"
    assert parsed["data"][2]["Телефон"] == "+380501112233"


def test_journal_worker_imports_trainees_from_zv_sheet_and_updates_group_status(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    drive_folders = lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-46-26",
                "name": "46-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-46-26",
                "modified_time": "2026-03-02T10:00:00Z",
            },
        ]
    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", drive_folders)
    monkeypatch.setattr("app.tasks.worker.list_drive_child_folders", drive_folders)
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": "zv-xlsx", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes(
            [
                (1, "З-СНН-001", "Петренко Іван Іванович", "ч", "01.02.1990", "1234567890", "м. Львів, вул. Зелена 1", "+380501112233"),
                (2, "З-СНН-002", "Коваль Олена Петрівна", "ж", "03.04.1992", "0987654321", "м. Львів, вул. Шевченка 2", "+380672223344"),
            ]
        ),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    section = db_session.get(journal_monitor.JournalMonitorSection, section_id)
    section.workload_auto_enabled = True
    section.workload_auto_year = 2026
    db_session.add(section)
    db_session.commit()

    from app.tasks.worker import process_journal_monitor_auto_task

    process_journal_monitor_auto_task.run()
    detail_response = client.get(f"/api/v1/journal-monitors/{section_id}", headers=auth_headers)

    assert detail_response.status_code == 200
    entry = detail_response.json()["entries"][0]
    assert entry["group_code"] == "46-26"
    assert entry["has_trainees"] is True
    assert entry["trainee_count"] == 2
    assert entry["processing_status"] == "trainees_only"
    assert entry["trainees_status"] == "processed"
    assert entry["trainees_message"] == "Додано/оновлено слухачів із журналу: 2"

    trainees = db_session.query(Trainee).order_by(Trainee.last_name).all()
    assert [(trainee.last_name, trainee.first_name, trainee.group_code) for trainee in trainees] == [
        ("Коваль", "Олена Петрівна", "46-26"),
        ("Петренко", "Іван Іванович", "46-26"),
    ]
    assert cipher.decrypt(trainees[1].tax_id_encrypted) == "1234567890"
    assert cipher.decrypt(trainees[1].address_encrypted) == "м. Львів, вул. Зелена 1"
    assert cipher.decrypt(trainees[1].phone_encrypted) == "+380501112233"


def test_empty_zv_marks_journal_without_trainees_even_when_group_had_existing_rows(
    db_session,
    monkeypatch,
):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    group = Group(branch_id="main", code="46-26", name="Група 46-26", status=GroupStatus.ACTIVE)
    trainee = Trainee(
        branch_id="main",
        first_name="Іван Іванович",
        last_name="Петренко",
        status="active",
        group_code="46-26",
    )
    db_session.add_all([section, group, trainee])
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-46-26",
        journal_name="46-26 Журнал",
        group_code="46-26",
        trainees_status="pending",
    )
    db_session.add(entry)
    db_session.commit()

    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": "zv-empty", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes(
            [
                (1, "З-СНН-001", "", "", "", "", "", ""),
                (2, "З-СНН-002", "", "", "", "", "", ""),
            ]
        ),
        raising=False,
    )

    result = journal_monitor.process_journal_trainees_for_section(db_session, section, limit=1, target_year=2026)
    groups_by_code, schedule_counts, trainee_counts = journal_monitor._group_maps(db_session, "main")
    journal_monitor._refresh_entry_project_state(db_session, entry, groups_by_code, schedule_counts, trainee_counts)
    db_session.commit()

    assert result["no_data"] == 1
    assert entry.trainees_status == "no_data"
    assert entry.has_trainees is False
    assert entry.trainee_count == 0
    assert trainee.is_deleted is True


def test_journal_monitor_uses_zv_row_count_not_registry_group_count(db_session):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    group = Group(branch_id="main", code="47-26", name="Група 47-26", status=GroupStatus.ACTIVE)
    db_session.add_all(
        [
            section,
            group,
            Trainee(branch_id="main", first_name="Один", last_name="Старий", status="active", group_code="47-26"),
            Trainee(branch_id="main", first_name="Два", last_name="Старий", status="active", group_code="47-26"),
        ]
    )
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-47-26",
        journal_name="47-26 Журнал",
        group_code="47-26",
        trainees_status="processed",
        trainees_message="Додано/оновлено слухачів із журналу: 1",
    )
    db_session.add(entry)
    db_session.commit()

    groups_by_code, schedule_counts, trainee_counts = journal_monitor._group_maps(db_session, "main")
    journal_monitor._refresh_entry_project_state(db_session, entry, groups_by_code, schedule_counts, trainee_counts)

    assert entry.has_trainees is True
    assert entry.trainee_count == 1


def test_drive_sync_removes_journal_when_folder_has_no_workbook(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    group = Group(branch_id="main", code="48-26", name="Група 48-26", status=GroupStatus.ACTIVE)
    trainee = Trainee(
        branch_id="main",
        first_name="Іван Іванович",
        last_name="Петренко",
        status="active",
        group_code="48-26",
    )
    teacher = Teacher(branch_id="main", first_name="Олена Петрівна", last_name="Коваль", annual_load_hours=100, is_active=True)
    db_session.add_all([section, group, trainee, teacher])
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-48-26",
        journal_name="48-26 Журнал",
        group_code="48-26",
        has_trainees=True,
        trainee_count=1,
        workload_status="processed",
        workload_year=2026,
        workload_hours=12,
        trainees_status="processed",
        trainees_message="Додано/оновлено слухачів із журналу: 1",
        trainees_processed_at=datetime(2026, 5, 1, 10, 0, tzinfo=timezone.utc),
        trainees_source_names=["48-26 Журнал"],
    )
    db_session.add(entry)
    db_session.flush()
    db_session.add(
        JournalWorkloadEntry(
            journal_monitor_entry_id=entry.id,
            branch_id="main",
            teacher_id=teacher.id,
            subject_name="Журнальна дисципліна",
            hours=12,
        )
    )
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [],
        raising=False,
    )

    journal_monitor.sync_journal_monitor_section(
        db_session,
        section,
        folder_lister=lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-48-26",
                "name": "48-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-48-26",
                "modified_time": "2026-05-02T10:00:00Z",
            }
        ],
        process_workload=False,
        process_trainees=False,
    )
    db_session.commit()

    db_session.refresh(group)
    db_session.refresh(trainee)
    assert group.hidden_from_registry is True
    assert trainee.is_deleted is True
    assert trainee.group_code is None
    assert db_session.query(journal_monitor.JournalMonitorEntry).count() == 0
    assert db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.teacher_id == teacher.id).count() == 0
    summary = {row["teacher_id"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert summary[teacher.id]["total_hours"] == 0


def test_journal_trainees_import_keeps_all_rows_with_shared_group_contract(
    db_session,
    monkeypatch,
):
    entry = journal_monitor.JournalMonitorEntry(
        section_id=1,
        branch_id="main",
        drive_file_id="drive-1-26",
        drive_url="https://drive.google.com/drive/folders/drive-1-26",
        journal_name="1-26 Організація трудових відносин",
        group_code="1-26",
    )
    db_session.add(entry)
    db_session.flush()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": "shared-contract-zv", "name": "1-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes(
            [
                (1, "1(З-СНН) від 02.01.2026", "Петренко Іван Іванович", "ч", "01.02.1990", "", "м. Львів", ""),
                (2, "1(З-СНН) від 02.01.2026", "Коваль Олена Петрівна", "ж", "03.04.1992", "", "м. Київ", ""),
                (3, "1(З-СНН) від 02.01.2026", "Шевченко Марія Іванівна", "ж", "05.06.1994", "", "м. Одеса", ""),
            ]
        ),
        raising=False,
    )

    result = journal_monitor.process_journal_trainees_entry(db_session, entry)

    assert result["entries"] == 3
    assert result["import_result"]["inserted"] == 3
    assert db_session.query(Trainee).filter(Trainee.group_code == "1-26", Trainee.is_deleted.is_(False)).count() == 3


def test_journal_response_hides_redundant_workbook_names_but_keeps_distinct_files():
    redundant_entry = journal_monitor.JournalMonitorEntry(
        id=1,
        section_id=1,
        branch_id="main",
        drive_file_id="drive-1-26",
        journal_name="1-26 Організація трудових відносин в умовах воєнного стану",
        group_code="1-26",
        workload_source_names=[
            '"1-26 Організація трудових відносин в умовах воєнного стану правові аспекти"',
            "Теорія",
        ],
    )

    payload = journal_monitor.entry_to_response_payload(redundant_entry)

    assert payload["workload_source_names"] == ["Теорія"]


def test_journal_response_hides_source_names_that_repeat_journal_title_variants():
    tractor_entry = journal_monitor.JournalMonitorEntry(
        id=1,
        section_id=1,
        branch_id="main",
        drive_file_id="drive-18p-26",
        journal_name="18п-26 Трактористи кат.А2",
        group_code="18п-26",
        workload_source_names=[
            "18п-26 Трактори A2",
            "18п-26 Трактори A2 виробн",
            "Практичні роботи",
        ],
    )
    long_title_entry = journal_monitor.JournalMonitorEntry(
        id=2,
        section_id=1,
        branch_id="main",
        drive_file_id="drive-2-26",
        journal_name="2-26 Соціально-психологічна адаптація цивільного населення під час роботи в умовах воєнного стану",
        group_code="2-26",
        workload_source_names=[
            '2-26 "Соціально-психологічна адаптація цивільного населення в умовах воєнного стану"',
        ],
    )

    tractor_payload = journal_monitor.entry_to_response_payload(tractor_entry)
    long_title_payload = journal_monitor.entry_to_response_payload(long_title_entry)

    assert tractor_payload["workload_source_names"] == ["Практичні роботи"]
    assert long_title_payload["workload_source_names"] == []


def test_journal_response_includes_processed_workload_teachers_for_status_tooltip():
    first_teacher = Teacher(
        id=11,
        branch_id="main",
        last_name="Брикін",
        first_name="Віктор Євгенович",
        hourly_rate=0,
        is_active=True,
    )
    second_teacher = Teacher(
        id=12,
        branch_id="main",
        last_name="Старожук",
        first_name="Людмила Василівна",
        hourly_rate=0,
        is_active=True,
    )
    entry = journal_monitor.JournalMonitorEntry(
        id=1,
        section_id=1,
        branch_id="main",
        drive_file_id="drive-82-26",
        journal_name="82-26 Журнал",
        group_code="82-26",
        workload_status="processed",
        workload_hours=14,
    )
    entry.workload_entries = [
        JournalWorkloadEntry(teacher_id=first_teacher.id, teacher=first_teacher, subject_name="Теорія", hours=4),
        JournalWorkloadEntry(teacher_id=first_teacher.id, teacher=first_teacher, subject_name="Практика", hours=5),
        JournalWorkloadEntry(teacher_id=second_teacher.id, teacher=second_teacher, subject_name="Практика", hours=5),
    ]

    payload = journal_monitor.entry_to_response_payload(entry)

    assert payload["workload_teachers"] == [
        {"teacher_id": 11, "teacher_name": "Брикін В. Є.", "hours": 9},
        {"teacher_id": 12, "teacher_name": "Старожук Л. В.", "hours": 5},
    ]


def test_background_tick_reprocesses_processed_trainees_when_import_count_is_too_low(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-1-26",
                "name": "1-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-1-26",
                "modified_time": "2026-05-01T10:00:00Z",
            },
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": "shared-contract-zv", "name": "1-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes(
            [
                (1, "1(З-СНН) від 02.01.2026", "Петренко Іван Іванович", "ч", "01.02.1990", "", "м. Львів", ""),
                (2, "1(З-СНН) від 02.01.2026", "Коваль Олена Петрівна", "ж", "03.04.1992", "", "м. Київ", ""),
                (3, "1(З-СНН) від 02.01.2026", "Шевченко Марія Іванівна", "ж", "05.06.1994", "", "м. Одеса", ""),
            ]
        ),
        raising=False,
    )
    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)
    entry_id = sync_response.json()["entries"][0]["id"]
    db_session.add(Trainee(branch_id="main", first_name="Іван Іванович", last_name="Петренко", group_code="1-26"))
    entry = db_session.get(journal_monitor.JournalMonitorEntry, entry_id)
    entry.trainees_status = "processed"
    entry.trainees_message = "Додано/оновлено слухачів із журналу: 3"
    entry.trainee_count = 1
    db_session.add(entry)
    db_session.commit()

    tick_response = client.post(
        f"/api/v1/journal-monitors/{section_id}/processing/background-tick?year=2026&sync=true",
        headers=auth_headers,
    )

    assert tick_response.status_code == 200
    entry_payload = tick_response.json()["entries"][0]
    assert entry_payload["trainees_status"] == "processed"
    assert entry_payload["trainee_count"] == 3


def test_journal_sync_refreshes_folders_without_inline_processing(client, auth_headers, monkeypatch):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-46-26",
                "name": "46-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-46-26",
                "modified_time": "2026-03-02T10:00:00Z",
            },
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "process_journal_trainees_for_section",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("sync must not process trainees inline")),
    )
    monkeypatch.setattr(
        journal_monitor,
        "process_next_journal_workload",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("sync must not process workload inline")),
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]

    sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)

    assert sync_response.status_code == 200
    entry = sync_response.json()["entries"][0]
    assert entry["group_code"] == "46-26"
    assert entry["trainees_status"] == "pending"
    assert entry["workload_status"] == "pending"


def test_journal_worker_marks_trainees_no_data_when_zv_sheet_has_no_rows(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    drive_folders = lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-47-26",
                "name": "47-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-47-26",
                "modified_time": "2026-03-02T10:00:00Z",
            },
        ]
    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", drive_folders)
    monkeypatch.setattr("app.tasks.worker.list_drive_child_folders", drive_folders)
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": "empty-zv-xlsx", "name": "47-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes([]),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    section = db_session.get(journal_monitor.JournalMonitorSection, section_id)
    section.workload_auto_enabled = True
    section.workload_auto_year = 2026
    db_session.add(section)
    db_session.commit()

    from app.tasks.worker import process_journal_monitor_auto_task

    process_journal_monitor_auto_task.run()
    detail_response = client.get(f"/api/v1/journal-monitors/{section_id}", headers=auth_headers)

    assert detail_response.status_code == 200
    entry = detail_response.json()["entries"][0]
    assert entry["has_trainees"] is False
    assert entry["trainee_count"] == 0
    assert entry["processing_status"] == "not_processed"
    assert entry["trainees_status"] == "no_data"
    assert entry["trainees_message"] == "На аркуші «ЗВ» не знайдено рядків зі слухачами"


def test_journal_worker_marks_trainees_no_data_when_zv_has_only_empty_positions(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Journals 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-846-26",
        journal_name="846-26 Journal",
        group_code="846-26",
    )
    db_session.add(entry)
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": "empty-positions-zv", "name": "846-26.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes(
            [(index, "", "", "", "", "", "", "") for index in range(1, 29)]
        ),
        raising=False,
    )

    result = journal_monitor.process_journal_trainees_for_section(db_session, section, limit=1, target_year=2026)
    db_session.commit()

    db_session.refresh(entry)
    assert result == {"processed": 0, "no_data": 1, "failed": 0}
    assert entry.trainees_status == "no_data"
    assert entry.trainee_count == 0
    assert db_session.query(Trainee).filter(Trainee.group_code == "846-26", Trainee.is_deleted.is_(False)).count() == 0


def test_journal_trainees_no_data_archives_previously_imported_group_trainees(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Journals 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-85-26",
        journal_name="85-26 Journal",
        group_code="85-26",
        trainees_status="processed",
        trainees_processed_at=datetime(2026, 5, 1, 10, 0, tzinfo=timezone.utc),
    )
    db_session.add_all(
        [
            entry,
            Trainee(branch_id="main", first_name="Ivan", last_name="Petrenko", status="active", group_code="85-26"),
            Trainee(branch_id="main", first_name="Olena", last_name="Koval", status="active", group_code="85-26"),
        ]
    )
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {
                "id": "empty-zv",
                "name": "85-26.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "modifiedTime": "2026-05-02T10:00:00Z",
            }
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes(
            [(index, "", "", "", "", "", "", "") for index in range(1, 20)]
        ),
        raising=False,
    )

    result = journal_monitor.process_journal_trainees_for_section(db_session, section, limit=1, target_year=2026)
    db_session.commit()

    db_session.refresh(entry)
    assert result == {"processed": 0, "no_data": 1, "failed": 0}
    assert entry.trainees_status == "no_data"
    assert db_session.query(Trainee).filter(Trainee.group_code == "85-26", Trainee.is_deleted.is_(False)).count() == 0
    assert db_session.query(Trainee).filter(Trainee.is_deleted.is_(True)).count() == 2


def test_journal_workload_no_data_clears_previously_imported_hours(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Journals 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_year=2026,
    )
    teacher = Teacher(branch_id="main", first_name="Olena", last_name="Koval", annual_load_hours=100, is_active=True)
    db_session.add_all([section, teacher])
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-846-26",
        journal_name="846-26 Journal",
        group_code="846-26",
        workload_status="processed",
        workload_year=2026,
        workload_processed_at=datetime(2026, 5, 1, 10, 0, tzinfo=timezone.utc),
        workload_hours=12,
    )
    db_session.add(entry)
    db_session.flush()
    db_session.add(
        JournalWorkloadEntry(
            journal_monitor_entry_id=entry.id,
            branch_id="main",
            teacher_id=teacher.id,
            subject_name="Subject",
            hours=12,
        )
    )
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {
                "id": "empty-disciplines",
                "name": "846-26.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "modifiedTime": "2026-05-02T10:00:00Z",
            }
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_workbook_bytes([]),
        raising=False,
    )

    result = journal_monitor.process_next_journal_workload(db_session, section, limit=1, target_year=2026)
    db_session.commit()

    db_session.refresh(entry)
    assert result == {"processed": 0, "failed": 1, "skipped_year": 0}
    assert entry.workload_status == "no_data"
    assert entry.workload_hours == 0
    assert db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).count() == 0
    summary = {row["teacher_id"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert summary[teacher.id]["total_hours"] == 0


def test_journal_workload_processes_hours_and_teacher_when_subject_names_are_blank(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Journals 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-85-26",
        journal_name="85-26 Journal",
        group_code="85-26",
        workload_status="pending",
        workload_year=2026,
    )
    db_session.add(entry)
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_workbook_bytes(
            [
                ("", 4, "3", "Паращук Олег Леонідович"),
                ("", 5, "5", "Паращук Олег Леонідович"),
                ("", 5, "7", "Паращук Олег Леонідович"),
                ("", 5, "9", "Паращук Олег Леонідович"),
                ("", 5, "11", "Паращук Олег Леонідович"),
                ("", 2, "13", "Паращук Олег Леонідович"),
                ("", 2, "15", "Паращук Олег Леонідович"),
                ("", 2, "19", "Паращук Олег Леонідович"),
            ]
        ),
        raising=False,
    )

    result = journal_monitor.process_next_journal_workload(db_session, section, limit=1, target_year=2026)
    db_session.commit()

    db_session.refresh(entry)
    assert result == {"processed": 1, "failed": 0, "skipped_year": 0}
    assert entry.workload_status == "processed"
    assert entry.workload_hours == 30
    workload_rows = db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).all()
    assert [(row.subject_name, row.hours) for row in workload_rows] == [("Без назви предмета", 30)]
    summary = collect_teacher_workload_summary(db_session, "main")
    assert [(row["teacher_name"], row["total_hours"]) for row in summary] == [("Паращук Олег Леонідович", 30)]


def test_journal_workload_processes_hours_and_teacher_when_subject_column_is_missing(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Journals 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-85-26",
        journal_name="85-26 Journal",
        group_code="85-26",
        workload_status="pending",
        workload_year=2026,
    )
    db_session.add(entry)
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_workbook_without_subject_column_bytes(
            [
                (12, "3-8", "Коваль Олена Петрівна"),
                (18, "9-19", "Коваль Олена Петрівна"),
            ]
        ),
        raising=False,
    )

    result = journal_monitor.process_next_journal_workload(db_session, section, limit=1, target_year=2026)
    db_session.commit()

    db_session.refresh(entry)
    assert result == {"processed": 1, "failed": 0, "skipped_year": 0}
    assert entry.workload_status == "processed"
    assert entry.workload_hours == 30
    workload_rows = db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).all()
    assert [(row.subject_name, row.hours) for row in workload_rows] == [("Без назви предмета", 30)]


def test_journal_workload_no_data_keeps_visible_hours_when_teacher_missing(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Journals 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-85-26",
        journal_name="85-26 Journal",
        group_code="85-26",
        workload_status="pending",
        workload_year=2026,
    )
    db_session.add(entry)
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_workbook_bytes(
            [
                ("Охорона праці", 4, "3", ""),
                ("Роль AI у сучасному ринку праці", 5, "5", ""),
                ("Практичне використання AI у повсякденній роботі", 5, "7", ""),
                ("AI у творчих і технічних професіях", 5, "9", ""),
                ("Працевлаштування та кар'єрний ріст завдяки AI", 5, "11", ""),
                ("Етика, ризики та відповідальне", 2, "13", ""),
                ("Фінальне тестування", 2, "15", ""),
                ("Підсумкове заняття", 2, "19", ""),
            ]
        ),
        raising=False,
    )

    result = journal_monitor.process_next_journal_workload(db_session, section, limit=1, target_year=2026)
    db_session.commit()

    db_session.refresh(entry)
    assert result == {"processed": 0, "failed": 1, "skipped_year": 0}
    assert entry.workload_status == "no_data"
    assert entry.workload_hours == 30
    assert "ПІБ викладачів" in (entry.workload_message or "")
    assert db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).count() == 0
    assert collect_teacher_workload_summary(db_session, "main") == []


def test_journal_workload_uses_complete_workbook_when_another_file_has_incomplete_rows(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Journals 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-85-26",
        journal_name="85-26 Journal",
        group_code="85-26",
        workload_status="pending",
        workload_year=2026,
    )
    db_session.add(entry)
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {
                "id": "complete-disciplines",
                "name": "85-26 Журнал.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
            {
                "id": "incomplete-disciplines",
                "name": "85-26 Журнал копія.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
        ],
        raising=False,
    )

    def workbook_bytes(file_id, mime_type=None, service_account_json=None):
        if file_id == "complete-disciplines":
            return _journal_workbook_bytes(
                [
                    ("Охорона праці", 4, "3", "Паращук Олег Леонідович"),
                    ("Роль AI у сучасному ринку праці", 5, "5", "Паращук Олег Леонідович"),
                    ("Практичне використання AI у повсякденній роботі", 5, "7", "Паращук Олег Леонідович"),
                    ("AI у творчих і технічних професіях", 5, "9", "Паращук Олег Леонідович"),
                    ("Працевлаштування та кар'єрний ріст завдяки AI", 5, "11", "Паращук Олег Леонідович"),
                    ("Етика, ризики та відповідальне", 2, "13", "Паращук Олег Леонідович"),
                    ("Фінальне тестування", 2, "15", "Паращук Олег Леонідович"),
                    ("Підсумкове заняття", 2, "19", "Паращук Олег Леонідович"),
                ]
            )
        return _journal_workbook_bytes(
            [
                ("", 4, "3", ""),
                ("", 5, "5", ""),
                ("", 5, "7", ""),
                ("", 5, "9", ""),
                ("", 5, "11", ""),
                ("", 2, "13", ""),
                ("", 2, "15", ""),
                ("", 2, "19", ""),
            ]
        )

    monkeypatch.setattr(journal_monitor, "download_drive_file_bytes", workbook_bytes, raising=False)

    result = journal_monitor.process_next_journal_workload(db_session, section, limit=1, target_year=2026)
    db_session.commit()

    db_session.refresh(entry)
    assert result == {"processed": 1, "failed": 0, "skipped_year": 0}
    assert entry.workload_status == "processed"
    assert entry.workload_hours == 30
    assert db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).count() == 8


def test_journal_workload_ignores_disciplines_total_row_without_teacher(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Journals 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-85-26",
        journal_name="85-26 Journal",
        group_code="85-26",
        workload_status="pending",
        workload_year=2026,
    )
    db_session.add(entry)
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_workbook_with_total_row_bytes(
            [
                ("Охорона праці", 4, "3", "Паращук Олег Леонідович"),
                ("Роль AI у сучасному ринку праці", 5, "5", "Паращук Олег Леонідович"),
                ("Практичне використання AI у повсякденній роботі", 5, "7", "Паращук Олег Леонідович"),
                ("AI у творчих і технічних професіях", 5, "9", "Паращук Олег Леонідович"),
                ("Працевлаштування та кар'єрний ріст завдяки AI", 5, "11", "Паращук Олег Леонідович"),
                ("Етика, ризики та відповідальне", 2, "13", "Паращук Олег Леонідович"),
                ("Фінальне тестування", 2, "15", "Паращук Олег Леонідович"),
                ("Підсумкове заняття", 2, "19", "Паращук Олег Леонідович"),
            ],
            30,
        ),
        raising=False,
    )

    result = journal_monitor.process_next_journal_workload(db_session, section, limit=1, target_year=2026)
    db_session.commit()

    db_session.refresh(entry)
    assert result == {"processed": 1, "failed": 0, "skipped_year": 0}
    assert entry.workload_status == "processed"
    assert entry.workload_hours == 30
    assert db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).count() == 8


def test_journal_workload_auto_start_processes_one_2026_journal_and_updates_teacher_workload(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    existing_teacher = Teacher(
        branch_id="main",
        first_name="Олена Петрівна",
        last_name="Коваль",
        hourly_rate=0,
        is_active=True,
    )
    db_session.add(existing_teacher)
    db_session.commit()

    drive_folders = lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-73-26",
                "name": "73-26 Журнал з кібербезпеки",
                "url": "https://drive.google.com/drive/folders/drive-73-26",
                "modified_time": "2026-02-01T10:00:00Z",
            },
            {
                "id": "drive-74-26",
                "name": "74-26 Журнал з обліку",
                "url": "https://drive.google.com/drive/folders/drive-74-26",
                "modified_time": "2026-02-02T10:00:00Z",
            },
            {
                "id": "drive-10-25",
                "name": "10-25 Архівний журнал",
                "url": "https://drive.google.com/drive/folders/drive-10-25",
                "modified_time": "2025-02-01T10:00:00Z",
            },
        ]
    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", drive_folders)
    monkeypatch.setattr("app.tasks.worker.list_drive_child_folders", drive_folders)
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": f"{folder_id}.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_workbook_bytes(
            [
                ("Основи кібербезпеки", 12, "1-18", "Коваль Олена Петрівна"),
                ("Безпечна робота", 8, "19-25", "Шевченко Марія Іванівна"),
            ]
        ),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]

    sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/workload-auto/start?year=2026", headers=auth_headers)

    assert sync_response.status_code == 200
    assert sync_response.json()["workload_auto_enabled"] is True
    assert sync_response.json()["workload_auto_year"] == 2026
    entries = {item["drive_file_id"]: item for item in sync_response.json()["entries"]}
    assert entries["drive-73-26"]["workload_status"] == "processed"
    assert entries["drive-73-26"]["workload_hours"] == 20
    assert entries["drive-74-26"]["workload_status"] == "pending"
    assert entries["drive-10-25"]["workload_status"] == "skipped_year"

    summary = {row["teacher_name"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert summary["Коваль Олена Петрівна"]["total_hours"] == 12
    assert summary["Шевченко Марія Іванівна"]["total_hours"] == 8

    from app.tasks.worker import process_journal_monitor_auto_task

    process_journal_monitor_auto_task.run()
    process_journal_monitor_auto_task.run()
    second_response = client.get(f"/api/v1/journal-monitors/{section_id}", headers=auth_headers)
    second_entries = {item["drive_file_id"]: item for item in second_response.json()["entries"]}
    assert second_entries["drive-74-26"]["workload_status"] == "processed"

    refreshed_summary = {row["teacher_name"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert refreshed_summary["Коваль Олена Петрівна"]["total_hours"] == 24
    assert refreshed_summary["Шевченко Марія Іванівна"]["total_hours"] == 16


def test_journal_processing_start_is_fast_and_tick_processes_trainees(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("processing/start must use existing journal entries")),
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes(
            [(1, "З-СНН-001", "Петренко Іван Іванович", "ч", "01.02.1990", "1234567890", "м. Львів", "+380501112233")]
        ),
        raising=False,
    )
    monkeypatch.setattr(
        "app.tasks.worker.process_journal_monitor_auto_task.delay",
        lambda: (_ for _ in ()).throw(AssertionError("processing/start must not call Celery in the web request")),
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    db_session.add(
        journal_monitor.JournalMonitorEntry(
            section_id=section_id,
            branch_id="main",
            drive_file_id="drive-46-26",
            drive_url="https://drive.google.com/drive/folders/drive-46-26",
            journal_name="46-26 Журнал",
            group_code="46-26",
            workload_status="processed",
            workload_year=2026,
            workload_hours=8,
        )
    )
    db_session.commit()

    response = client.post(f"/api/v1/journal-monitors/{section_id}/processing/start?year=2026", headers=auth_headers)

    assert response.status_code == 200
    assert response.json()["workload_auto_enabled"] is True
    assert response.json()["workload_auto_year"] == 2026
    entry = response.json()["entries"][0]
    assert entry["workload_status"] == "processed"
    assert entry["trainees_status"] == "pending"

    tick_response = client.post(f"/api/v1/journal-monitors/{section_id}/processing/tick", headers=auth_headers)

    assert tick_response.status_code == 200
    assert tick_response.json()["workload_auto_enabled"] is True
    tick_entry = tick_response.json()["entries"][0]
    assert tick_entry["trainees_status"] == "processed"
    assert tick_entry["trainee_count"] == 1
    assert "слухачі" in tick_response.json()["last_sync_message"]


def test_journal_processing_tick_processes_pending_workload_before_trainees(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("processing/start must use existing journal entries")),
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    db_session.add(
        journal_monitor.JournalMonitorEntry(
            section_id=section_id,
            branch_id="main",
            drive_file_id="drive-46-26",
            drive_url="https://drive.google.com/drive/folders/drive-46-26",
            journal_name="46-26 Журнал",
            group_code="46-26",
            workload_status="pending",
            trainees_status="pending",
        )
    )
    db_session.commit()

    response = client.post(f"/api/v1/journal-monitors/{section_id}/processing/start?year=2026", headers=auth_headers)
    assert response.status_code == 200

    tick_response = client.post(f"/api/v1/journal-monitors/{section_id}/processing/tick", headers=auth_headers)

    assert tick_response.status_code == 200
    tick_entry = tick_response.json()["entries"][0]
    assert tick_entry["workload_status"] == "processed"
    assert tick_entry["workload_hours"] == 8
    assert tick_entry["trainees_status"] == "processed"
    assert tick_entry["trainee_count"] == 1
    assert tick_response.json()["workload_auto_enabled"] is True
    summary = {row["teacher_name"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert summary["Коваль Олена Петрівна"]["total_hours"] == 8


def test_background_tick_processes_existing_queue_when_drive_folder_sync_fails(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    drive_sync_calls = 0

    def forbidden_drive_sync(*args, **kwargs):
        nonlocal drive_sync_calls
        drive_sync_calls += 1
        raise AssertionError("background tick must process the existing queue before Drive sync")

    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", forbidden_drive_sync)
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    db_session.add(
        journal_monitor.JournalMonitorEntry(
            section_id=section_id,
            branch_id="main",
            drive_file_id="drive-46-26",
            drive_url="https://drive.google.com/drive/folders/drive-46-26",
            journal_name="46-26 Журнал",
            group_code="46-26",
            workload_status="pending",
            trainees_status="pending",
        )
    )
    db_session.commit()

    response = client.post(f"/api/v1/journal-monitors/{section_id}/processing/background-tick?year=2026&sync=true", headers=auth_headers)

    assert response.status_code == 200
    assert drive_sync_calls == 1
    assert "Drive" in response.json()["last_sync_message"]
    entry = response.json()["entries"][0]
    assert entry["workload_status"] == "processed"
    assert entry["trainees_status"] == "processed"


def test_background_tick_skips_drive_sync_by_default_for_existing_queue(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    drive_sync_calls = 0

    def forbidden_drive_sync(*args, **kwargs):
        nonlocal drive_sync_calls
        drive_sync_calls += 1
        raise AssertionError("background tick should process the existing queue without Drive sync by default")

    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", forbidden_drive_sync)
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    db_session.add(
        journal_monitor.JournalMonitorEntry(
            section_id=section_id,
            branch_id="main",
            drive_file_id="drive-46-26",
            drive_url="https://drive.google.com/drive/folders/drive-46-26",
            journal_name="46-26 Журнал",
            group_code="46-26",
            workload_status="pending",
            trainees_status="pending",
        )
    )
    db_session.commit()

    response = client.post(f"/api/v1/journal-monitors/{section_id}/processing/background-tick?year=2026", headers=auth_headers)

    assert response.status_code == 200
    assert drive_sync_calls == 0
    entry = response.json()["entries"][0]
    assert entry["workload_status"] == "processed"
    assert entry["trainees_status"] == "processed"


def test_background_tick_processes_only_first_pending_journal_by_default(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("background tick should not sync Drive by default")),
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": f"{folder_id}.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    db_session.add_all(
        [
            journal_monitor.JournalMonitorEntry(
                section_id=section_id,
                branch_id="main",
                drive_file_id="drive-1-26",
                drive_url="https://drive.google.com/drive/folders/drive-1-26",
                journal_name="1-26 Журнал",
                group_code="1-26",
                workload_status="pending",
                trainees_status="pending",
            ),
            journal_monitor.JournalMonitorEntry(
                section_id=section_id,
                branch_id="main",
                drive_file_id="drive-10-26",
                drive_url="https://drive.google.com/drive/folders/drive-10-26",
                journal_name="10-26 Журнал",
                group_code="10-26",
                workload_status="pending",
                trainees_status="pending",
            ),
        ]
    )
    db_session.commit()

    response = client.post(f"/api/v1/journal-monitors/{section_id}/processing/background-tick?year=2026", headers=auth_headers)

    assert response.status_code == 200
    entries = {entry["group_code"]: entry for entry in response.json()["entries"]}
    assert entries["1-26"]["workload_status"] == "processed"
    assert entries["1-26"]["trainees_status"] == "processed"
    assert entries["10-26"]["workload_status"] == "pending"
    assert entries["10-26"]["trainees_status"] == "pending"


def test_background_tick_finishes_one_partially_processed_journal_before_next_one(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("background tick should not sync Drive by default")),
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": f"{folder_id}.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    db_session.add_all(
        [
            journal_monitor.JournalMonitorEntry(
                section_id=section_id,
                branch_id="main",
                drive_file_id="drive-1-26",
                drive_url="https://drive.google.com/drive/folders/drive-1-26",
                journal_name="1-26 Журнал",
                group_code="1-26",
                workload_status="processed",
                workload_year=2026,
                workload_hours=8,
                trainees_status="pending",
            ),
            journal_monitor.JournalMonitorEntry(
                section_id=section_id,
                branch_id="main",
                drive_file_id="drive-10-26",
                drive_url="https://drive.google.com/drive/folders/drive-10-26",
                journal_name="10-26 Журнал",
                group_code="10-26",
                workload_status="pending",
                trainees_status="pending",
            ),
        ]
    )
    db_session.commit()

    response = client.post(f"/api/v1/journal-monitors/{section_id}/processing/background-tick?year=2026", headers=auth_headers)

    assert response.status_code == 200
    entries = {entry["group_code"]: entry for entry in response.json()["entries"]}
    assert entries["1-26"]["workload_status"] == "processed"
    assert entries["1-26"]["trainees_status"] == "processed"
    assert entries["10-26"]["workload_status"] == "pending"
    assert entries["10-26"]["trainees_status"] == "pending"


def test_background_tick_discovers_new_drive_folder_for_existing_section(
    client,
    auth_headers,
    monkeypatch,
):
    initial_folders = [
        {
            "id": "drive-46-26",
            "name": "46-26 Журнал",
            "url": "https://drive.google.com/drive/folders/drive-46-26",
            "modified_time": "2026-05-01T10:00:00Z",
        },
    ]
    folders_with_new_journal = [
        *initial_folders,
        {
            "id": "drive-85-26",
            "name": "85-26 Журнал",
            "url": "https://drive.google.com/drive/folders/drive-85-26",
            "modified_time": "2026-05-01T11:00:00Z",
        },
    ]
    drive_sync_calls = 0

    def list_folders(_folder_id, service_account_json=None):
        nonlocal drive_sync_calls
        drive_sync_calls += 1
        return initial_folders if drive_sync_calls == 1 else folders_with_new_journal

    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", list_folders)
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": f"{folder_id}.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    first_sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)
    assert first_sync_response.status_code == 200
    assert [entry["group_code"] for entry in first_sync_response.json()["entries"]] == ["46-26"]

    tick_response = client.post(
        f"/api/v1/journal-monitors/{section_id}/processing/background-tick?year=2026&sync=true&workload_limit=20&trainees_limit=20",
        headers=auth_headers,
    )

    assert tick_response.status_code == 200
    entries = {entry["group_code"]: entry for entry in tick_response.json()["entries"]}
    assert set(entries) == {"46-26", "85-26"}
    assert entries["85-26"]["workload_status"] == "processed"
    assert entries["85-26"]["trainees_status"] == "processed"


def test_journal_auto_worker_processes_trainees_and_workload_together(db_session, monkeypatch):
    monkeypatch.setattr(
        "app.tasks.worker.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-46-26",
                "name": "46-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-46-26",
                "modified_time": "2026-03-02T10:00:00Z",
            },
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_enabled=True,
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.commit()

    from app.tasks.worker import process_journal_monitor_auto_task

    result = process_journal_monitor_auto_task.run()

    assert result == {"processed_sections": 1, "failed_sections": 0}
    entry = section.entries[0]
    assert entry.trainees_status == "processed"
    assert entry.trainee_count == 1
    assert entry.workload_status == "processed"
    assert entry.workload_hours == 8
    summary = {row["teacher_name"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert summary["Коваль Олена Петрівна"]["total_hours"] == 8


def test_journal_auto_worker_processes_active_sections_without_manual_auto_toggle(db_session, monkeypatch):
    monkeypatch.setattr(
        "app.tasks.worker.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-1-26",
                "name": "1-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-1-26",
                "modified_time": "2026-03-02T10:00:00Z",
            },
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "1-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_enabled=False,
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.commit()

    from app.tasks.worker import process_journal_monitor_auto_task

    result = process_journal_monitor_auto_task.run()

    assert result == {"processed_sections": 1, "failed_sections": 0}
    entry = section.entries[0]
    assert entry.group_code == "1-26"
    assert entry.workload_status == "processed"
    assert entry.trainees_status == "processed"
    assert entry.trainee_count == 1


def test_journal_auto_cron_endpoint_processes_active_sections(client, auth_headers, db_session, monkeypatch):
    monkeypatch.setattr(journal_monitor.settings, "cron_secret", "cron-secret")
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-1-26",
                "name": "1-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-1-26",
                "modified_time": "2026-03-02T10:00:00Z",
            },
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "1-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_enabled=False,
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.commit()

    response = client.post(
        "/api/v1/journal-monitors/auto-cron",
        headers={"Authorization": "Bearer cron-secret"},
    )

    assert response.status_code == 202
    assert response.json() == {"processed_sections": 1, "failed_sections": 0}
    db_session.refresh(section)
    entry = section.entries[0]
    assert entry.group_code == "1-26"
    assert entry.workload_status == "processed"
    assert entry.trainees_status == "processed"
    assert entry.trainee_count == 1


def test_journal_auto_tick_endpoint_processes_current_branch_sections(client, auth_headers, db_session, monkeypatch):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-1-26",
                "name": "1-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-1-26",
                "modified_time": "2026-03-02T10:00:00Z",
            },
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "1-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.process_next_drive_intake_file",
        lambda db, **kwargs: {"processed": 0, "skipped_already_processed": 0, "skipped_unsupported": 0},
    )
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_enabled=False,
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.commit()

    response = client.post("/api/v1/journal-monitors/auto-tick", headers=auth_headers)

    assert response.status_code == 202
    payload = response.json()
    assert payload["processed_sections"] == 1
    assert payload["failed_sections"] == 0
    assert payload["drive_intake_processed"] == 0
    assert payload["drive_intake_failed"] == 0
    db_session.refresh(section)
    entry = section.entries[0]
    assert entry.group_code == "1-26"
    assert entry.workload_status == "processed"
    assert entry.trainees_status == "processed"
    assert entry.trainee_count == 1


def test_journal_auto_tick_endpoint_processes_one_drive_intake_file(client, auth_headers, monkeypatch):
    captured: dict[str, object] = {}

    def fake_process_next_drive_intake_file(db, **kwargs):
        captured["branch_id"] = kwargs.get("branch_id")
        captured["import_job_runner"] = kwargs.get("import_job_runner")
        return {
            "processed": 1,
            "skipped_already_processed": 0,
            "skipped_unsupported": 0,
            "filename": "46-26 Розклад.docx",
        }

    monkeypatch.setattr(
        "app.api.routes.journal_monitors.process_next_drive_intake_file",
        fake_process_next_drive_intake_file,
        raising=False,
    )

    response = client.post("/api/v1/journal-monitors/auto-tick", headers=auth_headers)

    assert response.status_code == 202
    payload = response.json()
    assert payload["processed_sections"] == 0
    assert payload["failed_sections"] == 0
    assert payload["drive_intake_processed"] == 1
    assert payload["drive_intake_failed"] == 0
    assert payload["drive_intake_filename"] == "46-26 Розклад.docx"
    assert captured["branch_id"] == "main"
    assert captured["import_job_runner"] is not None


def test_journal_auto_tick_reports_drive_intake_error_message(client, auth_headers, monkeypatch):
    def fail_drive_intake(db, **kwargs):
        raise RuntimeError("Не вдалося отримати доступ до папки Google Drive")

    monkeypatch.setattr(
        "app.api.routes.journal_monitors.process_next_drive_intake_file",
        fail_drive_intake,
        raising=False,
    )

    response = client.post("/api/v1/journal-monitors/auto-tick", headers=auth_headers)

    assert response.status_code == 202
    payload = response.json()
    assert payload["drive_intake_processed"] == 0
    assert payload["drive_intake_failed"] == 1
    assert payload["drive_intake_message"] == "Не вдалося отримати доступ до папки Google Drive"


def test_journal_auto_tick_reuses_section_drive_credentials_for_intake(client, auth_headers, db_session, monkeypatch):
    captured: dict[str, object] = {}
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        service_account_json_encrypted=cipher.encrypt("section-service-account-json"),
    )
    db_session.add(section)
    db_session.commit()

    monkeypatch.setattr(
        "app.api.routes.journal_monitors.process_journal_monitor_background_step",
        lambda db, section, **kwargs: None,
    )

    def fake_process_next_drive_intake_file(db, **kwargs):
        captured["service_account_json"] = kwargs.get("service_account_json")
        return {"processed": 0, "skipped_already_processed": 0, "skipped_unsupported": 0}

    monkeypatch.setattr(
        "app.api.routes.journal_monitors.process_next_drive_intake_file",
        fake_process_next_drive_intake_file,
        raising=False,
    )

    response = client.post("/api/v1/journal-monitors/auto-tick", headers=auth_headers)

    assert response.status_code == 202
    assert captured["service_account_json"] == "section-service-account-json"


def test_journal_auto_worker_processes_pending_trainees_one_journal_per_tick(db_session, monkeypatch):
    drive_folders = lambda _folder_id, service_account_json=None: [
        {
            "id": "drive-51-26",
            "name": "51-26 Журнал",
            "url": "https://drive.google.com/drive/folders/drive-51-26",
            "modified_time": "2026-03-02T10:00:00Z",
        },
        {
            "id": "drive-52-26",
            "name": "52-26 Журнал",
            "url": "https://drive.google.com/drive/folders/drive-52-26",
            "modified_time": "2026-03-03T10:00:00Z",
        },
    ]
    monkeypatch.setattr("app.tasks.worker.list_drive_child_folders", drive_folders)
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": f"{folder_id}.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    downloaded: list[str] = []

    def fake_download(file_id, mime_type=None, service_account_json=None):
        downloaded.append(file_id)
        is_first = "51-26" in file_id
        contract = "З-СНН-051" if is_first else "З-СНН-052"
        full_name = "Перший Іван Іванович" if is_first else "Другий Іван Іванович"
        return _journal_zv_workbook_bytes(
            [(1, contract, full_name, "ч", "01.02.1990", "", "", "")]
        )

    monkeypatch.setattr(journal_monitor, "download_drive_file_bytes", fake_download, raising=False)
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_enabled=True,
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.flush()
    for code, drive_id in (("51-26", "drive-51-26"), ("52-26", "drive-52-26")):
        db_session.add(
            journal_monitor.JournalMonitorEntry(
                section_id=section.id,
                branch_id="main",
                drive_file_id=drive_id,
                drive_url=f"https://drive.google.com/drive/folders/{drive_id}",
                journal_name=f"{code} Журнал",
                group_code=code,
                workload_status="processed",
                workload_year=2026,
                workload_hours=8,
                trainees_status="pending",
            )
        )
    db_session.commit()

    from app.tasks.worker import process_journal_monitor_auto_task

    result = process_journal_monitor_auto_task.run()

    db_session.expire_all()
    assert result == {"processed_sections": 1, "failed_sections": 0}
    entries = {entry.group_code: entry for entry in section.entries}
    assert entries["51-26"].trainees_status == "processed"
    assert entries["52-26"].trainees_status == "pending"

    second_result = process_journal_monitor_auto_task.run()

    db_session.expire_all()
    assert second_result == {"processed_sections": 1, "failed_sections": 0}
    entries = {entry.group_code: entry for entry in section.entries}
    assert entries["52-26"].trainees_status == "processed"
    assert downloaded == ["drive-51-26-xlsx", "drive-52-26-xlsx"]
    assert db_session.query(Trainee).filter(Trainee.group_code == "51-26").count() == 1
    assert db_session.query(Trainee).filter(Trainee.group_code == "52-26").count() == 1


def test_start_requeue_keeps_already_imported_trainees_out_of_front_of_queue(db_session):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_enabled=True,
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.flush()
    imported = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-1-26",
        journal_name="1-26 Журнал",
        group_code="1-26",
        has_trainees=True,
        trainee_count=33,
        trainees_status="processed",
    )
    pending = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-2-26",
        journal_name="2-26 Журнал",
        group_code="2-26",
        has_trainees=False,
        trainee_count=0,
        trainees_status="no_data",
    )
    db_session.add_all([imported, pending])
    db_session.commit()

    changed = journal_monitor.requeue_journal_trainees_for_year(db_session, section, 2026)

    assert changed == 1
    assert imported.trainees_status == "processed"
    assert pending.trainees_status == "pending"


def test_force_requeue_marks_processed_workload_and_trainees_pending(db_session):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    teacher = Teacher(branch_id="main", first_name="Іванович", last_name="Викладач", is_active=True)
    db_session.add_all([section, teacher])
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-1-26",
        journal_name="1-26 Журнал",
        group_code="1-26",
        has_trainees=True,
        trainee_count=33,
        workload_status="processed",
        workload_year=2026,
        workload_hours=8,
        trainees_status="processed",
    )
    db_session.add(entry)
    db_session.flush()
    db_session.add(
        JournalWorkloadEntry(
            journal_monitor_entry_id=entry.id,
            branch_id="main",
            teacher_id=teacher.id,
            subject_name="Предмет",
            hours=8,
        )
    )
    db_session.commit()

    workload_changed = journal_monitor.requeue_journal_workload_for_year(db_session, section, 2026, force=True)
    trainees_changed = journal_monitor.requeue_journal_trainees_for_year(db_session, section, 2026, force=True)

    assert workload_changed == 1
    assert trainees_changed == 1
    assert entry.workload_status == "pending"
    assert entry.trainees_status == "pending"
    assert entry.workload_hours == 0
    assert db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).count() == 0


def test_reprocess_all_queues_existing_entries_without_drive_sync(client, auth_headers, db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    teacher = Teacher(branch_id="main", first_name="Іванович", last_name="Викладач", is_active=True)
    db_session.add_all([section, teacher])
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-1-26",
        journal_name="1-26 Журнал",
        group_code="1-26",
        has_trainees=True,
        trainee_count=12,
        workload_status="processed",
        workload_year=2026,
        workload_hours=8,
        trainees_status="processed",
    )
    db_session.add(entry)
    db_session.flush()
    db_session.add(
        JournalWorkloadEntry(
            journal_monitor_entry_id=entry.id,
            branch_id="main",
            teacher_id=teacher.id,
            subject_name="Предмет",
            hours=8,
        )
    )
    db_session.commit()

    drive_sync_calls = 0

    def forbidden_drive_lister(_folder_id, service_account_json=None):
        nonlocal drive_sync_calls
        drive_sync_calls += 1
        raise AssertionError("full reprocessing must queue existing entries without Drive sync")

    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", forbidden_drive_lister)

    response = client.post(
        f"/api/v1/journal-monitors/{section.id}/processing/reprocess-all?year=2026",
        headers=auth_headers,
    )

    assert response.status_code == 200
    payload = response.json()
    assert drive_sync_calls == 0
    assert payload["workload_auto_enabled"] is True
    assert "Повна переобробка 2026" in payload["last_sync_message"]
    assert payload["entries"][0]["workload_status"] == "pending"
    assert payload["entries"][0]["trainees_status"] == "pending"
    assert payload["entries"][0]["workload_hours"] == 0
    assert db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).count() == 0


def test_journal_step_message_does_not_grow_past_database_limit(db_session):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_enabled=True,
        workload_auto_year=2026,
        last_sync_message="x" * 500,
    )
    db_session.add(section)
    db_session.commit()

    journal_monitor.process_journal_monitor_section_step(db_session, section)
    db_session.commit()

    assert section.last_sync_message is not None
    assert len(section.last_sync_message) <= 500


def test_journal_workload_can_be_run_for_2025_on_demand(client, auth_headers, monkeypatch):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-10-25",
                "name": "10-25 Архівний журнал",
                "url": "https://drive.google.com/drive/folders/drive-10-25",
                "modified_time": "2025-02-01T10:00:00Z",
            },
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": f"{folder_id}.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_workbook_bytes(
            [("Архівна дисципліна", 6, "1-6", "Петренко Ігор Степанович")]
        ),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2025", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)
    assert sync_response.json()["entries"][0]["workload_status"] == "pending"

    process_response = client.post(
        f"/api/v1/journal-monitors/{section_id}/workload-auto/start?year=2025",
        headers=auth_headers,
    )

    assert process_response.status_code == 200
    entry = process_response.json()["entries"][0]
    assert entry["workload_status"] == "processed"
    assert entry["workload_year"] == 2025
    assert entry["workload_hours"] == 6


def test_journal_workload_parses_two_workbooks_lowercase_sheet_and_splits_multiple_teachers(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    db_session.add(
        Teacher(
            branch_id="main",
            first_name="Віктор Євгенович",
            last_name="Брикін",
            hourly_rate=0,
            is_active=True,
        )
    )
    db_session.commit()

    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-80-26",
                "name": "80-26 Журнал теорії і виробничого навчання",
                "url": "https://drive.google.com/drive/folders/drive-80-26",
                "modified_time": "2026-03-01T10:00:00Z",
            },
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": "theory-xlsx", "name": "Теорія.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
            {"id": "practice-xlsx", "name": "Виробниче навчання.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        ],
        raising=False,
    )

    def fake_download(file_id, mime_type=None, service_account_json=None):
        if file_id == "theory-xlsx":
            return _journal_workbook_bytes(
                [("Матеріалознавство", 5, "1-5", "Брикін Віктор Євгенович, Старожук Людмила Василівна")],
                sheet_name="дисципліни",
                discipline_sheet_index=3,
            )
        return _journal_workbook_bytes(
            [("Практика", 4, "6-9", "БрикінВіктор Євгенович")],
            sheet_name="ДИСЦИПЛІНИ",
            discipline_sheet_index=4,
        )

    monkeypatch.setattr(journal_monitor, "download_drive_file_bytes", fake_download, raising=False)

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    response = client.post(f"/api/v1/journal-monitors/{section_id}/workload-auto/start?year=2026", headers=auth_headers)

    assert response.status_code == 200
    entry = response.json()["entries"][0]
    assert entry["workload_status"] == "processed"
    assert entry["workload_source_names"] == ["Теорія", "Виробниче навчання"]
    summary = {row["teacher_name"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert summary["Брикін Віктор Євгенович"]["total_hours"] == 9
    assert summary["Старожук Людмила Василівна"]["total_hours"] == 5
    assert "Брикін Віктор Євгенович, Старожук Людмила Василівна" not in summary


def test_auto_sync_retries_failed_journal_after_access_is_fixed(client, auth_headers, db_session, monkeypatch):
    drive_folders = lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-82-26",
                "name": "82-26 Журнал з тимчасово закритим доступом",
                "url": "https://drive.google.com/drive/folders/drive-82-26",
                "modified_time": "2026-03-03T10:00:00Z",
            },
        ]
    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", drive_folders)
    monkeypatch.setattr("app.tasks.worker.list_drive_child_folders", drive_folders)
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": "restricted-xlsx", "name": "82-26 Трактористи виробн.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    access_fixed = {"value": False}

    def fake_download(file_id, mime_type=None, service_account_json=None):
        if not access_fixed["value"]:
            raise PermissionError("Немає доступу до файлу")
        return _journal_workbook_bytes([("Доступ відновлено", 7, "1-7", "Доступний Викладач")])

    monkeypatch.setattr(journal_monitor, "download_drive_file_bytes", fake_download, raising=False)

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    first_response = client.post(f"/api/v1/journal-monitors/{section_id}/workload-auto/start?year=2026", headers=auth_headers)
    assert first_response.json()["entries"][0]["workload_status"] == "failed"
    assert first_response.json()["entries"][0]["workload_source_names"] == ["82-26 Трактористи виробн"]

    access_fixed["value"] = True
    from app.tasks.worker import process_journal_monitor_auto_task

    process_journal_monitor_auto_task.run()
    retry_response = client.get(f"/api/v1/journal-monitors/{section_id}", headers=auth_headers)

    entry = retry_response.json()["entries"][0]
    assert entry["workload_status"] == "processed"
    assert entry["workload_hours"] == 7


def test_deleting_teacher_marks_related_journal_for_regeneration(client, auth_headers, db_session, monkeypatch):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-81-26",
                "name": "81-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-81-26",
                "modified_time": "2026-03-02T10:00:00Z",
            },
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": "journal-xlsx", "name": "Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_workbook_bytes(
            [("Дисципліна", 3, "1-3", "Тимчасовий Викладач")]
        ),
        raising=False,
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    client.post(f"/api/v1/journal-monitors/{section_id}/workload-auto/start?year=2026", headers=auth_headers)
    teacher_id = db_session.query(Teacher).filter(Teacher.last_name == "Тимчасовий").one().id

    delete_response = client.delete(f"/api/v1/teachers/{teacher_id}", headers=auth_headers)

    assert delete_response.status_code == 204
    detail_response = client.get(f"/api/v1/journal-monitors/{section_id}", headers=auth_headers)
    entry = detail_response.json()["entries"][0]
    assert entry["workload_status"] == "needs_regeneration"
    assert entry["workload_hours"] == 0
    assert db_session.query(JournalWorkloadEntry).count() == 0


def test_journal_monitor_sync_compares_drive_folders_with_project_data(client, auth_headers, db_session, monkeypatch):
    complete_group = Group(branch_id="main", code="180-25", name="Штучний інтелект", status=GroupStatus.ACTIVE)
    schedule_group = Group(branch_id="main", code="167-25", name="Трудові відносини", status=GroupStatus.ACTIVE)
    db_session.add_all([complete_group, schedule_group])
    db_session.flush()
    _seed_schedule(db_session, complete_group, "180")
    _seed_schedule(db_session, schedule_group, "167")
    db_session.add(Trainee(branch_id="main", first_name="Іван", last_name="Повний", status="active", group_code="180-25"))
    db_session.add(Trainee(branch_id="main", first_name="Олена", last_name="ТількиСлухачі", status="active", group_code="162-25"))
    db_session.commit()

    def fake_drive_folders(_folder_id: str, service_account_json: str | None = None):
        return [
            {
                "id": "drive-180",
                "name": "180-25 Штучний інтелект: розвиток кар'єри",
                "url": "https://drive.google.com/drive/folders/drive-180",
                "modified_time": "2026-05-01T10:00:00Z",
            },
            {
                "id": "drive-167",
                "name": "167-25 Організація трудових відносин",
                "url": "https://drive.google.com/drive/folders/drive-167",
                "modified_time": "2026-05-01T11:00:00Z",
            },
            {
                "id": "drive-162",
                "name": "162-25 Штучний інтелект",
                "url": "https://drive.google.com/drive/folders/drive-162",
                "modified_time": "2026-05-01T12:00:00Z",
            },
            {
                "id": "drive-999",
                "name": "999-25 Немає в системі",
                "url": "https://drive.google.com/drive/folders/drive-999",
                "modified_time": "2026-05-01T13:00:00Z",
            },
        ]

    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", fake_drive_folders)

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    assert create_response.status_code == 201
    section_id = create_response.json()["id"]

    sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)
    assert sync_response.status_code == 200
    assert sync_response.json()["stats"]["total"] == 4
    assert sync_response.json()["stats"]["complete"] == 0
    assert sync_response.json()["stats"]["schedule_only"] == 2
    assert sync_response.json()["stats"]["trainees_only"] == 0
    assert sync_response.json()["stats"]["not_processed"] == 2
    assert sync_response.json()["stats"]["workload_and_trainees"] == 0
    assert sync_response.json()["stats"]["workload_trainees_schedule"] == 0

    for entry in db_session.query(journal_monitor.JournalMonitorEntry).filter(
        journal_monitor.JournalMonitorEntry.section_id == section_id,
        journal_monitor.JournalMonitorEntry.group_code.in_(["180-25", "162-25"]),
    ):
        entry.workload_status = "processed"
        entry.workload_hours = 30
        entry.trainees_status = "processed"
        entry.trainees_message = "Додано/оновлено слухачів із журналу: 1"
        db_session.add(entry)
    db_session.commit()

    detail_response = client.get(f"/api/v1/journal-monitors/{section_id}", headers=auth_headers)
    assert detail_response.status_code == 200
    assert detail_response.json()["stats"]["trainees_only"] == 0
    assert detail_response.json()["stats"]["workload_and_trainees"] == 1
    assert detail_response.json()["stats"]["workload_trainees_schedule"] == 1
    entries = {item["group_code"]: item for item in detail_response.json()["entries"]}
    assert entries["180-25"]["has_schedule"] is True
    assert entries["180-25"]["has_trainees"] is True
    assert entries["180-25"]["processing_status"] == "complete"
    assert entries["167-25"]["processing_status"] == "schedule_only"
    assert entries["162-25"]["processing_status"] == "trainees_only"
    assert entries["999-25"]["processing_status"] == "not_processed"


def test_journal_monitor_exports_csv(client, auth_headers, db_session, monkeypatch):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-180",
                "name": "180-25 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-180",
                "modified_time": None,
            }
        ],
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)

    export_response = client.get(f"/api/v1/journal-monitors/{section_id}/export?format=csv", headers=auth_headers)
    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith("text/csv")
    assert "Номер групи" in export_response.text
    assert "180-25" in export_response.text


def test_journal_monitor_export_respects_filters(client, auth_headers, db_session, monkeypatch):
    complete_group = Group(branch_id="main", code="180-25", name="Штучний інтелект", status=GroupStatus.ACTIVE)
    schedule_group = Group(branch_id="main", code="167-25", name="Трудові відносини", status=GroupStatus.ACTIVE)
    db_session.add_all([complete_group, schedule_group])
    db_session.flush()
    _seed_schedule(db_session, complete_group, "180-filter")
    _seed_schedule(db_session, schedule_group, "167-filter")
    db_session.add(Trainee(branch_id="main", first_name="Іван", last_name="Повний", status="active", group_code="180-25"))
    db_session.add(Trainee(branch_id="main", first_name="Олена", last_name="Слухачі", status="active", group_code="162-25"))
    db_session.commit()

    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-180",
                "name": "180-25 Штучний інтелект",
                "url": "https://drive.google.com/drive/folders/drive-180",
                "modified_time": None,
            },
            {
                "id": "drive-167",
                "name": "167-25 Організація трудових відносин",
                "url": "https://drive.google.com/drive/folders/drive-167",
                "modified_time": None,
            },
            {
                "id": "drive-162",
                "name": "162-25 Штучний інтелект",
                "url": "https://drive.google.com/drive/folders/drive-162",
                "modified_time": None,
            },
            {
                "id": "drive-999",
                "name": "999-25 Не опрацьовано",
                "url": "https://drive.google.com/drive/folders/drive-999",
                "modified_time": None,
            },
        ],
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)

    export_response = client.get(
        f"/api/v1/journal-monitors/{section_id}/export?format=csv&status=schedule_only&has_schedule=true&has_trainees=false&q=відносин",
        headers=auth_headers,
    )

    assert export_response.status_code == 200
    assert "167-25" in export_response.text
    assert "180-25" not in export_response.text
    assert "162-25" not in export_response.text
    assert "999-25" not in export_response.text


def test_journal_monitor_export_filters_by_workload_presence(client, auth_headers, db_session, monkeypatch):
    scheduled_group = Group(branch_id="main", code="167-25", name="Трудові відносини", status=GroupStatus.ACTIVE)
    db_session.add(scheduled_group)
    db_session.flush()
    _seed_schedule(db_session, scheduled_group, "167-workload-filter")
    db_session.add(Trainee(branch_id="main", first_name="Олена", last_name="Слухачі", status="active", group_code="162-25"))
    db_session.commit()

    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-167",
                "name": "167-25 Є розклад і педнавантаження",
                "url": "https://drive.google.com/drive/folders/drive-167",
                "modified_time": None,
            },
            {
                "id": "drive-162",
                "name": "162-25 Немає педнавантаження",
                "url": "https://drive.google.com/drive/folders/drive-162",
                "modified_time": None,
            },
            {
                "id": "drive-999",
                "name": "999-25 Тільки педнавантаження",
                "url": "https://drive.google.com/drive/folders/drive-999",
                "modified_time": None,
            },
        ],
    )

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)

    for entry in db_session.query(journal_monitor.JournalMonitorEntry).filter(
        journal_monitor.JournalMonitorEntry.section_id == section_id,
        journal_monitor.JournalMonitorEntry.group_code.in_(["167-25", "999-25"]),
    ):
        entry.workload_status = "processed"
        entry.workload_hours = 10
        db_session.add(entry)
    db_session.commit()

    workload_only_response = client.get(
        f"/api/v1/journal-monitors/{section_id}/export?format=csv&workload=workload_only",
        headers=auth_headers,
    )
    without_workload_response = client.get(
        f"/api/v1/journal-monitors/{section_id}/export?format=csv&workload=without_workload",
        headers=auth_headers,
    )

    assert workload_only_response.status_code == 200
    assert "999-25" in workload_only_response.text
    assert "167-25" not in workload_only_response.text
    assert "162-25" not in workload_only_response.text
    assert without_workload_response.status_code == 200
    assert "162-25" in without_workload_response.text
    assert "167-25" not in without_workload_response.text
    assert "999-25" not in without_workload_response.text


def test_drive_folder_listing_uses_service_account_bearer_token(monkeypatch):
    monkeypatch.setattr(journal_monitor.settings, "google_drive_api_key", "")
    monkeypatch.setattr(
        journal_monitor.settings,
        "google_drive_service_account_json",
        '{"client_email":"drive-reader@example.iam.gserviceaccount.com","private_key":"-----BEGIN PRIVATE KEY-----\\nKEY\\n-----END PRIVATE KEY-----\\n","token_uri":"https://oauth2.googleapis.com/token"}',
    )
    monkeypatch.setattr(journal_monitor, "_get_service_account_access_token", lambda _raw_json=None: "service-token")

    captured: dict[str, object] = {}

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self):
            return (
                b'{"files":[{"id":"folder-1","name":"180-25 Journal","webViewLink":"https://drive/folder-1",'
                b'"modifiedTime":"2026-05-01T10:00:00Z"}]}'
            )

    def fake_urlopen(request, timeout):
        captured["url"] = request.full_url
        captured["authorization"] = request.headers.get("Authorization")
        captured["timeout"] = timeout
        return FakeResponse()

    monkeypatch.setattr(journal_monitor, "urlopen", fake_urlopen)

    folders = journal_monitor.list_drive_child_folders("root-folder")

    assert folders[0]["id"] == "folder-1"
    assert "key=" not in str(captured["url"])
    assert captured["authorization"] == "Bearer service-token"


def test_journal_monitor_sync_without_credentials_explains_service_account_setup(client, auth_headers, monkeypatch):
    monkeypatch.setattr(journal_monitor.settings, "google_drive_api_key", "")
    monkeypatch.setattr(journal_monitor.settings, "google_drive_service_account_json", "")

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали без ключа", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    assert create_response.status_code == 201

    sync_response = client.post(
        f"/api/v1/journal-monitors/{create_response.json()['id']}/sync",
        headers=auth_headers,
    )

    assert sync_response.status_code == 502
    detail = sync_response.json()["detail"]
    assert "GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON" in detail
    assert "suptc-drive-journal-monitor" in detail


def test_journal_monitor_can_store_section_service_account_json(client, auth_headers, monkeypatch):
    monkeypatch.setattr(journal_monitor.settings, "google_drive_api_key", "")
    monkeypatch.setattr(journal_monitor.settings, "google_drive_service_account_json", "")

    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали з ключем", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    assert create_response.status_code == 201
    section_id = create_response.json()["id"]

    patch_response = client.patch(
        f"/api/v1/journal-monitors/{section_id}",
        json={
            "service_account_json": '{"client_email":"drive-reader@example.iam.gserviceaccount.com","private_key":"key"}'
        },
        headers=auth_headers,
    )
    assert patch_response.status_code == 200
    assert patch_response.json()["has_service_account_credentials"] is True
    assert "service_account_json" not in patch_response.json()

    captured: dict[str, object] = {}

    def fake_lister(folder_id: str, service_account_json: str | None = None):
        captured["folder_id"] = folder_id
        captured["service_account_json"] = service_account_json
        return []

    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", fake_lister)

    sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)
    assert sync_response.status_code == 200
    assert captured["folder_id"] == "root-folder"
    assert "drive-reader@example.iam.gserviceaccount.com" in str(captured["service_account_json"])


def test_journal_sync_creates_groups_from_drive_folders(client, auth_headers, db_session, monkeypatch):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-1-26",
                "name": "1-26 Організація трудових відносин в умовах воєнного стану",
                "url": "https://drive.google.com/drive/folders/drive-1-26",
                "modified_time": "2026-05-01T10:00:00Z",
            }
        ],
    )
    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]

    sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)

    assert sync_response.status_code == 200
    entry = sync_response.json()["entries"][0]
    assert entry["group_code"] == "1-26"
    assert entry["has_group"] is True
    group = db_session.query(Group).filter(Group.code == "1-26").one()
    assert group.name == "Організація трудових відносин в умовах воєнного стану"
    groups_response = client.get("/api/v1/groups", headers=auth_headers)
    groups_payload = {item["code"]: item for item in groups_response.json()}
    assert groups_payload["1-26"]["year"] == 2026


def test_delete_journal_entry_and_background_tick_reimports_it(client, auth_headers, monkeypatch):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-46-26",
                "name": "46-26 Журнал",
                "url": "https://drive.google.com/drive/folders/drive-46-26",
                "modified_time": "2026-05-01T10:00:00Z",
            }
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )
    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)
    entry_id = sync_response.json()["entries"][0]["id"]

    delete_response = client.delete(f"/api/v1/journal-monitors/{section_id}/entries/{entry_id}", headers=auth_headers)
    assert delete_response.status_code == 204
    assert client.get(f"/api/v1/journal-monitors/{section_id}", headers=auth_headers).json()["entries"] == []

    tick_response = client.post(
        f"/api/v1/journal-monitors/{section_id}/processing/background-tick?year=2026&sync=true",
        headers=auth_headers,
    )

    assert tick_response.status_code == 200
    entry = tick_response.json()["entries"][0]
    assert entry["group_code"] == "46-26"
    assert entry["workload_status"] == "processed"
    assert entry["trainees_status"] == "processed"
    assert entry["trainee_count"] == 1


def test_delete_journal_entry_removes_workload_from_teacher_summary(client, auth_headers, db_session):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    teacher = Teacher(branch_id="main", first_name="Олена Петрівна", last_name="Коваль", annual_load_hours=100, is_active=True)
    db_session.add_all([section, teacher])
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-46-26",
        journal_name="46-26 Журнал",
        group_code="46-26",
        workload_status="processed",
        workload_year=2026,
        workload_hours=12,
    )
    db_session.add(entry)
    db_session.flush()
    db_session.add(
        JournalWorkloadEntry(
            journal_monitor_entry_id=entry.id,
            branch_id="main",
            teacher_id=teacher.id,
            subject_name="Журнальна дисципліна",
            hours=12,
        )
    )
    db_session.commit()

    before = {row["teacher_id"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert before[teacher.id]["total_hours"] == 12

    response = client.delete(f"/api/v1/journal-monitors/{section.id}/entries/{entry.id}", headers=auth_headers)

    assert response.status_code == 204
    assert db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.teacher_id == teacher.id).count() == 0
    after = {row["teacher_id"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert after[teacher.id]["total_hours"] == 0
    assert after[teacher.id]["remaining_hours"] == 100


def test_background_tick_processes_existing_pending_entries_when_folder_sync_fails(
    client,
    auth_headers,
    db_session,
    monkeypatch,
):
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("Drive root temporarily unavailable")),
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": f"{folder_id}.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )
    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    db_session.add(
        journal_monitor.JournalMonitorEntry(
            section_id=section_id,
            branch_id="main",
            drive_file_id="drive-1-26",
            drive_url="https://drive.google.com/drive/folders/drive-1-26",
            journal_name="1-26 Журнал",
            group_code="1-26",
            workload_status="pending",
            trainees_status="pending",
        )
    )
    db_session.commit()

    tick_response = client.post(
        f"/api/v1/journal-monitors/{section_id}/processing/background-tick?year=2026&sync=true",
        headers=auth_headers,
    )

    assert tick_response.status_code == 200
    payload = tick_response.json()
    assert "Drive root temporarily unavailable" in payload["last_sync_message"]
    entry = payload["entries"][0]
    assert entry["workload_status"] == "processed"
    assert entry["trainees_status"] == "processed"
    assert entry["trainee_count"] == 1


def test_background_tick_clears_all_pending_workloads_for_resynced_journals(
    client,
    auth_headers,
    monkeypatch,
):
    folders = [
        {
            "id": "drive-1-26",
            "name": "1-26 Журнал",
            "url": "https://drive.google.com/drive/folders/drive-1-26",
            "modified_time": "2026-05-01T10:00:00Z",
        },
        {
            "id": "drive-10-26",
            "name": "10-26 Журнал",
            "url": "https://drive.google.com/drive/folders/drive-10-26",
            "modified_time": "2026-05-01T11:00:00Z",
        },
        {
            "id": "drive-11-26",
            "name": "11-26 Журнал",
            "url": "https://drive.google.com/drive/folders/drive-11-26",
            "modified_time": "2026-05-01T12:00:00Z",
        },
    ]
    monkeypatch.setattr("app.api.routes.journal_monitors.list_drive_child_folders", lambda _folder_id, service_account_json=None: folders)
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": f"{folder_id}.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_combined_workbook_bytes(),
        raising=False,
    )
    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]

    tick_response = client.post(
        f"/api/v1/journal-monitors/{section_id}/processing/background-tick?year=2026&sync=true&workload_limit=20&trainees_limit=20",
        headers=auth_headers,
    )

    assert tick_response.status_code == 200
    entries = {entry["group_code"]: entry for entry in tick_response.json()["entries"]}
    assert {entries[code]["workload_status"] for code in ("1-26", "10-26", "11-26")} == {"processed"}
    assert {entries[code]["trainees_status"] for code in ("1-26", "10-26", "11-26")} == {"processed"}


def test_delete_journal_entry_archives_group_trainees_and_resync_restores_them(client, auth_headers, db_session, monkeypatch):
    group = Group(branch_id="main", code="46-26", name="Журнал із листів", status=GroupStatus.ACTIVE)
    db_session.add(group)
    db_session.flush()
    _seed_schedule(db_session, group, "46")
    db_session.add(Trainee(branch_id="main", first_name="Іван", last_name="Петренко", status="active", group_code="46-26"))
    db_session.commit()
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-46-26",
                "name": "46-26 Журнал із листів",
                "url": "https://drive.google.com/drive/folders/drive-46-26",
                "modified_time": "2026-05-01T10:00:00Z",
            }
        ],
    )
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes(
            [
                (1, "З-СНН-001", "Петренко Іван Іванович", "ч", "01.02.1990", "1234567890", "м. Львів", ""),
                (2, "З-СНН-002", "Коваль Олена Петрівна", "ж", "03.04.1992", "0987654321", "м. Київ", ""),
            ]
        ),
        raising=False,
    )
    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)
    entry_id = sync_response.json()["entries"][0]["id"]

    delete_response = client.delete(f"/api/v1/journal-monitors/{section_id}/entries/{entry_id}", headers=auth_headers)

    assert delete_response.status_code == 204
    assert "46-26" not in {item["code"] for item in client.get("/api/v1/groups", headers=auth_headers).json()}
    db_session.refresh(group)
    assert group.hidden_from_registry is True
    assert db_session.query(ScheduleSlot).filter(ScheduleSlot.group_id == group.id).count() == 1
    assert db_session.query(Trainee).filter(Trainee.group_code == "46-26", Trainee.is_deleted.is_(False)).count() == 0
    assert db_session.query(Trainee).filter(Trainee.is_deleted.is_(True)).count() == 1

    resync_response = client.post(
        f"/api/v1/journal-monitors/{section_id}/processing/background-tick?year=2026&sync=true",
        headers=auth_headers,
    )

    assert resync_response.status_code == 200
    db_session.refresh(group)
    assert group.hidden_from_registry is False
    groups_payload = {item["code"]: item for item in client.get("/api/v1/groups", headers=auth_headers).json()}
    assert groups_payload["46-26"]["year"] == 2026
    assert db_session.query(Trainee).filter(Trainee.group_code == "46-26", Trainee.is_deleted.is_(False)).count() == 2


def test_processed_journal_reimports_trainees_when_group_roster_was_deleted(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    entry = journal_monitor.JournalMonitorEntry(
        section=section,
        branch_id="main",
        drive_file_id="drive-46-26",
        journal_name="46-26 Журнал",
        group_code="46-26",
        workload_status="processed",
        workload_year=2026,
        trainees_status="processed",
        trainees_message="Додано/оновлено слухачів із журналу: 2",
        trainee_count=2,
        has_trainees=True,
    )
    db_session.add_all(
        [
            section,
            entry,
            Trainee(
                branch_id="main",
                first_name="Іван Іванович",
                last_name="Петренко",
                contract_number="З-СНН-001",
                status="active",
                is_deleted=True,
                deleted_at=datetime.now(timezone.utc),
            ),
            Trainee(
                branch_id="main",
                first_name="Олена Петрівна",
                last_name="Коваль",
                contract_number="З-СНН-002",
                status="active",
                is_deleted=True,
                deleted_at=datetime.now(timezone.utc),
            ),
        ]
    )
    db_session.commit()

    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": f"{folder_id}-xlsx", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes(
            [
                (1, "З-СНН-001", "Петренко Іван Іванович", "ч", "01.02.1990", "1234567890", "м. Львів", ""),
                (2, "З-СНН-002", "Коваль Олена Петрівна", "ж", "03.04.1992", "0987654321", "м. Київ", ""),
            ]
        ),
        raising=False,
    )

    result = journal_monitor.process_journal_trainees_for_section(db_session, section, limit=None, target_year=2026)

    assert result["processed"] == 1
    assert db_session.query(Trainee).filter(Trainee.group_code == "46-26", Trainee.is_deleted.is_(False)).count() == 2
    assert db_session.query(Trainee).count() == 2
    db_session.refresh(entry)
    assert entry.trainees_status == "processed"
    assert entry.trainee_count == 2


def test_drive_sync_cleanup_hides_group_and_archives_trainees_when_folder_removed(db_session):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    group = Group(branch_id="main", code="46-26", name="46-26 Журнал", status=GroupStatus.ACTIVE)
    trainee = Trainee(branch_id="main", first_name="Іван Іванович", last_name="Петренко", status="active", group_code="46-26")
    teacher = Teacher(branch_id="main", first_name="Олена Петрівна", last_name="Коваль", annual_load_hours=100, is_active=True)
    db_session.add_all([section, group, trainee, teacher])
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-46-26",
        journal_name="46-26 Журнал",
        group_code="46-26",
        trainees_status="processed",
        workload_status="processed",
        workload_year=2026,
        workload_hours=8,
    )
    db_session.add(entry)
    db_session.flush()
    db_session.add(
        JournalWorkloadEntry(
            journal_monitor_entry_id=entry.id,
            branch_id="main",
            teacher_id=teacher.id,
            subject_name="Журнальна дисципліна",
            hours=8,
        )
    )
    db_session.commit()

    before = {row["teacher_id"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert before[teacher.id]["total_hours"] == 8

    journal_monitor.sync_journal_monitor_section(
        db_session,
        section,
        folder_lister=lambda _folder_id, service_account_json=None: [],
        process_workload=False,
        process_trainees=False,
    )
    db_session.commit()

    db_session.refresh(group)
    db_session.refresh(trainee)
    assert group.hidden_from_registry is True
    assert trainee.is_deleted is True
    assert trainee.group_code is None
    assert db_session.query(journal_monitor.JournalMonitorEntry).count() == 0
    assert db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.teacher_id == teacher.id).count() == 0
    after = {row["teacher_id"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert after[teacher.id]["total_hours"] == 0
    assert after[teacher.id]["remaining_hours"] == 100


def test_drive_sync_renames_group_and_requeues_processed_entry_when_folder_changes(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    group = Group(branch_id="main", code="46-26", name="Стара назва", status=GroupStatus.ACTIVE)
    db_session.add_all([section, group])
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-46-26",
        journal_name="46-26 Стара назва",
        group_code="46-26",
        trainees_status="processed",
        workload_status="processed",
        workload_processed_at=datetime(2026, 5, 1, 10, 0, tzinfo=timezone.utc),
        trainees_processed_at=datetime(2026, 5, 1, 10, 0, tzinfo=timezone.utc),
        drive_modified_at=datetime(2026, 5, 1, 9, 0, tzinfo=timezone.utc),
    )
    db_session.add(entry)
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {
                "id": "sheet-46-26",
                "name": "46-26.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "modifiedTime": "2026-05-02T10:00:00Z",
            }
        ],
        raising=False,
    )

    journal_monitor.sync_journal_monitor_section(
        db_session,
        section,
        folder_lister=lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-46-26",
                "name": "46-26 Нова назва",
                "url": "https://drive.google.com/drive/folders/drive-46-26",
                "modified_time": "2026-05-02T10:00:00Z",
            }
        ],
        process_workload=False,
        process_trainees=False,
    )
    db_session.commit()

    db_session.refresh(group)
    db_session.refresh(entry)
    assert group.name == "Нова назва"
    assert entry.journal_name == "46-26 Нова назва"
    assert entry.workload_status == "pending"
    assert entry.trainees_status == "pending"


def test_drive_sync_keeps_processed_workload_when_only_folder_metadata_changes(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Journals 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    group = Group(branch_id="main", code="46-26", name="Old name", status=GroupStatus.ACTIVE)
    teacher = Teacher(branch_id="main", first_name="Ivan", last_name="Teacher", is_active=True)
    db_session.add_all([section, group, teacher])
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-46-26",
        journal_name="46-26 Old name",
        group_code="46-26",
        trainees_status="processed",
        workload_status="processed",
        workload_year=2026,
        workload_processed_at=datetime(2026, 5, 1, 10, 0, tzinfo=timezone.utc),
        trainees_processed_at=datetime(2026, 5, 1, 10, 0, tzinfo=timezone.utc),
        drive_modified_at=datetime(2026, 5, 1, 9, 0, tzinfo=timezone.utc),
        workload_hours=4,
        workload_source_names=["46-26.xlsx"],
        trainees_source_names=["46-26.xlsx"],
    )
    db_session.add(entry)
    db_session.flush()
    db_session.add(
        JournalWorkloadEntry(
            journal_monitor_entry_id=entry.id,
            branch_id="main",
            teacher_id=teacher.id,
            subject_name="Existing subject",
            hours=4,
        )
    )
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {
                "id": "sheet-46-26",
                "name": "46-26.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "modifiedTime": "2026-05-01T09:30:00Z",
            }
        ],
        raising=False,
    )

    journal_monitor.sync_journal_monitor_section(
        db_session,
        section,
        folder_lister=lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-46-26",
                "name": "46-26 New name",
                "url": "https://drive.google.com/drive/folders/drive-46-26",
                "modified_time": "2026-05-02T10:00:00Z",
            }
        ],
        process_workload=False,
        process_trainees=False,
    )
    db_session.commit()

    db_session.refresh(group)
    db_session.refresh(entry)
    workload_rows = db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).all()
    summary = {row["teacher_id"]: row for row in collect_teacher_workload_summary(db_session, "main")}
    assert group.name == "New name"
    assert entry.journal_name == "46-26 New name"
    assert entry.workload_status == "processed"
    assert entry.trainees_status == "processed"
    assert entry.workload_hours == 4
    assert [(row.subject_name, row.hours) for row in workload_rows] == [("Existing subject", 4)]
    assert summary[teacher.id]["total_hours"] == 4


def test_journal_daily_activity_lists_created_and_changed_since_8_kyiv(db_session):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    db_session.add(section)
    db_session.flush()
    db_session.add_all(
        [
            journal_monitor.JournalMonitorEntry(
                section_id=section.id,
                branch_id="main",
                drive_file_id="created-today",
                journal_name="46-26 Новий журнал",
                group_code="46-26",
                drive_created_at=datetime(2026, 5, 13, 5, 30, tzinfo=timezone.utc),
                drive_modified_at=datetime(2026, 5, 13, 5, 30, tzinfo=timezone.utc),
            ),
            journal_monitor.JournalMonitorEntry(
                section_id=section.id,
                branch_id="main",
                drive_file_id="changed-today",
                journal_name="47-26 Змінений журнал",
                group_code="47-26",
                drive_created_at=datetime(2026, 5, 12, 7, 0, tzinfo=timezone.utc),
                drive_modified_at=datetime(2026, 5, 13, 8, 45, tzinfo=timezone.utc),
                drive_change_started_at=datetime(2026, 5, 13, 6, 10, tzinfo=timezone.utc),
            ),
            journal_monitor.JournalMonitorEntry(
                section_id=section.id,
                branch_id="main",
                drive_file_id="before-cutoff",
                journal_name="48-26 Ранній журнал",
                group_code="48-26",
                drive_created_at=datetime(2026, 5, 13, 4, 59, tzinfo=timezone.utc),
                drive_modified_at=datetime(2026, 5, 13, 4, 59, tzinfo=timezone.utc),
                drive_change_started_at=datetime(2026, 5, 13, 4, 59, tzinfo=timezone.utc),
            ),
        ]
    )
    db_session.commit()
    db_session.refresh(section)

    activity = journal_monitor.collect_daily_journal_activity(
        section,
        now=datetime(2026, 5, 13, 9, 0, tzinfo=timezone.utc),
    )

    assert activity["cutoff_at"] == datetime(2026, 5, 13, 5, 0, tzinfo=timezone.utc)
    assert [item["journal_name"] for item in activity["created"]] == ["46-26 Новий журнал"]
    assert activity["created"][0]["created_at"] == datetime(2026, 5, 13, 5, 30, tzinfo=timezone.utc)
    assert [item["journal_name"] for item in activity["changed"]] == ["47-26 Змінений журнал"]
    assert activity["changed"][0]["change_started_at"] == datetime(2026, 5, 13, 6, 10, tzinfo=timezone.utc)
    assert activity["changed"][0]["modified_at"] == datetime(2026, 5, 13, 8, 45, tzinfo=timezone.utc)


def test_journal_sync_tracks_drive_created_time_and_daily_change_start(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    db_session.add(section)
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {
                "id": "sheet-49-26",
                "name": "49-26 Журнал.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "modifiedTime": "2026-05-13T06:20:00Z",
            }
        ],
        raising=False,
    )

    journal_monitor.sync_journal_monitor_section(
        db_session,
        section,
        folder_lister=lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-49-26",
                "name": "49-26 Денний журнал",
                "url": "https://drive.google.com/drive/folders/drive-49-26",
                "created_time": "2026-05-13T05:15:00Z",
                "modified_time": "2026-05-13T06:20:00Z",
            }
        ],
        process_workload=False,
        process_trainees=False,
    )
    db_session.commit()

    entry = db_session.query(journal_monitor.JournalMonitorEntry).one()
    assert journal_monitor._as_aware_utc(entry.drive_created_at) == datetime(2026, 5, 13, 5, 15, tzinfo=timezone.utc)
    assert journal_monitor._as_aware_utc(entry.drive_change_started_at) == datetime(2026, 5, 13, 6, 20, tzinfo=timezone.utc)


def test_reprocessing_journal_trainees_updates_changed_phone_and_archives_removed_rows(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
    )
    db_session.add(section)
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-46-26",
        journal_name="46-26 Журнал",
        group_code="46-26",
    )
    db_session.add_all(
        [
            entry,
            Trainee(
                branch_id="main",
                first_name="Іван Іванович",
                last_name="Петренко",
                birth_date=datetime(1990, 2, 1).date(),
                status="active",
                group_code="46-26",
                phone_encrypted=cipher.encrypt("+380000000000"),
            ),
            Trainee(
                branch_id="main",
                first_name="Олена Петрівна",
                last_name="Коваль",
                birth_date=datetime(1992, 4, 3).date(),
                status="active",
                group_code="46-26",
            ),
        ]
    )
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {"id": "zv-xlsx", "name": "46-26 Журнал.xlsx", "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_zv_workbook_bytes(
            [
                (1, "З-СНН-001", "Петренко Іван Іванович", "ч", "01.02.1990", "1234567890", "м. Львів", "+380501112233"),
            ]
        ),
        raising=False,
    )

    result = journal_monitor.process_journal_trainees_entry(db_session, entry)
    db_session.commit()

    assert result["changed"] >= 1
    active_trainees = db_session.query(Trainee).filter(Trainee.is_deleted.is_(False)).all()
    archived_trainees = db_session.query(Trainee).filter(Trainee.is_deleted.is_(True)).all()
    assert [(item.last_name, item.group_code) for item in active_trainees] == [("Петренко", "46-26")]
    assert [item.last_name for item in archived_trainees] == ["Коваль"]
    assert cipher.decrypt(active_trainees[0].phone_encrypted) == "+380501112233"


def test_processed_workload_reimports_when_workbook_modified_time_is_newer(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_year=2026,
    )
    teacher = Teacher(branch_id="main", first_name="Іванович", last_name="Старий", is_active=True)
    db_session.add_all([section, teacher])
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-46-26",
        journal_name="46-26 Журнал",
        group_code="46-26",
        workload_status="processed",
        workload_year=2026,
        workload_processed_at=datetime(2026, 5, 1, 10, 0, tzinfo=timezone.utc),
        workload_hours=4,
    )
    db_session.add(entry)
    db_session.flush()
    db_session.add(
        JournalWorkloadEntry(
            journal_monitor_entry_id=entry.id,
            branch_id="main",
            teacher_id=teacher.id,
            subject_name="Старий предмет",
            hours=4,
        )
    )
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {
                "id": "disciplines-xlsx",
                "name": "46-26 Журнал.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "modifiedTime": "2026-05-02T10:00:00Z",
            }
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_workbook_bytes(
            [("Новий предмет", 10, "1-10", "Новий Викладач")]
        ),
        raising=False,
    )

    result = journal_monitor.process_next_journal_workload(db_session, section, limit=1, target_year=2026)
    db_session.commit()

    db_session.refresh(entry)
    assert result["processed"] == 1
    assert entry.workload_status == "processed"
    assert entry.workload_hours == 10
    rows = db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).all()
    assert [(row.subject_name, row.hours) for row in rows] == [("Новий предмет", 10)]


def test_no_data_workload_retries_without_newer_modified_time(db_session, monkeypatch):
    section = journal_monitor.JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root-folder",
        folder_id="root-folder",
        workload_auto_year=2026,
    )
    db_session.add(section)
    db_session.flush()
    entry = journal_monitor.JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="drive-85-26",
        journal_name="85-26 Журнал",
        group_code="85-26",
        workload_status="no_data",
        workload_year=2026,
        workload_processed_at=datetime(2026, 5, 2, 10, 0, tzinfo=timezone.utc),
        workload_hours=30,
    )
    db_session.add(entry)
    db_session.commit()
    monkeypatch.setattr(
        journal_monitor,
        "list_drive_journal_workbook_files",
        lambda folder_id, service_account_json=None: [
            {
                "id": "disciplines-xlsx",
                "name": "85-26 Журнал.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "modifiedTime": "2026-05-01T10:00:00Z",
            }
        ],
        raising=False,
    )
    monkeypatch.setattr(
        journal_monitor,
        "download_drive_file_bytes",
        lambda file_id, mime_type=None, service_account_json=None: _journal_workbook_bytes(
            [("Штучний інтелект", 30, "3-19", "Коваль Олена Петрівна")]
        ),
        raising=False,
    )

    result = journal_monitor.process_next_journal_workload(db_session, section, limit=1, target_year=2026)
    db_session.commit()

    db_session.refresh(entry)
    assert result["processed"] == 1
    assert entry.workload_status == "processed"
    assert entry.workload_hours == 30
    rows = db_session.query(JournalWorkloadEntry).filter(JournalWorkloadEntry.journal_monitor_entry_id == entry.id).all()
    assert [(row.subject_name, row.hours) for row in rows] == [("Штучний інтелект", 30)]


def test_bulk_delete_journal_entries_hides_related_groups(client, auth_headers, db_session, monkeypatch):
    db_session.add_all(
        [
            Trainee(branch_id="main", first_name="Іван", last_name="Перший", status="active", group_code="31-26"),
            Trainee(branch_id="main", first_name="Олена", last_name="Друга", status="active", group_code="32-26"),
        ]
    )
    db_session.commit()
    monkeypatch.setattr(
        "app.api.routes.journal_monitors.list_drive_child_folders",
        lambda _folder_id, service_account_json=None: [
            {
                "id": "drive-31-26",
                "name": "31-26 Перша група",
                "url": "https://drive.google.com/drive/folders/drive-31-26",
                "modified_time": "2026-05-01T10:00:00Z",
            },
            {
                "id": "drive-32-26",
                "name": "32-26 Друга група",
                "url": "https://drive.google.com/drive/folders/drive-32-26",
                "modified_time": "2026-05-01T11:00:00Z",
            },
        ],
    )
    create_response = client.post(
        "/api/v1/journal-monitors",
        json={"name": "Журнали 2026", "folder_url": "https://drive.google.com/drive/folders/root-folder"},
        headers=auth_headers,
    )
    section_id = create_response.json()["id"]
    sync_response = client.post(f"/api/v1/journal-monitors/{section_id}/sync", headers=auth_headers)
    entry_ids = [item["id"] for item in sync_response.json()["entries"]]

    delete_response = client.post(
        f"/api/v1/journal-monitors/{section_id}/entries/bulk-delete",
        json={"entry_ids": entry_ids},
        headers=auth_headers,
    )

    assert delete_response.status_code == 200
    assert delete_response.json()["deleted_count"] == 2
    assert set(delete_response.json()["deleted_ids"]) == set(entry_ids)
    assert {group.code for group in db_session.query(Group).filter(Group.hidden_from_registry.is_(True)).all()} == {"31-26", "32-26"}
    assert db_session.query(Trainee).filter(Trainee.group_code.in_(["31-26", "32-26"]), Trainee.is_deleted.is_(False)).count() == 0
    assert db_session.query(Trainee).filter(Trainee.is_deleted.is_(True)).count() == 2
    assert client.get(f"/api/v1/journal-monitors/{section_id}", headers=auth_headers).json()["entries"] == []
