from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.models import (
    Document,
    DocumentType,
    DraftStatus,
    Group,
    GroupMembership,
    GroupStatus,
    ImportJob,
    JournalMonitorEntry,
    JournalMonitorSection,
    JournalWorkloadEntry,
    JobStatus,
    MembershipStatus,
    OCRResult,
    Performance,
    Room,
    ScheduleSlot,
    Subject,
    Teacher,
    Trainee,
)


def test_auth_login_and_me(client):
    login_response = client.post("/api/v1/auth/login", json={"username": "admin", "password": "Admin123!"})
    assert login_response.status_code == 200
    payload = login_response.json()
    assert payload["access_token"]
    assert payload["refresh_token"]

    me_response = client.get(
        "/api/v1/auth/me",
        headers={"Authorization": f"Bearer {payload['access_token']}"},
    )
    assert me_response.status_code == 200
    me_payload = me_response.json()
    role_names = [role["name"] for role in me_payload["roles"]]
    assert "admin" in role_names


def test_group_trainee_enrollment_flow(client, auth_headers):
    trainee_response = client.post(
        "/api/v1/trainees",
        json={"first_name": "Марина", "last_name": "Іваненко", "status": "active"},
        headers=auth_headers,
    )
    assert trainee_response.status_code == 201
    trainee_id = trainee_response.json()["id"]

    group_response = client.post(
        "/api/v1/groups",
        json={"code": "GRP-001", "name": "Тестова група", "capacity": 25, "status": "planned"},
        headers=auth_headers,
    )
    assert group_response.status_code == 201
    group_id = group_response.json()["id"]

    enroll_response = client.post(
        f"/api/v1/groups/{group_id}/enroll",
        json={"trainee_id": trainee_id},
        headers=auth_headers,
    )
    assert enroll_response.status_code == 201
    assert enroll_response.json()["status"] == "active"


def test_group_audit_returns_group_actions(client, auth_headers):
    trainee_response = client.post(
        "/api/v1/trainees",
        json={"first_name": "Олег", "last_name": "Петренко", "status": "active"},
        headers=auth_headers,
    )
    assert trainee_response.status_code == 201
    trainee_id = trainee_response.json()["id"]

    group_response = client.post(
        "/api/v1/groups",
        json={"code": "AUD-001", "name": "Група з історією", "capacity": 25, "status": "planned"},
        headers=auth_headers,
    )
    assert group_response.status_code == 201
    group_id = group_response.json()["id"]

    enroll_response = client.post(
        f"/api/v1/groups/{group_id}/enroll",
        json={"trainee_id": trainee_id},
        headers=auth_headers,
    )
    assert enroll_response.status_code == 201

    audit_response = client.get(f"/api/v1/groups/{group_id}/audit", headers=auth_headers)
    assert audit_response.status_code == 200
    actions = [item["action"] for item in audit_response.json()]
    assert "group.create" in actions
    assert "group.enroll" in actions


def test_groups_api_falls_back_to_schedule_dates_when_group_dates_are_empty(client, auth_headers, db_session):
    group = Group(branch_id="main", code="46-26", name="Група з розкладом", status=GroupStatus.ACTIVE)
    teacher = Teacher(branch_id="main", first_name="Тест", last_name="Викладач", hourly_rate=0, is_active=True)
    subject = Subject(branch_id="main", name="Предмет", hours_total=4)
    room = Room(branch_id="main", name="Аудиторія", capacity=20)
    db_session.add_all([group, teacher, subject, room])
    db_session.flush()
    db_session.add_all(
        [
            ScheduleSlot(
                group_id=group.id,
                teacher_id=teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=datetime(2026, 3, 11, 9, 30, tzinfo=timezone.utc),
                ends_at=datetime(2026, 3, 11, 11, 5, tzinfo=timezone.utc),
                pair_number=1,
                academic_hours=2,
            ),
            ScheduleSlot(
                group_id=group.id,
                teacher_id=teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=datetime(2026, 3, 12, 11, 10, tzinfo=timezone.utc),
                ends_at=datetime(2026, 3, 12, 12, 45, tzinfo=timezone.utc),
                pair_number=2,
                academic_hours=2,
            ),
        ]
    )
    db_session.commit()

    list_response = client.get("/api/v1/groups", headers=auth_headers)
    assert list_response.status_code == 200
    listed_group = next(item for item in list_response.json() if item["code"] == "46-26")
    assert listed_group["start_date"] == "2026-03-11"
    assert listed_group["end_date"] == "2026-03-12"

    detail_response = client.get(f"/api/v1/groups/{group.id}", headers=auth_headers)
    assert detail_response.status_code == 200
    assert detail_response.json()["start_date"] == "2026-03-11"
    assert detail_response.json()["end_date"] == "2026-03-12"


def test_schedule_workload_and_kpi_flow(client, auth_headers):
    teacher_response = client.post(
        "/api/v1/teachers",
        json={"first_name": "Тест", "last_name": "Викладач", "hourly_rate": 0, "annual_load_hours": 100, "is_active": True},
        headers=auth_headers,
    )
    assert teacher_response.status_code == 201

    group_response = client.post(
        "/api/v1/groups",
        json={"code": "GRP-002", "name": "Група для розкладу", "capacity": 20, "status": "active"},
        headers=auth_headers,
    )
    assert group_response.status_code == 201

    schedule_response = client.post(
        "/api/v1/schedule/generate",
        json={"start_date": date.today().isoformat(), "days": 3},
        headers=auth_headers,
    )
    assert schedule_response.status_code == 200
    assert len(schedule_response.json()) > 0

    workload_response = client.get("/api/v1/teacher-workload", headers=auth_headers)
    assert workload_response.status_code == 200
    assert isinstance(workload_response.json(), list)

    kpi_response = client.get("/api/v1/dashboard/kpi", headers=auth_headers)
    assert kpi_response.status_code == 200
    kpi_payload = kpi_response.json()
    assert kpi_payload["active_groups"] >= 1
    assert "facility_load_pct" not in kpi_payload


def test_teacher_workload_summary_export_is_single_printable_sheet(client, auth_headers, db_session):
    first_teacher = Teacher(
        branch_id="main",
        first_name="Галина Михайлівна",
        last_name="Войтехівська",
        hourly_rate=0,
        annual_load_hours=180,
        is_active=True,
    )
    second_teacher = Teacher(
        branch_id="main",
        first_name="Роман Йосипович",
        last_name="Бубняк",
        hourly_rate=0,
        annual_load_hours=0,
        is_active=True,
    )
    group = Group(branch_id="main", code="PRINT-001", name="Група для друку", status=GroupStatus.ACTIVE)
    subject = Subject(branch_id="main", name="Предмет для друку", hours_total=20)
    room = Room(branch_id="main", name="Аудиторія друку", capacity=20)
    db_session.add_all([first_teacher, second_teacher, group, subject, room])
    db_session.flush()
    starts_at = datetime(2026, 4, 1, 9, 30, tzinfo=timezone.utc)
    db_session.add_all(
        [
            ScheduleSlot(
                group_id=group.id,
                teacher_id=first_teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=starts_at,
                ends_at=starts_at + timedelta(hours=2),
                pair_number=1,
                academic_hours=10,
            ),
            ScheduleSlot(
                group_id=group.id,
                teacher_id=second_teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=starts_at,
                ends_at=starts_at + timedelta(hours=2),
                pair_number=2,
                academic_hours=4,
            ),
        ]
    )
    db_session.commit()

    response = client.get("/api/v1/teacher-workload/export-summary", headers=auth_headers)

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["Педнавантаження"]
    sheet = workbook["Педнавантаження"]
    assert isinstance(sheet["A1"].value, str)
    assert sheet["A1"].value.startswith("Дата формування:")
    assert [cell.value for cell in sheet[3]] == ["Викладач", "Поточні години", "Річний план", "Залишок годин"]
    values = {
        row[0]: row[1:]
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, values_only=True)
        if row[0]
    }
    assert values["Войтехівська Галина Михайлівна"] == (10, 180, 170)
    assert values["Бубняк Роман Йосипович"] == (4, 0, -4)


def test_dashboard_kpi_excludes_archived_trainees_from_active_count(client, auth_headers, db_session):
    group = Group(branch_id="main", code="KPI-ARCH", name="KPI архів", status=GroupStatus.ACTIVE)
    visible = Trainee(branch_id="main", first_name="Активний", last_name="Слухач", status="active")
    archived = Trainee(
        branch_id="main",
        first_name="Архівний",
        last_name="Слухач",
        status="active",
        is_deleted=True,
        deleted_at=datetime.now(timezone.utc),
    )
    db_session.add_all([group, visible, archived])
    db_session.flush()
    db_session.add_all(
        [
            GroupMembership(group_id=group.id, trainee_id=visible.id, status=MembershipStatus.ACTIVE),
            GroupMembership(group_id=group.id, trainee_id=archived.id, status=MembershipStatus.ACTIVE),
        ]
    )
    db_session.commit()

    response = client.get("/api/v1/dashboard/kpi", headers=auth_headers)

    assert response.status_code == 200
    assert response.json()["active_trainees"] == 1


def test_dashboard_kpi_excludes_hidden_groups_from_active_count(client, auth_headers, db_session):
    visible_group = Group(branch_id="main", code="KPI-VISIBLE", name="Видима група", status=GroupStatus.ACTIVE)
    hidden_group = Group(
        branch_id="main",
        code="KPI-HIDDEN",
        name="Прихована група",
        status=GroupStatus.ACTIVE,
        hidden_from_registry=True,
    )
    db_session.add_all([visible_group, hidden_group])
    db_session.commit()

    response = client.get("/api/v1/dashboard/kpi", headers=auth_headers)

    assert response.status_code == 200
    assert response.json()["active_groups"] == 1


def test_dashboard_kpi_uses_current_year_journal_count_when_section_exists(client, auth_headers, db_session):
    groups = [
        Group(branch_id="main", code=f"KPI-JOURNAL-{index}", name=f"Група {index}", status=GroupStatus.ACTIVE)
        for index in range(3)
    ]
    section = JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root",
        folder_id="root",
    )
    db_session.add_all([*groups, section])
    db_session.flush()
    db_session.add_all(
        [
            JournalMonitorEntry(
                section_id=section.id,
                branch_id="main",
                drive_file_id="journal-1",
                journal_name="1-26 Перша група",
                group_code="1-26",
            ),
            JournalMonitorEntry(
                section_id=section.id,
                branch_id="main",
                drive_file_id="journal-2",
                journal_name="2-26 Друга група",
                group_code="2-26",
            ),
        ]
    )
    db_session.commit()

    response = client.get("/api/v1/dashboard/kpi?year=2026", headers=auth_headers)

    assert response.status_code == 200
    assert response.json()["active_groups"] == 2


def test_dashboard_student_plan_uses_processed_group_trainees(client, auth_headers, db_session):
    for index in range(24):
        db_session.add(
            Trainee(
                branch_id="main",
                first_name=f"Слухач{index}",
                last_name="План",
                status="active",
                group_code="180-25",
            )
        )
    db_session.add(Trainee(branch_id="main", first_name="Без", last_name="Групи", status="active"))
    db_session.add(
        Trainee(
            branch_id="main",
            first_name="Архів",
            last_name="Не рахуємо",
            status="active",
            group_code="180-25",
            is_deleted=True,
            deleted_at=datetime.now(timezone.utc),
        )
    )
    db_session.commit()

    plan_response = client.put(
        "/api/v1/dashboard/student-plan",
        json={"year": 2026, "target_trainees": 100},
        headers=auth_headers,
    )
    assert plan_response.status_code == 200
    assert plan_response.json()["target_trainees"] == 100

    response = client.get("/api/v1/dashboard/kpi?year=2026", headers=auth_headers)
    assert response.status_code == 200
    payload = response.json()
    assert payload["student_plan_year"] == 2026
    assert payload["student_plan_target"] == 100
    assert payload["student_plan_processed"] == 24
    assert payload["training_plan_progress_pct"] == 24


def test_schedule_generation_does_not_require_auditoriums(client, auth_headers, db_session):
    db_session.query(Room).delete()
    db_session.add(Teacher(branch_id="main", first_name="Без", last_name="Аудиторій", hourly_rate=0, is_active=True))
    db_session.add(Group(branch_id="main", code="REMOTE-1", name="Дистанційна група", status=GroupStatus.ACTIVE))
    db_session.add(Subject(branch_id="main", name="Дистанційний предмет", hours_total=12))
    db_session.commit()

    response = client.post(
        "/api/v1/schedule/generate",
        json={"start_date": date.today().isoformat(), "days": 1},
        headers=auth_headers,
    )

    assert response.status_code == 200
    assert len(response.json()) >= 1


def test_dashboard_attention_collects_actionable_items(client, auth_headers, db_session):
    document = Document(
        branch_id="main",
        file_name="attention.docx",
        file_path="tmp/attention.docx",
        file_type=DocumentType.DOCX,
        source="upload",
    )
    scheduled_group = Group(branch_id="main", code="OK-1", name="З розкладом", status=GroupStatus.ACTIVE)
    group_without_schedule = Group(branch_id="main", code="NO-SCHEDULE", name="Без розкладу", status=GroupStatus.ACTIVE)
    teacher = Teacher(branch_id="main", first_name="Тест", last_name="Викладач", hourly_rate=0, is_active=True)
    subject = Subject(branch_id="main", name="Предмет", hours_total=2)
    room = Room(branch_id="main", name="Аудиторія", capacity=20)
    db_session.add_all([document, scheduled_group, group_without_schedule, teacher, subject, room])
    db_session.flush()
    db_session.add_all(
        [
            ScheduleSlot(
                group_id=scheduled_group.id,
                teacher_id=teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=datetime.now(timezone.utc),
                ends_at=datetime.now(timezone.utc) + timedelta(hours=2),
                pair_number=1,
                academic_hours=2,
            ),
            ImportJob(
                branch_id="main",
                idempotency_key="main:attention-failed",
                document_id=document.id,
                status=JobStatus.FAILED,
                message="Помилка тесту",
            ),
            OCRResult(
                branch_id="main",
                document_id=document.id,
                extracted_text="draft",
                status=DraftStatus.PENDING,
                confidence=0.7,
            ),
            Trainee(branch_id="main", first_name="Без", last_name="Групи", status="active"),
            Trainee(branch_id="main", first_name="Сирітський", last_name="Код", status="active", group_code="MISSING"),
        ]
    )
    db_session.commit()

    response = client.get("/api/v1/dashboard/attention", headers=auth_headers)

    assert response.status_code == 200
    payload = response.json()
    items = {item["key"]: item for item in payload["items"]}
    assert items["failed_jobs"]["count"] == 1
    assert items["pending_drafts"]["count"] == 1
    assert items["unassigned_trainees"]["count"] == 1
    assert items["orphan_group_codes"]["count"] == 1
    assert items["groups_without_schedule"]["count"] == 1
    assert payload["total_count"] == 5


def test_active_groups_between_dates_and_excel_export(client, auth_headers, db_session):
    group = Group(
        branch_id="main",
        code="167-25",
        name="Організація трудових відносин",
        capacity=25,
        status=GroupStatus.ACTIVE,
        start_date=date(2025, 10, 21),
        end_date=date(2025, 10, 24),
    )
    other_group = Group(branch_id="main", code="999-25", name="Поза періодом", capacity=25, status=GroupStatus.ACTIVE)
    teacher_one = Teacher(branch_id="main", first_name="Лілія", last_name="Штогрин", hourly_rate=0, is_active=True)
    teacher_two = Teacher(branch_id="main", first_name="Артур", last_name="Костів", hourly_rate=0, is_active=True)
    subject = Subject(branch_id="main", name="Трудовий договір", hours_total=10)
    room = Room(branch_id="main", name="Імпорт: 167-25", capacity=25)
    db_session.add_all([group, other_group, teacher_one, teacher_two, subject, room])
    db_session.flush()
    db_session.add_all(
        [
            ScheduleSlot(
                group_id=group.id,
                teacher_id=teacher_one.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=datetime(2025, 10, 21, 9, 30, tzinfo=timezone.utc),
                ends_at=datetime(2025, 10, 21, 11, 5, tzinfo=timezone.utc),
                pair_number=1,
                academic_hours=2,
            ),
            ScheduleSlot(
                group_id=group.id,
                teacher_id=teacher_two.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=datetime(2025, 10, 22, 11, 10, tzinfo=timezone.utc),
                ends_at=datetime(2025, 10, 22, 12, 45, tzinfo=timezone.utc),
                pair_number=2,
                academic_hours=1.5,
            ),
            ScheduleSlot(
                group_id=other_group.id,
                teacher_id=teacher_one.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=datetime(2025, 11, 1, 9, 30, tzinfo=timezone.utc),
                ends_at=datetime(2025, 11, 1, 11, 5, tzinfo=timezone.utc),
                pair_number=1,
                academic_hours=2,
            ),
        ]
    )
    db_session.commit()

    response = client.get(
        "/api/v1/groups/active-between?date_from=2025-10-20&date_to=2025-10-25",
        headers=auth_headers,
    )
    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["code"] == "167-25"
    assert payload[0]["total_hours"] == 3.5
    assert {item["teacher_name"] for item in payload[0]["teachers"]} == {"Штогрин Лілія", "Костів Артур"}

    all_time_response = client.get("/api/v1/groups/active-between", headers=auth_headers)
    assert all_time_response.status_code == 200
    assert {item["code"] for item in all_time_response.json()} == {"167-25", "999-25"}

    partial_search_response = client.get(
        "/api/v1/groups/active-between?search=трудових відносин",
        headers=auth_headers,
    )
    assert partial_search_response.status_code == 200
    partial_search_payload = partial_search_response.json()
    assert len(partial_search_payload) == 1
    assert partial_search_payload[0]["code"] == "167-25"

    export_response = client.get(
        "/api/v1/groups/active-between/export?date_from=2025-10-20&date_to=2025-10-25",
        headers=auth_headers,
    )
    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert export_response.content


def test_bulk_group_code_update_flow(client, auth_headers):
    first = client.post(
        "/api/v1/trainees",
        json={"first_name": "Іван", "last_name": "Перший", "status": "active"},
        headers=auth_headers,
    )
    second = client.post(
        "/api/v1/trainees",
        json={"first_name": "Олена", "last_name": "Друга", "status": "active"},
        headers=auth_headers,
    )
    assert first.status_code == 201
    assert second.status_code == 201
    trainee_ids = [first.json()["id"], second.json()["id"]]

    bulk_response = client.post(
        "/api/v1/trainees/bulk/group-code",
        json={"trainee_ids": trainee_ids, "group_code": "73-26"},
        headers=auth_headers,
    )
    assert bulk_response.status_code == 200
    assert bulk_response.json()["updated_count"] == 2
    assert bulk_response.json()["group_code"] == "73-26"

    trainees_response = client.get("/api/v1/trainees", headers=auth_headers)
    assert trainees_response.status_code == 200
    rows = trainees_response.json()
    updated = [item for item in rows if item["id"] in trainee_ids]
    assert len(updated) == 2
    assert all(item["group_code"] == "73-26" for item in updated)


def test_bulk_status_update_flow(client, auth_headers):
    first = client.post(
        "/api/v1/trainees",
        json={"first_name": "Степан", "last_name": "Перший", "status": "active"},
        headers=auth_headers,
    )
    second = client.post(
        "/api/v1/trainees",
        json={"first_name": "Марія", "last_name": "Друга", "status": "active"},
        headers=auth_headers,
    )
    assert first.status_code == 201
    assert second.status_code == 201
    trainee_ids = [first.json()["id"], second.json()["id"]]

    bulk_response = client.post(
        "/api/v1/trainees/bulk/status",
        json={"trainee_ids": trainee_ids, "status": "completed"},
        headers=auth_headers,
    )
    assert bulk_response.status_code == 200
    assert bulk_response.json()["updated_count"] == 2
    assert bulk_response.json()["status"] == "completed"

    trainees_response = client.get("/api/v1/trainees", headers=auth_headers)
    assert trainees_response.status_code == 200
    rows = trainees_response.json()
    updated = [item for item in rows if item["id"] in trainee_ids]
    assert len(updated) == 2
    assert all(item["status"] == "completed" for item in updated)


def test_bulk_archive_restore_flow(client, auth_headers):
    first = client.post(
        "/api/v1/trainees",
        json={"first_name": "Анна", "last_name": "Видалити1", "status": "active"},
        headers=auth_headers,
    )
    second = client.post(
        "/api/v1/trainees",
        json={"first_name": "Петро", "last_name": "Видалити2", "status": "active"},
        headers=auth_headers,
    )
    assert first.status_code == 201
    assert second.status_code == 201
    trainee_ids = [first.json()["id"], second.json()["id"]]

    bulk_response = client.post(
        "/api/v1/trainees/bulk/delete",
        json={"trainee_ids": trainee_ids},
        headers=auth_headers,
    )
    assert bulk_response.status_code == 200
    assert bulk_response.json()["deleted_count"] == 2
    assert set(bulk_response.json()["deleted_ids"]) == set(trainee_ids)

    active_response = client.get("/api/v1/trainees", headers=auth_headers)
    assert active_response.status_code == 200
    active_rows = active_response.json()
    active_ids = {item["id"] for item in active_rows}
    assert all(trainee_id not in active_ids for trainee_id in trainee_ids)

    archived_response = client.get("/api/v1/trainees?include_deleted=true", headers=auth_headers)
    assert archived_response.status_code == 200
    archived_rows = archived_response.json()
    archived_by_id = {item["id"]: item for item in archived_rows if item["id"] in trainee_ids}
    assert len(archived_by_id) == 2
    assert all(row["is_deleted"] is True for row in archived_by_id.values())
    assert all(row["deleted_at"] is not None for row in archived_by_id.values())

    restore_response = client.post(
        "/api/v1/trainees/bulk/restore",
        json={"trainee_ids": trainee_ids},
        headers=auth_headers,
    )
    assert restore_response.status_code == 200
    assert restore_response.json()["restored_count"] == 2
    assert set(restore_response.json()["restored_ids"]) == set(trainee_ids)

    final_response = client.get("/api/v1/trainees", headers=auth_headers)
    assert final_response.status_code == 200
    final_rows = final_response.json()
    restored_by_id = {item["id"]: item for item in final_rows if item["id"] in trainee_ids}
    assert len(restored_by_id) == 2
    assert all(row["is_deleted"] is False for row in restored_by_id.values())
    assert all(row["deleted_at"] is None for row in restored_by_id.values())


def test_bulk_purge_trainees_removes_rows_and_related_history(client, auth_headers, db_session):
    group = Group(branch_id="main", code="PURGE-001", name="Група для видалення слухачів", status=GroupStatus.ACTIVE)
    db_session.add(group)
    db_session.commit()

    first = client.post(
        "/api/v1/trainees",
        json={"first_name": "Анна", "last_name": "Видалити", "status": "active", "group_code": "PURGE-001"},
        headers=auth_headers,
    )
    second = client.post(
        "/api/v1/trainees",
        json={"first_name": "Петро", "last_name": "Видалити", "status": "active", "group_code": "PURGE-001"},
        headers=auth_headers,
    )
    kept = client.post(
        "/api/v1/trainees",
        json={"first_name": "Марія", "last_name": "Залишити", "status": "active", "group_code": "PURGE-001"},
        headers=auth_headers,
    )
    assert first.status_code == 201
    assert second.status_code == 201
    assert kept.status_code == 201
    purge_ids = [first.json()["id"], second.json()["id"]]
    kept_id = kept.json()["id"]
    db_session.add_all(
        [
            GroupMembership(group_id=group.id, trainee_id=purge_ids[0], status=MembershipStatus.ACTIVE),
            GroupMembership(group_id=group.id, trainee_id=purge_ids[1], status=MembershipStatus.ACTIVE),
            Performance(branch_id="main", trainee_id=purge_ids[0], group_id=group.id, progress_pct=50, attendance_pct=80),
        ]
    )
    db_session.commit()

    purge_response = client.post(
        "/api/v1/trainees/bulk/purge",
        json={"trainee_ids": purge_ids},
        headers=auth_headers,
    )

    assert purge_response.status_code == 200
    payload = purge_response.json()
    assert payload["purged_count"] == 2
    assert set(payload["purged_ids"]) == set(purge_ids)
    assert payload["missing_ids"] == []
    assert all(db_session.get(Trainee, trainee_id) is None for trainee_id in purge_ids)
    assert db_session.get(Trainee, kept_id) is not None
    assert db_session.query(GroupMembership).filter(GroupMembership.trainee_id.in_(purge_ids)).count() == 0
    assert db_session.query(Performance).filter(Performance.trainee_id.in_(purge_ids)).count() == 0

    purge_all_response = client.post("/api/v1/trainees/bulk/purge-all", headers=auth_headers)

    assert purge_all_response.status_code == 200
    assert purge_all_response.json()["purged_count"] == 1
    assert db_session.query(Trainee).count() == 0


def test_delete_group_cleans_related_rows(client, auth_headers, db_session):
    trainee_response = client.post(
        "/api/v1/trainees",
        json={"first_name": "Оксана", "last_name": "Тест", "status": "active"},
        headers=auth_headers,
    )
    assert trainee_response.status_code == 201
    trainee_id = trainee_response.json()["id"]

    group_response = client.post(
        "/api/v1/groups",
        json={"code": "GRP-DEL-001", "name": "Група на видалення", "capacity": 20, "status": "active"},
        headers=auth_headers,
    )
    assert group_response.status_code == 201
    group_id = group_response.json()["id"]

    enroll_response = client.post(
        f"/api/v1/groups/{group_id}/enroll",
        json={"trainee_id": trainee_id},
        headers=auth_headers,
    )
    assert enroll_response.status_code == 201

    teacher = Teacher(branch_id="main", first_name="Тест", last_name="Викладач", hourly_rate=0.0, annual_load_hours=10.0)
    subject = Subject(branch_id="main", name="Тестовий предмет", hours_total=12)
    room = Room(branch_id="main", name="Аудиторія 999", capacity=20)
    db_session.add_all([teacher, subject, room])
    db_session.flush()

    slot_start = datetime.now(timezone.utc).replace(microsecond=0)
    db_session.add(
        ScheduleSlot(
            group_id=group_id,
            teacher_id=teacher.id,
            subject_id=subject.id,
            room_id=room.id,
            starts_at=slot_start,
            ends_at=slot_start + timedelta(hours=2),
            pair_number=1,
            academic_hours=2.0,
        )
    )
    db_session.add(
        Performance(
            branch_id="main",
            trainee_id=trainee_id,
            group_id=group_id,
            progress_pct=10.0,
            attendance_pct=90.0,
            employment_flag=False,
        )
    )
    db_session.commit()

    delete_response = client.delete(f"/api/v1/groups/{group_id}?delete_trainees=true", headers=auth_headers)
    assert delete_response.status_code == 204

    assert db_session.get(Group, group_id) is None
    assert db_session.query(GroupMembership).filter(GroupMembership.group_id == group_id).count() == 0
    assert db_session.query(ScheduleSlot).filter(ScheduleSlot.group_id == group_id).count() == 0
    assert db_session.query(Performance).filter(Performance.group_id == group_id).count() == 0


def test_delete_group_clears_trainee_group_code_when_trainees_kept(client, auth_headers, db_session):
    trainee_response = client.post(
        "/api/v1/trainees",
        json={"first_name": "Іван", "last_name": "Код", "status": "active", "group_code": "GRP-ORPH-001"},
        headers=auth_headers,
    )
    assert trainee_response.status_code == 201
    trainee_id = trainee_response.json()["id"]

    group_response = client.post(
        "/api/v1/groups",
        json={"code": "GRP-ORPH-001", "name": "Група для очищення коду", "capacity": 20, "status": "active"},
        headers=auth_headers,
    )
    assert group_response.status_code == 201
    group_id = group_response.json()["id"]

    delete_response = client.delete(f"/api/v1/groups/{group_id}", headers=auth_headers)
    assert delete_response.status_code == 204

    trainee = db_session.get(Trainee, trainee_id)
    assert trainee is not None
    assert trainee.is_deleted is False
    assert trainee.group_code is None


def test_delete_journal_backed_group_hides_it_and_keeps_monitor_match(client, auth_headers, db_session):
    group_response = client.post(
        "/api/v1/groups",
        json={"code": "GRP-JOURNAL-001", "name": "Група з журналом", "capacity": 20, "status": "active"},
        headers=auth_headers,
    )
    assert group_response.status_code == 201
    group_id = group_response.json()["id"]

    section = JournalMonitorSection(
        branch_id="main",
        name="Журнали",
        folder_url="https://drive.google.com/drive/folders/test",
        folder_id="test",
    )
    db_session.add(section)
    db_session.flush()
    entry = JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="journal-file-1",
        journal_name="Журнал GRP-JOURNAL-001",
        group_code="GRP-JOURNAL-001",
        matched_group_id=group_id,
        has_group=True,
    )
    db_session.add(entry)
    db_session.commit()

    delete_response = client.delete(f"/api/v1/groups/{group_id}", headers=auth_headers)
    assert delete_response.status_code == 204

    db_session.expire_all()
    group = db_session.get(Group, group_id)
    assert group is not None
    assert group.hidden_from_registry is True
    db_session.refresh(entry)
    assert entry.matched_group_id == group_id
    assert entry.has_group is True


def test_bulk_delete_groups_archives_matching_trainees(client, auth_headers, db_session):
    kept_group = Group(branch_id="main", code="KEEP-001", name="Залишається", status=GroupStatus.ACTIVE)
    db_session.add(kept_group)
    db_session.commit()

    first_group = client.post(
        "/api/v1/groups",
        json={"code": "BULK-DEL-001", "name": "Перша група", "capacity": 20, "status": "active"},
        headers=auth_headers,
    )
    second_group = client.post(
        "/api/v1/groups",
        json={"code": "BULK-DEL-002", "name": "Друга група", "capacity": 20, "status": "active"},
        headers=auth_headers,
    )
    assert first_group.status_code == 201
    assert second_group.status_code == 201
    group_ids = [first_group.json()["id"], second_group.json()["id"]]

    first_trainee = client.post(
        "/api/v1/trainees",
        json={"first_name": "Ірина", "last_name": "Перша", "status": "active", "group_code": "BULK-DEL-001"},
        headers=auth_headers,
    )
    second_trainee = client.post(
        "/api/v1/trainees",
        json={"first_name": "Олег", "last_name": "Другий", "status": "active", "group_code": "BULK-DEL-002"},
        headers=auth_headers,
    )
    kept_trainee = client.post(
        "/api/v1/trainees",
        json={"first_name": "Марія", "last_name": "Залишити", "status": "active", "group_code": "KEEP-001"},
        headers=auth_headers,
    )
    assert first_trainee.status_code == 201
    assert second_trainee.status_code == 201
    assert kept_trainee.status_code == 201

    response = client.post(
        "/api/v1/groups/bulk/delete",
        json={"group_ids": group_ids, "delete_trainees": True},
        headers=auth_headers,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["deleted_count"] == 2
    assert set(payload["deleted_ids"]) == set(group_ids)
    assert payload["missing_ids"] == []
    assert all(db_session.get(Group, group_id) is None for group_id in group_ids)
    assert db_session.get(Group, kept_group.id) is not None

    archived_rows = client.get("/api/v1/trainees?include_deleted=true", headers=auth_headers).json()
    archived_by_id = {item["id"]: item for item in archived_rows}
    assert archived_by_id[first_trainee.json()["id"]]["is_deleted"] is True
    assert archived_by_id[second_trainee.json()["id"]]["is_deleted"] is True
    assert archived_by_id[kept_trainee.json()["id"]]["is_deleted"] is False


def test_pages_endpoints_survive_group_delete_with_journal_workload(client, auth_headers, db_session):
    teacher = Teacher(branch_id="main", first_name="Іван", last_name="Викладач", hourly_rate=0.0, annual_load_hours=100.0)
    group = Group(branch_id="main", code="80-26", name="Група 80-26", status=GroupStatus.ACTIVE)
    subject = Subject(branch_id="main", name="Журнальний предмет", hours_total=12)
    room = Room(branch_id="main", name="Аудиторія 80", capacity=20)
    section = JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root",
        folder_id="root",
    )
    db_session.add_all([teacher, group, subject, room, section])
    db_session.flush()
    entry = JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="journal-80-26",
        journal_name="80-26 Група",
        group_code="80-26",
        matched_group_id=group.id,
        has_group=True,
        workload_status="processed",
        workload_year=2026,
        trainees_status="processed",
        trainees_message="Додано/оновлено слухачів із журналу: 1",
        trainee_count=1,
        has_trainees=True,
    )
    trainee = Trainee(branch_id="main", first_name="Петро", last_name="Слухач", status="active", group_code="80-26")
    db_session.add_all([entry, trainee])
    db_session.flush()
    db_session.add_all(
        [
            GroupMembership(group_id=group.id, trainee_id=trainee.id, status=MembershipStatus.ACTIVE),
            ScheduleSlot(
                group_id=group.id,
                teacher_id=teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=datetime(2026, 5, 12, 9, 30, tzinfo=timezone.utc),
                ends_at=datetime(2026, 5, 12, 11, 5, tzinfo=timezone.utc),
                pair_number=1,
                academic_hours=2,
            ),
            JournalWorkloadEntry(
                journal_monitor_entry_id=entry.id,
                branch_id="main",
                teacher_id=teacher.id,
                subject_name="Журнальний предмет",
                hours=2,
            ),
        ]
    )
    db_session.commit()

    delete_response = client.post(
        "/api/v1/groups/bulk/delete",
        json={"group_ids": [group.id], "delete_trainees": True},
        headers=auth_headers,
    )
    assert delete_response.status_code == 200

    for path in (
        "/api/v1/groups",
        "/api/v1/trainees?include_deleted=true",
        "/api/v1/trainees?search=80-26",
        "/api/v1/schedule",
        "/api/v1/teacher-workload",
    ):
        response = client.get(path, headers=auth_headers)
        assert response.status_code == 200, path


def test_clear_orphan_group_codes_endpoint(client, auth_headers):
    valid_group_response = client.post(
        "/api/v1/groups",
        json={"code": "GRP-VALID-001", "name": "Валідна група", "capacity": 20, "status": "active"},
        headers=auth_headers,
    )
    assert valid_group_response.status_code == 201

    orphan_trainee = client.post(
        "/api/v1/trainees",
        json={"first_name": "Олена", "last_name": "Сирота", "status": "active", "group_code": "NO-SUCH-GROUP"},
        headers=auth_headers,
    )
    valid_trainee = client.post(
        "/api/v1/trainees",
        json={"first_name": "Марія", "last_name": "Валідна", "status": "active", "group_code": "GRP-VALID-001"},
        headers=auth_headers,
    )
    assert orphan_trainee.status_code == 201
    assert valid_trainee.status_code == 201

    cleanup_response = client.post("/api/v1/trainees/bulk/clear-orphan-group-codes", headers=auth_headers)
    assert cleanup_response.status_code == 200
    payload = cleanup_response.json()
    assert payload["cleared_count"] == 1

    trainees_response = client.get("/api/v1/trainees", headers=auth_headers)
    assert trainees_response.status_code == 200
    rows = trainees_response.json()
    orphan_row = next(item for item in rows if item["id"] == orphan_trainee.json()["id"])
    valid_row = next(item for item in rows if item["id"] == valid_trainee.json()["id"])
    assert orphan_row["group_code"] is None
    assert valid_row["group_code"] == "GRP-VALID-001"


def test_archive_unassigned_group_trainees_endpoint(client, auth_headers):
    no_group_one = client.post(
        "/api/v1/trainees",
        json={"first_name": "А", "last_name": "БезГрупи1", "status": "active"},
        headers=auth_headers,
    )
    no_group_two = client.post(
        "/api/v1/trainees",
        json={"first_name": "Б", "last_name": "БезГрупи2", "status": "active", "group_code": ""},
        headers=auth_headers,
    )
    with_group = client.post(
        "/api/v1/trainees",
        json={"first_name": "В", "last_name": "ЗГрупою", "status": "active", "group_code": "73-26"},
        headers=auth_headers,
    )
    assert no_group_one.status_code == 201
    assert no_group_two.status_code == 201
    assert with_group.status_code == 201

    archive_response = client.post("/api/v1/trainees/bulk/archive-unassigned-group", headers=auth_headers)
    assert archive_response.status_code == 200
    payload = archive_response.json()
    assert payload["deleted_count"] == 2

    active_rows = client.get("/api/v1/trainees", headers=auth_headers)
    assert active_rows.status_code == 200
    active_ids = {item["id"] for item in active_rows.json()}
    assert no_group_one.json()["id"] not in active_ids
    assert no_group_two.json()["id"] not in active_ids
    assert with_group.json()["id"] in active_ids
