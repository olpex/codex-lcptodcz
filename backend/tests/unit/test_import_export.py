from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.core.crypto import cipher
from app.models import Group, GroupMembership, JournalMonitorEntry, JournalMonitorSection, JournalWorkloadEntry, Room, ScheduleSlot, Subject, Teacher, Trainee
from app.models import DocumentType
from app.services.import_export import (
    _ensure_group_for_trainee,
    analyze_trainee_import_duplicates,
    collect_report_rows,
    collect_group_export_rows,
    collect_teacher_workload_summary,
    parse_document_content,
    reconcile_teacher_workload_sources,
    save_report_file,
    try_import_trainees,
)
from app.services.trainee_deduplication import deduplicate_trainees


def test_parse_xlsx_and_import_trainees(tmp_path: Path, db_session):
    file_path = tmp_path / "trainees.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["first_name", "last_name", "status"])
    sheet.append(["Олена", "Коваль", "active"])
    sheet.append(["Іван", "Сидоренко", "active"])
    workbook.save(file_path)

    parsed = parse_document_content(str(file_path), doc_type=DocumentType.XLSX)
    assert parsed["rows"] == 2
    assert "first_name" in [h.lower() for h in parsed["headers"]]

    result = try_import_trainees(db_session, parsed, "main")
    assert result["inserted"] == 2


def test_teacher_workload_summary_includes_all_active_teachers_with_negative_remaining(db_session):
    no_plan_with_hours = Teacher(
        branch_id="main",
        first_name="Петро Іванович",
        last_name="Бойко",
        annual_load_hours=0,
        is_active=True,
    )
    no_plan_without_hours = Teacher(
        branch_id="main",
        first_name="Олена Петрівна",
        last_name="Андрук",
        annual_load_hours=0,
        is_active=True,
    )
    planned_without_hours = Teacher(
        branch_id="main",
        first_name="Ірина Миколаївна",
        last_name="Шевченко",
        annual_load_hours=12,
        is_active=True,
    )
    inactive_teacher = Teacher(
        branch_id="main",
        first_name="Ігор Петрович",
        last_name="Ярема",
        annual_load_hours=12,
        is_active=False,
    )
    group = Group(branch_id="main", code="WG-001", name="Група навантаження", status="active")
    subject = Subject(branch_id="main", name="Предмет навантаження", hours_total=20)
    room = Room(branch_id="main", name="Аудиторія навантаження", capacity=20)
    db_session.add_all([no_plan_with_hours, no_plan_without_hours, planned_without_hours, inactive_teacher, group, subject, room])
    db_session.flush()

    starts_at = datetime(2026, 4, 1, 9, 30, tzinfo=timezone.utc)
    db_session.add(
        ScheduleSlot(
            group_id=group.id,
            teacher_id=no_plan_with_hours.id,
            subject_id=subject.id,
            room_id=room.id,
            starts_at=starts_at,
            ends_at=starts_at + timedelta(minutes=95),
            academic_hours=2.0,
            pair_number=1,
        )
    )
    db_session.commit()

    rows = collect_teacher_workload_summary(db_session, "main")

    assert [row["teacher_name"] for row in rows] == [
        "Андрук Олена Петрівна",
        "Бойко Петро Іванович",
        "Шевченко Ірина Миколаївна",
    ]
    assert [row["row_number"] for row in rows] == [1, 2, 3]
    assert rows[0]["remaining_hours"] == 0
    assert rows[1]["remaining_hours"] == -2
    assert rows[2]["remaining_hours"] == 12


def test_teacher_workload_summary_includes_group_breakdown_from_schedule_and_journals(db_session):
    teacher = Teacher(branch_id="main", first_name="Олег Леонідович", last_name="Паращук", is_active=True)
    group = Group(branch_id="main", code="32-26", name="Група 32", status="active")
    subject = Subject(branch_id="main", name="Предмет деталізації", hours_total=20)
    room = Room(branch_id="main", name="Аудиторія деталізації", capacity=20)
    section = JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root",
        folder_id="root",
    )
    db_session.add_all([teacher, group, subject, room, section])
    db_session.flush()
    journal = JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="journal-33-26",
        journal_name="33-26 Журнал",
        group_code="33-26",
        workload_status="processed",
        workload_year=2026,
        workload_hours=10,
    )
    db_session.add(journal)
    db_session.flush()

    starts_at = datetime(2026, 2, 1, 9, 30, tzinfo=timezone.utc)
    db_session.add_all(
        [
            ScheduleSlot(
                group_id=group.id,
                teacher_id=teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=starts_at,
                ends_at=starts_at + timedelta(minutes=95),
                academic_hours=2.0,
                pair_number=1,
            ),
            JournalWorkloadEntry(
                journal_monitor_entry_id=journal.id,
                branch_id="main",
                teacher_id=teacher.id,
                subject_name="Журнальна дисципліна",
                hours=10,
            ),
        ]
    )
    db_session.commit()

    rows = collect_teacher_workload_summary(db_session, "main")

    assert rows[0]["total_hours"] == 12
    assert rows[0]["groups"] == [
        {"group_code": "32-26", "group_name": "Група 32", "hours": 2.0},
        {"group_code": "33-26", "group_name": "33-26 Журнал", "hours": 10.0},
    ]


def test_teacher_workload_summary_does_not_double_count_journal_hours_for_scheduled_group(db_session):
    teacher = Teacher(branch_id="main", first_name="Олег Леонідович", last_name="Паращук", is_active=True)
    group = Group(branch_id="main", code="46-26", name="Група 46-26", status="active")
    subject = Subject(branch_id="main", name="Предмет 46-26", hours_total=68)
    room = Room(branch_id="main", name="Аудиторія 46-26", capacity=20)
    section = JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root",
        folder_id="root",
    )
    db_session.add_all([teacher, group, subject, room, section])
    db_session.flush()
    journal = JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="journal-46-26",
        journal_name="46-26 Журнал",
        group_code="46-26",
        workload_status="processed",
        workload_year=2026,
        workload_hours=68,
    )
    db_session.add(journal)
    db_session.flush()

    starts_at = datetime(2026, 3, 1, 9, 30, tzinfo=timezone.utc)
    db_session.add_all(
        [
            ScheduleSlot(
                group_id=group.id,
                teacher_id=teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=starts_at,
                ends_at=starts_at + timedelta(minutes=95),
                academic_hours=68.0,
                pair_number=1,
            ),
            JournalWorkloadEntry(
                journal_monitor_entry_id=journal.id,
                branch_id="main",
                teacher_id=teacher.id,
                subject_name="Предмет 46-26",
                hours=68,
            ),
        ]
    )
    db_session.commit()

    rows = collect_teacher_workload_summary(db_session, "main")

    assert rows[0]["teacher_name"] == "Паращук Олег Леонідович"
    assert rows[0]["total_hours"] == 68
    assert rows[0]["groups"] == [
        {"group_code": "46-26", "group_name": "Група 46-26", "hours": 68.0},
    ]


def test_teacher_workload_export_uses_card_summary_group_breakdown(db_session):
    teacher = Teacher(
        branch_id="main",
        first_name="Галина Михайлівна",
        last_name="Войтихівська",
        annual_load_hours=180,
        is_active=True,
    )
    scheduled_group = Group(branch_id="main", code="16-26", name="Група 16-26", status="active")
    subject = Subject(branch_id="main", name="Предмет 16-26", hours_total=20)
    room = Room(branch_id="main", name="Аудиторія 16-26", capacity=20)
    section = JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root",
        folder_id="root",
    )
    db_session.add_all([teacher, scheduled_group, subject, room, section])
    db_session.flush()
    journal = JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="journal-17-26",
        journal_name="17-26 Журнал",
        group_code="17-26",
        workload_status="processed",
        workload_year=2026,
        workload_hours=10,
    )
    db_session.add(journal)
    db_session.flush()

    starts_at = datetime(2026, 2, 1, 9, 30, tzinfo=timezone.utc)
    db_session.add_all(
        [
            ScheduleSlot(
                group_id=scheduled_group.id,
                teacher_id=teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=starts_at,
                ends_at=starts_at + timedelta(minutes=95),
                academic_hours=2.0,
                pair_number=1,
            ),
            JournalWorkloadEntry(
                journal_monitor_entry_id=journal.id,
                branch_id="main",
                teacher_id=teacher.id,
                subject_name="Журнальна дисципліна",
                hours=10,
            ),
        ]
    )
    db_session.commit()

    sheets = collect_report_rows(
        db_session,
        "teacher_workload",
        "main",
        {"teacher_ids": [teacher.id], "start_date": None, "end_date": None},
    )

    teacher_rows = sheets["Войтихівська Галина Михайлівна"]
    assert teacher_rows == [
        {
            "Номер за порядком": 1,
            "Прізвище, ім'я та по батькові викладача": "Войтихівська Галина Михайлівна",
            "Групи": "16-26",
            "Назва групи": "Група 16-26",
            "Годин у групі": 2.0,
            "Загальна кількість годин": 12.0,
            "Річне педнавантаження": 180.0,
            "Залишок годин": 168.0,
        },
        {
            "Номер за порядком": 1,
            "Прізвище, ім'я та по батькові викладача": "Войтихівська Галина Михайлівна",
            "Групи": "17-26",
            "Назва групи": "17-26 Журнал",
            "Годин у групі": 10.0,
            "Загальна кількість годин": 12.0,
            "Річне педнавантаження": 180.0,
            "Залишок годин": 168.0,
        },
    ]


def test_teacher_workload_xlsx_sheet_moves_teacher_totals_to_top(db_session):
    teacher = Teacher(
        branch_id="main",
        first_name="First Middle",
        last_name="Teacher",
        annual_load_hours=180,
        is_active=True,
    )
    first_group = Group(branch_id="main", code="16-26", name="Group 16", status="active")
    second_group = Group(branch_id="main", code="17-26", name="Group 17", status="active")
    subject = Subject(branch_id="main", name="Subject", hours_total=20)
    room = Room(branch_id="main", name="Room", capacity=20)
    db_session.add_all([teacher, first_group, second_group, subject, room])
    db_session.flush()

    starts_at = datetime(2026, 2, 1, 9, 30, tzinfo=timezone.utc)
    db_session.add_all(
        [
            ScheduleSlot(
                group_id=first_group.id,
                teacher_id=teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=starts_at,
                ends_at=starts_at + timedelta(minutes=95),
                academic_hours=2.0,
                pair_number=1,
            ),
            ScheduleSlot(
                group_id=second_group.id,
                teacher_id=teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=starts_at + timedelta(days=1),
                ends_at=starts_at + timedelta(days=1, minutes=95),
                academic_hours=6.0,
                pair_number=2,
            ),
        ]
    )
    db_session.commit()

    sheets = collect_report_rows(
        db_session,
        "teacher_workload",
        "main",
        {"teacher_ids": [teacher.id], "start_date": None, "end_date": None},
    )
    teacher_rows = next(iter(sheets.values()))
    headers = list(teacher_rows[0].keys())

    xlsx_path, doc_type = save_report_file(
        sheets,
        "teacher_workload",
        "xlsx",
        {"teacher_ids": [teacher.id], "start_date": None, "end_date": None},
    )

    assert doc_type == DocumentType.XLSX
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    assert [sheet.cell(1, column).value for column in range(2, 6)] == [
        headers[1],
        headers[5],
        headers[6],
        headers[7],
    ]
    assert [sheet.cell(2, column).value for column in range(2, 6)] == [
        teacher_rows[0][headers[1]],
        8.0,
        180.0,
        172.0,
    ]
    assert [sheet.cell(4, column).value for column in range(1, 4)] == headers[2:5]
    assert [sheet.cell(5, column).value for column in range(1, 4)] == ["16-26", "Group 16", 2.0]
    assert [sheet.cell(6, column).value for column in range(1, 4)] == ["17-26", "Group 17", 6.0]
    assert sheet["A1"].value is None
    assert sheet["A4"].value != headers[0]


def test_reconcile_teacher_workload_sources_removes_stale_rows_and_corrects_entry_hours(db_session):
    teacher = Teacher(branch_id="main", first_name="Олена Петрівна", last_name="Коваль", annual_load_hours=100, is_active=True)
    section = JournalMonitorSection(
        branch_id="main",
        name="Журнали 2026",
        folder_url="https://drive.google.com/drive/folders/root",
        folder_id="root",
    )
    db_session.add_all([teacher, section])
    db_session.flush()
    processed = JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="journal-10-26",
        journal_name="10-26 Журнал",
        group_code="10-26",
        workload_status="processed",
        workload_year=2026,
        workload_hours=99,
    )
    pending = JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="journal-11-26",
        journal_name="11-26 Журнал",
        group_code="11-26",
        workload_status="pending",
        workload_year=2026,
        workload_hours=12,
    )
    no_data = JournalMonitorEntry(
        section_id=section.id,
        branch_id="main",
        drive_file_id="journal-12-26",
        journal_name="12-26 Журнал",
        group_code="12-26",
        workload_status="no_data",
        workload_year=2026,
        workload_hours=30,
        workload_source_names=["12-26.xlsx"],
    )
    db_session.add_all([processed, pending, no_data])
    db_session.flush()
    db_session.add_all(
        [
            JournalWorkloadEntry(
                journal_monitor_entry_id=processed.id,
                branch_id="main",
                teacher_id=teacher.id,
                subject_name="Чинна дисципліна",
                hours=14,
            ),
            JournalWorkloadEntry(
                journal_monitor_entry_id=pending.id,
                branch_id="main",
                teacher_id=teacher.id,
                subject_name="Застаріла дисципліна",
                hours=12,
            ),
        ]
    )
    db_session.commit()

    report = reconcile_teacher_workload_sources(db_session, "main")
    db_session.commit()

    assert report == {
        "deleted_stale_workload_rows": 1,
        "corrected_processed_entries": 1,
        "reset_unprocessed_entries": 1,
    }
    db_session.refresh(processed)
    db_session.refresh(pending)
    db_session.refresh(no_data)
    assert processed.workload_hours == 14
    assert pending.workload_hours == 0
    assert no_data.workload_hours == 30
    assert no_data.workload_source_names == ["12-26.xlsx"]
    rows = collect_teacher_workload_summary(db_session, "main")
    assert rows[0]["total_hours"] == 14
    assert rows[0]["remaining_hours"] == 86


def test_group_export_rows_include_existing_groups_and_teacher_hours(db_session):
    scheduled_group = Group(branch_id="main", code="72-26", name="Група з розкладом", status="active")
    empty_group = Group(branch_id="main", code="73-26", name="Група без розкладу", status="active")
    first_teacher = Teacher(branch_id="main", first_name="Ірина Петрівна", last_name="Коваль", is_active=True)
    second_teacher = Teacher(branch_id="main", first_name="Марія Іванівна", last_name="Бондар", is_active=True)
    subject = Subject(branch_id="main", name="Предмет груп", hours_total=20)
    room = Room(branch_id="main", name="Аудиторія груп", capacity=20)
    db_session.add_all([scheduled_group, empty_group, first_teacher, second_teacher, subject, room])
    db_session.flush()

    starts_at = datetime(2026, 4, 1, 9, 30, tzinfo=timezone.utc)
    db_session.add_all(
        [
            ScheduleSlot(
                group_id=scheduled_group.id,
                teacher_id=first_teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=starts_at,
                ends_at=starts_at + timedelta(minutes=95),
                academic_hours=2.0,
                pair_number=1,
            ),
            ScheduleSlot(
                group_id=scheduled_group.id,
                teacher_id=second_teacher.id,
                subject_id=subject.id,
                room_id=room.id,
                starts_at=starts_at + timedelta(days=1),
                ends_at=starts_at + timedelta(days=1, minutes=95),
                academic_hours=3.0,
                pair_number=2,
            ),
        ]
    )
    db_session.commit()

    rows = collect_group_export_rows(db_session, "main")

    assert rows == [
        {
            "Номер групи": "72-26",
            "Назва групи": "Група з розкладом",
            "Кількість годин": 5,
            "Викладач": "Бондар Марія Іванівна",
            "Кількість годин викладача в групі": 3,
        },
        {
            "Номер групи": "72-26",
            "Назва групи": "Група з розкладом",
            "Кількість годин": 5,
            "Викладач": "Коваль Ірина Петрівна",
            "Кількість годин викладача в групі": 2,
        },
        {
            "Номер групи": "73-26",
            "Назва групи": "Група без розкладу",
            "Кількість годин": 0,
            "Викладач": "",
            "Кількість годин викладача в групі": 0,
        },
    ]


def _create_contract_like_workbook(file_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Додаток"
    sheet.append(["", "", "Список безробітних", "", "", "", "", "", "", "", "", "", "", "", ""])
    sheet.append(["", "", "Група 73-26  «Штучний інтелект»", "", "", "", "", "", "", "", "", "", "", "", ""])
    sheet.append(
        [
            "№",
            "Центр зайнятості, який направив безробітного  на професійне навчання",
            "ПІБ безробітного ",
            "Дата народження",
            "№ Договору",
            "Сертифікат",
            "Дата видачі сертифікату",
            "Індекс",
            "Адреса",
            "Паспорт: СЕРІЯ",
            "Паспорт: №",
            "Ким виданий",
            "Коли виданий",
            "Ідентифікаційний код",
            "Телефон",
        ]
    )
    sheet.append(
        [
            1,
            "Луцька філія",
            "Бортнік Тетяна Анатоліївна",
            "12.04.1984",
            "1499",
            "03032604210004401Н",
            "27.04.2026",
            "44500",
            "м. Луцьк, вул. Шевченка 1",
            "АЮ",
            "276684",
            "Луцьким РВ УДМС",
            "17.12.2015",
            "3079204140",
            "0978319450",
        ]
    )
    workbook.save(file_path)


def test_import_uses_dodatok_sheet_group_context_and_populates_fields(tmp_path: Path, db_session):
    file_path = tmp_path / "contracts.xlsx"
    _create_contract_like_workbook(file_path)

    parsed = parse_document_content(str(file_path), doc_type=DocumentType.XLSX)
    assert parsed["sheet_name"] == "Додаток"
    assert parsed["default_group_code"] == "73-26"

    result = try_import_trainees(db_session, parsed, "main")
    assert result["inserted"] == 1
    assert result["default_group_code"] == "73-26"

    trainee = db_session.query(Trainee).filter(Trainee.contract_number == "1499").first()
    assert trainee is not None
    assert trainee.group_code == "73-26"
    assert trainee.birth_date is not None
    assert trainee.employment_center_encrypted is not None
    assert trainee.address_encrypted is not None
    assert trainee.passport_series_encrypted is not None
    assert trainee.phone_encrypted is not None


def test_import_reads_group_context_with_colon_and_typographic_dash(tmp_path: Path):
    file_path = tmp_path / "contracts_typographic_group.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Додаток"
    sheet.append(["Група: 80–26 Організація трудових відносин"])
    sheet.append([])
    sheet.append(["№", "ПІБ безробітного", "Дата народження", "№ Договору"])
    sheet.append([1, "Кравченко Олена Іванівна", "05.03.1991", "80-26/001"])
    workbook.save(file_path)

    parsed = parse_document_content(str(file_path), doc_type=DocumentType.XLSX)

    assert parsed["default_group_code"] == "80-26"
    assert parsed["default_group_name"] == "Організація трудових відносин"


def test_import_prefers_dodatok_sheet_over_other_registry_like_sheets(tmp_path: Path, db_session):
    file_path = tmp_path / "listeners.xlsx"
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Договори"
    first_sheet.append(["Група 99-26 Архів"])
    first_sheet.append([])
    first_sheet.append(["№", "ПІБ безробітного", "Дата народження", "№ Договору"])
    first_sheet.append([1, "Архівний Петро Іванович", "01.01.1990", "99-26/001"])

    dodatok = workbook.create_sheet("Додаток")
    dodatok.append(["Група 73-26 Штучний інтелект"])
    dodatok.append([])
    dodatok.append(["№", "ПІБ безробітного", "Дата народження", "№ Договору"])
    dodatok.append([1, "Бортнік Тетяна Анатоліївна", "12.04.1984", "1499"])
    workbook.save(file_path)

    parsed = parse_document_content(str(file_path), doc_type=DocumentType.XLSX)

    assert parsed["sheet_name"] == "Додаток"
    assert parsed["default_group_code"] == "73-26"
    assert parsed["data"][0]["№ Договору"] == "1499"


def test_import_falls_back_to_first_registry_sheet_when_dodatok_is_absent(tmp_path: Path, db_session):
    file_path = tmp_path / "contracts_first_sheet.xlsx"
    _create_contract_like_workbook(file_path)

    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Перший аркуш"
    for row in load_workbook(file_path).active.iter_rows(values_only=True):
        first_sheet.append(list(row))
    service_sheet = workbook.create_sheet("Службовий аркуш")
    service_sheet.append(["Службовий аркуш без реєстру слухачів"])
    workbook.save(file_path)

    parsed = parse_document_content(str(file_path), doc_type=DocumentType.XLSX)
    assert parsed["sheet_name"] == "Перший аркуш"

    result = try_import_trainees(db_session, parsed, "main")
    assert result["inserted"] == 1
    trainee = db_session.query(Trainee).filter(Trainee.contract_number == "1499").first()
    assert trainee is not None


def test_import_updates_existing_missing_fields_instead_of_skipping(tmp_path: Path, db_session):
    db_session.add(
        Trainee(
            branch_id="main",
            first_name="Тетяна Анатоліївна",
            last_name="Бортнік",
            status="active",
        )
    )
    db_session.commit()

    file_path = tmp_path / "contracts_update.xlsx"
    _create_contract_like_workbook(file_path)
    parsed = parse_document_content(str(file_path), doc_type=DocumentType.XLSX)

    result = try_import_trainees(db_session, parsed, "main")
    assert result["inserted"] == 0
    assert result["updated_existing"] == 1

    trainees = db_session.query(Trainee).filter(Trainee.last_name == "Бортнік").all()
    assert len(trainees) == 1
    trainee = trainees[0]
    assert trainee.contract_number == "1499"
    assert trainee.group_code == "73-26"
    assert trainee.employment_center_encrypted is not None


def test_import_replaces_existing_phone_that_is_actually_address(db_session):
    db_session.add(
        Trainee(
            branch_id="main",
            first_name="Іван Іванович",
            last_name="Петренко",
            birth_date=datetime(1990, 2, 1).date(),
            status="active",
            address_encrypted=cipher.encrypt("м. Львів, вул. Зелена 1"),
            phone_encrypted=cipher.encrypt("м. Львів, вул. Зелена 1"),
        )
    )
    db_session.commit()

    result = try_import_trainees(
        db_session,
        {
            "headers": ["Прізвище", "Ім'я", "По батькові", "Дата народження", "Домашня адреса", "Телефон"],
            "data": [
                {
                    "Прізвище": "Петренко",
                    "Ім'я": "Іван",
                    "По батькові": "Іванович",
                    "Дата народження": "01.02.1990",
                    "Домашня адреса": "м. Львів, вул. Зелена 1",
                    "Телефон": "+380501112233",
                }
            ],
        },
        "main",
    )

    assert result["updated_existing"] == 1
    trainee = db_session.query(Trainee).filter(Trainee.last_name == "Петренко").one()
    assert cipher.decrypt(trainee.phone_encrypted) == "+380501112233"


def test_import_restores_archived_existing_trainee(tmp_path: Path, db_session):
    db_session.add(
        Trainee(
            branch_id="main",
            first_name="Тетяна Анатоліївна",
            last_name="Бортнік",
            contract_number="1499",
            status="active",
            is_deleted=True,
            deleted_at=datetime.now(timezone.utc),
        )
    )
    db_session.commit()

    file_path = tmp_path / "contracts_restore.xlsx"
    _create_contract_like_workbook(file_path)
    parsed = parse_document_content(str(file_path), doc_type=DocumentType.XLSX)

    result = try_import_trainees(db_session, parsed, "main")
    assert result["inserted"] == 0
    assert result["updated_existing"] == 1
    assert result["restored_deleted"] == 1

    trainee = db_session.query(Trainee).filter(Trainee.contract_number == "1499").one()
    assert trainee.is_deleted is False
    assert trainee.deleted_at is None
    assert trainee.group_code == "73-26"


def test_import_overwrite_mode_updates_existing_non_empty_fields(tmp_path: Path, db_session):
    db_session.add(
        Trainee(
            branch_id="main",
            first_name="Тетяна Анатоліївна",
            last_name="Бортнік",
            contract_number="OLD-1499",
            status="completed",
            group_code="OLD-GROUP",
        )
    )
    db_session.commit()

    file_path = tmp_path / "contracts_overwrite.xlsx"
    _create_contract_like_workbook(file_path)
    parsed = parse_document_content(str(file_path), doc_type=DocumentType.XLSX)

    result = try_import_trainees(db_session, parsed, "main", update_existing_mode="overwrite")
    assert result["inserted"] == 0
    assert result["updated_existing"] == 1
    assert result["update_existing_mode"] == "overwrite"

    trainee = db_session.query(Trainee).filter(Trainee.last_name == "Бортнік").first()
    assert trainee is not None
    assert trainee.contract_number == "1499"
    assert trainee.group_code == "73-26"
    assert trainee.status == "active"


def test_import_matches_existing_full_first_name_when_incoming_omits_middle_name(db_session):
    db_session.add(
        Trainee(
            branch_id="main",
            source_row_number=5,
            first_name="РћР»СЊРіР° Р®СЂС–С—РІРЅР°",
            last_name="Р’СЂСѓР±Р»РµРІСЃСЊРєР°",
            birth_date=datetime(1988, 5, 4).date(),
            group_code="90-26",
            tax_id_encrypted=cipher.encrypt("3226619663"),
            address_encrypted=cipher.encrypt("Р’РѕР»РёРЅСЊСЃРєР° РѕР±Р»Р°СЃС‚СЊ"),
            status="active",
        )
    )
    db_session.commit()
    parsed = {
        "rows": 1,
        "headers": [
            "source_row_number",
            "last_name",
            "first_name",
            "birth_date",
            "contract_number",
            "tax_id",
            "passport_number",
        ],
        "data": [
            {
                "source_row_number": 5,
                "last_name": "Р’СЂСѓР±Р»РµРІСЊРєР°",
                "first_name": "РћР»СЊРіР°",
                "birth_date": "04.05.1988",
                "contract_number": "1826",
                "tax_id": "3226619663",
                "passport_number": "937755",
            }
        ],
        "default_group_code": "90-26",
    }

    result = try_import_trainees(db_session, parsed, "main", update_existing_mode="overwrite")

    assert result["inserted"] == 0
    assert result["updated_existing"] == 1
    trainee = db_session.query(Trainee).filter(Trainee.last_name == "Р’СЂСѓР±Р»РµРІСЃСЊРєР°").one()
    assert trainee.contract_number == "1826"
    assert cipher.decrypt(trainee.passport_number_encrypted) == "937755"


def test_deduplicate_trainees_keeps_most_complete_group_record_and_removes_sparse_copy(db_session):
    sparse = Trainee(
        branch_id="main",
        source_row_number=5,
        first_name="РћР»СЊРіР° Р®СЂС–С—РІРЅР°",
        last_name="Р’СЂСѓР±Р»РµРІСЃСЊРєР°",
        birth_date=datetime(1988, 5, 4).date(),
        group_code="90-26",
        address_encrypted=cipher.encrypt("Р’РѕР»РёРЅСЊСЃРєР° РѕР±Р»Р°СЃС‚СЊ"),
        tax_id_encrypted=cipher.encrypt("3226619663"),
        status="active",
    )
    complete = Trainee(
        branch_id="main",
        source_row_number=5,
        first_name="РћР»СЊРіР° Р®СЂС–С—РІРЅР°",
        last_name="Р’СЂСѓР±Р»РµРІСЃСЊРєР°",
        birth_date=datetime(1988, 5, 4).date(),
        group_code="90-26",
        employment_center_encrypted=cipher.encrypt("Р›СѓС†СЊРєР° С„С–Р»С–СЏ Р’РћР¦Р—"),
        contract_number="1826",
        certificate_number="03502605140025601Рќ",
        certificate_issue_date=datetime(2026, 5, 14).date(),
        postal_index="43001",
        address_encrypted=cipher.encrypt("Р’РѕР»РёРЅСЊСЃРєР° РѕР±Р»Р°СЃС‚СЊ"),
        passport_series_encrypted=cipher.encrypt("РђРЎ"),
        passport_number_encrypted=cipher.encrypt("937755"),
        passport_issued_by_encrypted=cipher.encrypt("Р›СѓС†СЊРєРёРј Р Р’ РЈРњР’РЎ"),
        passport_issued_date=datetime(2004, 10, 18).date(),
        tax_id_encrypted=cipher.encrypt("3226619663"),
        status="active",
    )
    db_session.add_all([sparse, complete])
    db_session.commit()

    result = deduplicate_trainees(db_session, "main", commit=True)

    assert result["duplicate_groups"] == 1
    assert result["removed_count"] == 1
    remaining = db_session.query(Trainee).filter(Trainee.is_deleted.is_(False)).one()
    assert remaining.id == complete.id
    assert remaining.contract_number == "1826"
    assert cipher.decrypt(remaining.passport_number_encrypted) == "937755"


def test_deduplicate_trainees_keeps_memberships_attached_to_keeper(db_session):
    keeper = Trainee(
        branch_id="main",
        first_name="Ольга Юріївна",
        last_name="Врублевська",
        birth_date=datetime(1988, 5, 4).date(),
        group_code="90-26",
        tax_id_encrypted=cipher.encrypt("3226619663"),
        status="active",
    )
    duplicate = Trainee(
        branch_id="main",
        first_name="Ольга Юріївна",
        last_name="Врублевська",
        birth_date=datetime(1988, 5, 4).date(),
        group_code="90-26",
        tax_id_encrypted=cipher.encrypt("3226619663"),
        contract_number="1826",
        status="active",
    )
    group = Group(branch_id="main", code="90-26", name="Журнал 90-26", status="active")
    db_session.add_all([keeper, duplicate, group])
    db_session.flush()
    db_session.add(
        GroupMembership(
            group_id=group.id,
            trainee_id=duplicate.id,
            status="active",
        )
    )
    db_session.commit()

    result = deduplicate_trainees(db_session, "main", commit=True)

    assert result["duplicate_groups"] == 1
    assert result["removed_count"] == 1
    remaining = db_session.query(Trainee).filter(Trainee.is_deleted.is_(False)).one()
    memberships = db_session.query(GroupMembership).all()
    assert len(memberships) == 1
    assert memberships[0].trainee_id == remaining.id
    assert remaining.contract_number == "1826"


def test_import_overwrite_reuses_existing_group_membership(db_session):
    group = Group(branch_id="main", code="90-26", name="Журнал 90-26", status="active")
    trainee = Trainee(
        branch_id="main",
        first_name="Ольга Юріївна",
        last_name="Врублевська",
        birth_date=datetime(1988, 5, 4).date(),
        group_code="90-26",
        tax_id_encrypted=cipher.encrypt("3226619663"),
        status="active",
    )
    db_session.add_all([group, trainee])
    db_session.flush()
    db_session.add(
        GroupMembership(
            group_id=group.id,
            trainee_id=trainee.id,
            status="active",
        )
    )
    db_session.commit()

    parsed = {
        "rows": 1,
        "headers": ["Прізвище", "Ім'я", "По батькові", "Дата народження"],
        "sheet_name": "ЗВ",
        "default_group_code": "90-26",
        "default_group_name": "Журнал 90-26",
        "data": [
            {
                "Прізвище": "Врублевська",
                "Ім'я": "Ольга",
                "По батькові": "Юріївна",
                "Дата народження": "04.05.1988",
            }
        ],
    }

    result = try_import_trainees(db_session, parsed, "main", update_existing_mode="overwrite")

    assert result["inserted"] == 0
    assert result["updated_existing"] == 0
    assert result["skipped_existing"] == 1
    memberships = db_session.query(GroupMembership).filter(GroupMembership.group_id == group.id).all()
    assert len(memberships) == 1
    assert memberships[0].trainee_id == trainee.id


def test_ensure_group_for_trainee_ignores_conflicting_membership_insert(db_session, monkeypatch):
    group = Group(branch_id="main", code="90-26", name="Журнал 90-26", status="active")
    trainee = Trainee(
        branch_id="main",
        first_name="Ольга Юріївна",
        last_name="Врублевська",
        birth_date=datetime(1988, 5, 4).date(),
        group_code="90-26",
        tax_id_encrypted=cipher.encrypt("3226619663"),
        status="active",
    )
    db_session.add_all([group, trainee])
    db_session.flush()
    db_session.add(
        GroupMembership(
            group_id=group.id,
            trainee_id=trainee.id,
            status="active",
        )
    )
    db_session.commit()

    real_query = db_session.query

    class _FakeMembershipQuery:
        def filter(self, *_args, **_kwargs):
            return self

        def first(self):
            return None

    def fake_query(*entities, **kwargs):
        if len(entities) == 1 and entities[0] is GroupMembership:
            return _FakeMembershipQuery()
        return real_query(*entities, **kwargs)

    monkeypatch.setattr(db_session, "query", fake_query)

    memberships_added, group_changed = _ensure_group_for_trainee(
        db_session,
        trainee,
        "main",
        {"90-26": group},
        set(),
        "90-26",
        "Журнал 90-26",
        overwrite_group=True,
    )

    assert memberships_added == 0
    assert group_changed is False
    memberships = real_query(GroupMembership).filter(GroupMembership.group_id == group.id).all()
    assert len(memberships) == 1
    assert memberships[0].trainee_id == trainee.id


def test_import_skip_existing_mode_does_not_update_duplicate(tmp_path: Path, db_session):
    db_session.add(
        Trainee(
            branch_id="main",
            first_name="Тетяна Анатоліївна",
            last_name="Бортнік",
            contract_number="1499",
            status="completed",
            group_code="OLD-GROUP",
        )
    )
    db_session.commit()

    file_path = tmp_path / "contracts_skip_existing.xlsx"
    _create_contract_like_workbook(file_path)
    parsed = parse_document_content(str(file_path), doc_type=DocumentType.XLSX)

    result = try_import_trainees(db_session, parsed, "main", update_existing_mode="skip_existing")
    assert result["inserted"] == 0
    assert result["updated_existing"] == 0
    assert result["skipped_existing"] == 1
    assert result["update_existing_mode"] == "skip_existing"

    trainee = db_session.query(Trainee).filter(Trainee.contract_number == "1499").first()
    assert trainee is not None
    assert trainee.group_code == "OLD-GROUP"
    assert trainee.status == "completed"


def test_import_deduplicates_repeated_rows_for_same_group_membership(db_session):
    parsed = {
        "rows": 3,
        "headers": ["Прізвище", "Ім'я", "По батькові"],
        "sheet_name": "ЗВ",
        "default_group_code": "1-26",
        "default_group_name": "Журнал 1-26",
        "data": [
            {"Прізвище": "Петренко", "Ім'я": "Іван", "По батькові": "Іванович"},
            {"Прізвище": "Петренко", "Ім'я": "Іван", "По батькові": "Іванович"},
            {"Прізвище": "Петренко", "Ім'я": "Іван", "По батькові": "Іванович"},
        ],
    }

    result = try_import_trainees(db_session, parsed, "main")

    assert result["inserted"] == 1
    assert result["updated_existing"] == 0
    assert result["skipped_existing"] == 2
    assert result["memberships_created"] == 1
    assert db_session.query(Trainee).count() == 1
    assert db_session.query(GroupMembership).count() == 1


def test_analyze_trainee_import_duplicates_reports_existing_rows(tmp_path: Path, db_session):
    db_session.add(
        Trainee(
            branch_id="main",
            first_name="Тетяна Анатоліївна",
            last_name="Бортнік",
            contract_number="1499",
            status="active",
        )
    )
    db_session.commit()

    file_path = tmp_path / "contracts_duplicate_preview.xlsx"
    _create_contract_like_workbook(file_path)
    parsed = parse_document_content(str(file_path), doc_type=DocumentType.XLSX)

    result = analyze_trainee_import_duplicates(db_session, parsed, "main")
    assert result["duplicate_count"] == 1
    assert result["new_count"] == 0
    assert result["invalid_count"] == 0
    assert result["duplicate_preview"][0]["incoming_name"] == "Бортнік Тетяна Анатоліївна"
    assert result["duplicate_preview"][0]["match_reason"] == "contract_number"
